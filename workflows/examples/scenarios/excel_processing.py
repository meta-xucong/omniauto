"""场景6: 办公自动化 - 创建Excel，写入数据，读取并计算汇总."""

import asyncio
from pathlib import Path
from datetime import datetime
from omniauto.core.context import TaskContext, StepResult
from omniauto.core.state_machine import AtomicStep, Workflow
from openpyxl import Workbook, load_workbook

async def excel_office_auto(ctx: TaskContext) -> StepResult:
    output_dir = Path("./outputs")
    output_dir.mkdir(exist_ok=True)
    excel_path = output_dir / f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 1. 创建并写入数据
    wb = Workbook()
    ws = wb.active
    ws.title = "销售数据"
    ws.append(["产品", "销量", "单价"])
    ws.append(["A产品", 100, 50])
    ws.append(["B产品", 200, 30])
    ws.append(["C产品", 150, 40])
    wb.save(str(excel_path))

    # 2. 读取并计算汇总
    wb2 = load_workbook(str(excel_path))
    ws2 = wb2.active
    total_sales = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        qty = row[1] or 0
        price = row[2] or 0
        total_sales += qty * price

    # 3. 追加汇总行
    ws2.append(["汇总", "", total_sales])
    wb2.save(str(excel_path))

    return StepResult(success=True, data={
        "excel_path": str(excel_path),
        "total_sales": total_sales,
        "record_count": ws2.max_row - 1,
    })

workflow = Workflow(task_id="scenario_excel_processing")
workflow.add_step(AtomicStep("excel_auto", excel_office_auto, lambda r: r.success))
