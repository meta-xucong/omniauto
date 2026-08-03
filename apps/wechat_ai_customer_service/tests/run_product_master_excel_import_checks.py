"""Focused checks for local/manual V2 vehicle Excel import."""

from __future__ import annotations

import ast
import json
import sys
import tempfile
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook, load_workbook
from fastapi import FastAPI
from fastapi.testclient import TestClient


APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
for import_root in (APP_ROOT / "workflows", APP_ROOT / "adapters", APP_ROOT):
    if str(import_root) not in sys.path:
        sys.path.insert(0, str(import_root))

from apps.wechat_ai_customer_service.admin_backend.services.product_master_excel_import import (  # noqa: E402
    FIELD_SHEET,
    LEGACY_TEMPLATE_VERSION,
    LEGACY_VEHICLE_COLUMNS,
    MAX_IMPORT_ROWS,
    MAX_PICTURE_ROWS,
    PATH_VERTICAL_TEMPLATE_VERSION,
    PATH_VERTICAL_VEHICLE_HEADERS,
    PICTURE_SHEET,
    PICTURE_COLUMNS,
    RAW_VEHICLE_TEMPLATE_COLUMNS,
    SIMPLE_PICTURE_COLUMNS,
    SIMPLE_VERTICAL_TEMPLATE_VERSION,
    TEMPLATE_VERSION,
    VEHICLE_COLUMNS,
    VEHICLE_SHEET,
    VERTICAL_VEHICLE_HEADERS,
    ProductMasterExcelImportService,
    build_template_bytes,
    _simple_field_label,
)
from apps.wechat_ai_customer_service.admin_backend.api import product_console as product_console_api  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.auth_context import AuthTenantMiddleware  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.product_console_service import ProductConsoleService  # noqa: E402
from apps.wechat_ai_customer_service import product_master as product_master_module  # noqa: E402
from apps.wechat_ai_customer_service.product_master import ProductMasterStore  # noqa: E402
from apps.wechat_ai_customer_service.workflows.customer_service_brain import compact_product_item_for_brain_prompt  # noqa: E402
from packages.dafengche_product_master import create_manual_vehicle  # noqa: E402


def main() -> int:
    checks = [
        check_template_generates_current_version_and_standard_dafengche_name_shape,
        check_template_mirrors_frontend_v2_admin_fields_and_documents_image_url_boundary,
        check_template_layout_styles_preserve_parser_contract,
        check_preview_does_not_write_and_confirm_writes_manual_v2,
        check_complete_admin_fields_round_trip_and_customer_evidence_filter,
        check_missing_template_version_rejected_without_preview_artifact,
        check_legacy_template_version_remains_readable,
        check_horizontal_admin_template_version_remains_readable,
        check_path_vertical_template_version_remains_readable,
        check_old_simple_vertical_local_extension_fields_remain_readable,
        check_simple_vertical_url_picture_sheet_remains_readable,
        check_unknown_column_and_missing_required_field_are_row_errors,
        check_duplicate_header_is_rejected,
        check_vehicle_and_picture_row_limits_are_errors,
        check_duplicate_field_item_and_picture_row_errors,
        check_repeat_import_preserves_existing_manual_fields,
        check_confirm_rollback_prevents_half_batch_pollution,
        check_confirm_uses_postgres_batch_transaction,
        check_postgres_batch_failure_writes_zero_items,
        check_postgres_mirror_failure_warns_and_db_remains_canonical,
        check_product_master_batch_rejects_duplicate_item_ids,
        check_excel_manual_customer_evidence_filters_restricted_raw_payload_fields,
        check_excel_manual_customer_evidence_tenant_and_shop_scope_isolated,
        check_product_console_excel_api_contract,
        check_preview_is_tenant_scoped,
        check_import_service_has_no_network_llm_or_dafengche_client_imports,
        check_customer_evidence_does_not_leak_raw_source_payloads,
        check_unauthenticated_template_download_only_allows_template,
    ]
    failures: list[str] = []
    for check in checks:
        try:
            check()
            print(f"PASS {check.__name__}")
        except Exception as exc:  # pragma: no cover - script runner
            failures.append(f"{check.__name__}: {exc!r}")
            print(f"FAIL {check.__name__}: {exc!r}")
    if failures:
        raise SystemExit("\n".join(failures))
    print(f"PASS {len(checks)}/{len(checks)} product-master Excel import checks")
    return 0


def check_template_generates_current_version_and_standard_dafengche_name_shape() -> None:
    workbook = load_workbook(BytesIO(build_template_bytes()), data_only=True)
    vehicle = workbook[VEHICLE_SHEET]
    assert_true(PICTURE_SHEET not in workbook.sheetnames, "new download template must not expose an Excel image URL sheet")
    for row in range(2, vehicle.max_row + 1):
        assert_equal(vehicle.cell(row=row, column=1).value, None, "download template vehicle id cells should be blank")
        assert_equal(vehicle.cell(row=row, column=3).value, None, "download template editable value cells should be blank")
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_template_check", root=Path(directory) / "pm")
        parsed = ProductMasterExcelImportService(store).preview(filename="template.xlsx", content=workbook_bytes())
        assert_equal(parsed["ok"], True, "template preview should pass")
        assert_equal(parsed["template_version"], TEMPLATE_VERSION, "template version")
        assert_equal((parsed.get("summary") or {}).get("picture_count"), 0, "new template imports no Excel image URL rows")
        confirm = ProductMasterExcelImportService(store).confirm(preview_id=parsed["preview_id"])
        assert_equal(confirm["imported_count"], 1, "template sample import")
        record = store.get_item("local_camry_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        name = ((detail.get("baseCarInfo") or {}).get("name") or {})
        assert_true(isinstance(name, dict), "baseCarInfo.name must be the standard nested object")
        assert_equal(name.get("displayValue"), "丰田 凯美瑞 2.0G", "displayValue")
        assert_equal(name.get("brandName"), "丰田", "brandName nested under baseCarInfo.name")
        assert_true("brandName" not in (detail.get("baseCarInfo") or {}), "brandName must not be written as a competing top-level baseCarInfo field")


def check_template_mirrors_frontend_v2_admin_fields_and_documents_image_url_boundary() -> None:
    targets = {spec.target for spec in VEHICLE_COLUMNS}
    raw_template_targets = {spec.target for spec in RAW_VEHICLE_TEMPLATE_COLUMNS}
    expected_raw_template_targets = {spec.target for spec in RAW_VEHICLE_TEMPLATE_COLUMNS}
    assert_true({"annotations.category", "manual.sku"} <= targets, "old simple template must keep local extension compatibility")
    assert_equal(raw_template_targets, expected_raw_template_targets, "new download template must be generated from the raw Dafengche vehicle_detail matrix plus local vehicle id")
    assert_true("detail.carModelParam.gearBoxType" in raw_template_targets, "new raw template must use Dafengche contract gearBoxType")
    assert_true("detail.carModelParam.engineVolumeLiter" in raw_template_targets, "new raw template must use Dafengche contract engineVolumeLiter")
    assert_true("detail.baseCarInfo.innerColor" in raw_template_targets, "new raw template must use Dafengche contract innerColor")
    assert_true("detail.carModelParam.gearbox" not in raw_template_targets, "new raw template must not expose local compatibility-only gearbox")
    assert_true("detail.carLicenseInfo.licenseStatus" not in raw_template_targets, "new raw template must not expose non-contract licenseStatus")
    for technical_target in (
        "detail.baseCarInfo.name.brandCode",
        "detail.baseCarInfo.name.seriesCode",
        "detail.baseCarInfo.name.modelCode",
        "detail.baseCarInfo.area.cityCode",
        "detail.baseCarInfo.area.provinceCode",
        "detail.baseCarInfo.registerArea.cityCode",
        "detail.baseCarInfo.registerArea.provinceCode",
    ):
        assert_true(technical_target not in raw_template_targets, f"new raw template must not ask users to fill technical code field: {technical_target}")
    assert_true("detail.baseCarInfo.area.cityName" in raw_template_targets and "detail.baseCarInfo.area.displayValue" in raw_template_targets, "new raw template must keep readable city/location fields editable")
    app_js = (APP_ROOT / "admin_backend" / "static" / "app.js").read_text(encoding="utf-8")
    assert_true("vehicleV2RawFieldGroupsEditorHtml" in app_js and "data-dafengche-path" in app_js, "frontend V2 editor must render the backend raw field matrix dynamically")
    workbook = load_workbook(BytesIO(build_template_bytes()), data_only=True)
    vehicle_headers = [str(cell.value or "") for cell in workbook[VEHICLE_SHEET][1]]
    assert_equal(vehicle_headers, list(VERTICAL_VEHICLE_HEADERS), "new download template must use the vertical vehicle sheet headers")
    assert_equal(vehicle_headers[0], "车辆编号 *", "new simple template must visibly mark vehicle id required")
    vehicle_labels = {str(row[1].value or "") for row in workbook[VEHICLE_SHEET].iter_rows(min_row=2)}
    expected_raw_labels = {_field_label_for_target(target) for target in expected_raw_template_targets if target != "local_id"}
    forbidden_local_extension_labels = {_field_label_for_target(target) for target in targets if target.startswith("annotations.") or target.startswith("manual.")}
    assert_equal(vehicle_labels, expected_raw_labels, "new simple vertical template rows must expose only raw vehicle business labels")
    assert_equal(vehicle_labels & forbidden_local_extension_labels, set(), "new simple template must not expose local extension labels")
    assert_true(PICTURE_SHEET not in workbook.sheetnames, "new template must not include URL image columns")
    visible_text = _workbook_visible_text(workbook)
    for forbidden in (
        "detail.",
        "baseCarInfo",
        "carPriceInfo",
        "carModelParam",
        "carLicenseInfo",
        "annotations.",
        "manual.",
        "source_payloads",
        "pictureUrl",
        "bigPictureUrl",
        "local_id",
        "target",
        "字段路径",
        "字段值",
        "assetFile",
        "图片地址",
        "大图地址",
        "图片网址",
        "brandCode",
        "seriesCode",
        "modelCode",
        "cityCode",
        "provinceCode",
        "车辆所在地城市 code",
        "车辆归属地城市 code",
        "店铺编码",
    ):
        assert_true(forbidden not in visible_text, f"user-visible template must hide internal field/path token: {forbidden}")
    assert_true("车辆图片请在管理台车辆编辑页上传" in visible_text, "template must direct users to the frontend image upload entry")
    assert_true("一次选择多张" in visible_text, "template must explain multi-image frontend upload")
    for forbidden_label in ("客户可见补充", "客户回复话术", "本地类目", "客户常用叫法", "客户可见卖点", "默认回复", "报价回复", "议价回复", "物流回复", "售后回复", "内部备注", "本地单位", "库存数量", "内部编号"):
        assert_true(forbidden_label not in visible_text, f"new simple template must not expose local extension label: {forbidden_label}")
    doc_text = (APP_ROOT / "docs" / "local_manual_vehicle_product_master_excel_import_20260801.md").read_text(encoding="utf-8")
    assert_true("未发现官方接口文档" in doc_text and "最小核心必填规则" in doc_text, "local doc must record the required-field evidence basis")
    assert_true("车辆编号 *" in doc_text and "车辆编号` 必填" in doc_text and "车辆展示名称` 必填" in doc_text and "品牌、车系、车型" in doc_text, "local doc must scope required fields narrowly")
    assert_true("Excel 只维护原始车辆字段" in doc_text and "本地客服扩展" in doc_text and "高级选项" in doc_text, "local doc must explain non-raw fields are maintained in frontend advanced options, not Excel")


def check_template_layout_styles_preserve_parser_contract() -> None:
    workbook = load_workbook(BytesIO(build_template_bytes()))
    vehicle = workbook[VEHICLE_SHEET]
    fields = workbook[FIELD_SHEET]

    assert_equal([cell.value for cell in vehicle[1]], list(VERTICAL_VEHICLE_HEADERS), "vehicle row 1 must remain raw parser headers")
    assert_equal(list(vehicle.merged_cells.ranges), [], "vehicle sheet must not merge parser header row")
    assert_true(PICTURE_SHEET not in workbook.sheetnames, "current download template intentionally omits the legacy image URL sheet")
    assert_equal(vehicle.freeze_panes, "A2", "vehicle sheet freezes header")
    assert_equal(fields.freeze_panes, "A4", "field sheet freezes to table header")
    assert_equal(vehicle.sheet_view.showGridLines, False, "vehicle gridlines disabled")
    assert_equal(fields.sheet_view.showGridLines, False, "field sheet gridlines disabled")

    assert_true(bool(vehicle["A1"].font.bold), "vehicle header is bold")
    assert_equal(str(vehicle["A1"].font.color.rgb)[-6:], "FFFFFF", "vehicle header uses white font")
    assert_equal(vehicle["A1"].alignment.wrap_text, True, "vehicle header wraps")
    assert_equal(vehicle["A1"].alignment.horizontal, "center", "vehicle header centered")
    assert_true((vehicle.row_dimensions[1].height or 0) >= 45, "vehicle header row height increased")
    assert_true(str(vehicle["B2"].value or "") and "detail." not in str(vehicle["B2"].value or ""), "first vertical field row should expose a business item label, not a target path")
    display_name_row = _vertical_row_for_path(vehicle, "detail.baseCarInfo.name.displayValue")
    display_name_instruction = str(vehicle.cell(row=display_name_row, column=4).value or "")
    assert_equal(display_name_instruction.count("必填"), 1, "required instruction should not repeat the required marker")
    assert_true(str(vehicle.cell(row=display_name_row, column=2).value or "") not in display_name_instruction, "field instruction should not repeat the field title")
    assert_equal(_fill_rgb(vehicle.cell(row=display_name_row, column=3)), "FFF8E1", "editable value cells should use the input highlight")
    sale_price_row = _vertical_row_for_path(vehicle, "detail.carPriceInfo.salePrice")
    sale_price_instruction = str(vehicle.cell(row=sale_price_row, column=4).value or "")
    assert_equal(sale_price_instruction.count("填写数字"), 1, "numeric instruction should not repeat the numeric marker")
    assert_true(_fill_rgb(vehicle["A2"]) != "FFFFFF", "vehicle sample field rows should have light fill")
    price_row = _vertical_row_for_path(vehicle, "detail.carPriceInfo.salePrice")
    seat_row = _vertical_row_for_path(vehicle, "detail.carModelParam.seatNumber")
    model_row = _vertical_row_for_path(vehicle, "detail.carModelParam.gearBoxType")
    assert_equal(vehicle.cell(row=price_row, column=3).number_format, "#,##0.00", "vehicle price cells use readable numeric format")
    assert_equal(vehicle.cell(row=seat_row, column=3).number_format, "#,##0.00", "vehicle numeric cells use readable numeric format")
    assert_true(_fill_rgb(vehicle.cell(row=price_row, column=1)) != _fill_rgb(vehicle.cell(row=model_row, column=1)), "vertical field groups should use visible group colors")

    assert_true(bool(fields["A1"].font.bold), "field version row highlighted")
    assert_equal(fields.row_dimensions[1].hidden, True, "machine template version row should be hidden from users")
    assert_true(bool(fields["A3"].font.bold), "field header row highlighted")
    assert_true("复制整组字段行后填写" in str(fields["B2"].value or ""), "field instructions mention copying a blank field group")
    assert_true("车辆图片请在管理台车辆编辑页上传" in str(fields["B2"].value or ""), "field instructions mention frontend image upload boundary")
    assert_true("一次选择多张" in str(fields["B2"].value or ""), "field instructions mention multi-image upload")
    assert_equal([cell.value for cell in fields[3]], ["填写项", "是否必填", "填写说明"], "field sheet must not expose target paths")
    required_marks = {str(row[0].value or ""): str(row[1].value or "") for row in fields.iter_rows(min_row=4)}
    assert_equal(required_marks.get("车辆编号"), "是", "field sheet must mark vehicle id required in user-readable Chinese")
    title_label = _field_label_for_target("detail.baseCarInfo.name.displayValue")
    brand_label = _field_label_for_target("detail.baseCarInfo.name.brandName")
    assert_equal(required_marks.get(title_label), "是", "field sheet must mark vehicle display name required in user-readable Chinese")
    assert_equal(required_marks.get(brand_label), "否", "field sheet must not require brand without an official rule")
    assert_equal(required_marks.get("默认回复"), None, "field sheet must not list local reply templates in the raw-fields download template")
    assert_true(_fill_rgb(fields["A4"]) != _fill_rgb(fields["A5"]), "field data rows use alternating fill")

    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_style_parse_check", root=Path(directory) / "pm")
        blank_preview = ProductMasterExcelImportService(store).preview(filename="blank-template.xlsx", content=build_template_bytes())
        assert_equal(blank_preview.get("ok"), False, "blank download template should require user input before confirm")
        preview = ProductMasterExcelImportService(store).preview(filename="styled-template.xlsx", content=workbook_bytes())
        assert_equal(preview.get("ok"), True, "filled styled template must remain parseable")
        legacy_header_workbook = load_workbook(BytesIO(workbook_bytes()))
        legacy_header_workbook[VEHICLE_SHEET]["A1"] = "车辆编号"
        legacy_header_buffer = BytesIO()
        legacy_header_workbook.save(legacy_header_buffer)
        legacy_header_preview = ProductMasterExcelImportService(store).preview(filename="legacy-simple-header.xlsx", content=legacy_header_buffer.getvalue())
        assert_equal(legacy_header_preview.get("ok"), True, "parser must keep accepting already downloaded simple templates with the old vehicle-id header")


def check_preview_does_not_write_and_confirm_writes_manual_v2() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_preview_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes())
        assert_equal(preview["ok"], True, "preview should pass")
        assert_equal(store.list_items(include_archived=True), [], "preview must not write product items")
        result = service.confirm(preview_id=preview["preview_id"])
        assert_equal(result["saved_ids"], ["local_camry_001"], "confirm saved id")
        record = store.get_item("local_camry_001", include_archived=True) or {}
        source = record.get("source") or {}
        assert_equal(source.get("type"), "manual", "source.type")
        assert_equal(source.get("provider"), "manual", "source.provider")
        assert_equal(((source.get("marker") or {}).get("ingest_channel")), "manual_input", "manual marker")
        assert_equal((source.get("binding") or {}).get("state"), "unbound", "manual records stay unbound")
        assert_true("shopCode" not in (source.get("binding") or {}), "Excel import must not fake shopCode")
        pictures = ((((record.get("source_payloads") or {}).get("vehicle_pictures") or {}).get("payload")) or [])
        assert_equal(len(pictures), 0, "new Excel template should not import image URL rows")


def check_complete_admin_fields_round_trip_and_customer_evidence_filter() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_full_admin_round_trip", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="full-admin.xlsx", content=workbook_bytes(full_admin_values=True))
        assert_equal(preview["ok"], True, "full admin fields preview")
        service.confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_camry_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        base = detail.get("baseCarInfo") or {}
        model = detail.get("carModelParam") or {}
        license_info = detail.get("carLicenseInfo") or {}
        price = detail.get("carPriceInfo") or {}
        name = base.get("name") or {}
        assert_equal(name.get("displayValue"), "丰田 凯美瑞 2.0G 豪华版", "title round trip")
        assert_equal(name.get("brandName"), "丰田", "brand nested round trip")
        assert_equal(base.get("color"), "白色备用", "color backup round trip")
        assert_equal(base.get("innerColor"), "黑色内饰", "innerColor round trip")
        assert_equal((base.get("area") or {}).get("cityName"), "南京", "city name round trip without asking for cityCode")
        assert_equal((base.get("area") or {}).get("displayValue"), "江苏南京", "area display value round trip")
        assert_true("cityCode" not in (base.get("area") or {}), "new template must not synthesize cityCode from readable city names")
        assert_equal((base.get("registerArea") or {}).get("cityName"), "苏州", "register area city name round trip without asking for cityCode")
        assert_equal(model.get("gearBoxType"), "CVT", "gearBoxType round trip")
        assert_equal(model.get("engineVolumeLiter"), "2.0L", "engineVolumeLiter round trip")
        assert_equal(license_info.get("keysCount"), 2, "keysCount round trip")
        assert_equal(price.get("purchasePrice"), 10.88, "purchasePrice can be stored locally")
        assert_equal(price.get("dealPrice"), 13.5, "dealPrice can be stored locally")
        assert_equal(price.get("managerPrice"), 13.3, "managerPrice can be stored locally")
        assert_equal(price.get("wholesalePrice"), 12.8, "wholesalePrice can be stored locally")
        wechat = (((record.get("extensions") or {}).get("wechat_customer_service") or {}))
        assert_equal(wechat.get("customer_visible_annotations") or {}, {}, "new raw-fields Excel template must not write local customer annotations")
        assert_equal(wechat.get("manual_annotations") or {}, {}, "new raw-fields Excel template must not write local manual annotations")
        pictures = ((((record.get("source_payloads") or {}).get("vehicle_pictures") or {}).get("payload")) or [])
        assert_equal(len(pictures), 0, "new complete admin Excel template should leave images to frontend upload")
        item = store.get_customer_evidence_item("local_camry_001") or {}
        evidence = item.get("customer_evidence") or {}
        assert_equal(evidence.get("price"), 13.98, "customer evidence only exposes customer-visible salePrice")
        combined = json.dumps(item, ensure_ascii=False)
        for forbidden in ("10.88", "13.5", "13.3", "12.8", "purchasePrice", "dealPrice", "salesPrice", "managerPrice", "wholesalePrice"):
            assert_true(forbidden not in combined, f"restricted/internal Excel field leaked into customer evidence: {forbidden}")


def check_missing_template_version_rejected_without_preview_artifact() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_version_check", root=Path(directory) / "pm")
        content = workbook_bytes(remove_field_sheet=True)
        try:
            ProductMasterExcelImportService(store).preview(filename="bad.xlsx", content=content)
        except ValueError as exc:
            assert_true("字段说明" in str(exc), "must reject missing version sheet")
        else:
            raise AssertionError("missing template version should be rejected")
        preview_root = store.root / ".excel_import_previews"
        assert_true(not preview_root.exists() or not list(preview_root.glob("*.json")), "invalid template must not create confirmable preview artifacts")


def check_legacy_template_version_remains_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_legacy_template_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(filename="legacy-template.xlsx", content=legacy_workbook_bytes())
        assert_equal(preview["ok"], True, "legacy template preview should remain readable")
        assert_equal(preview["template_version"], LEGACY_TEMPLATE_VERSION, "legacy template version should be reported")
        ProductMasterExcelImportService(store).confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_legacy_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        name = ((detail.get("baseCarInfo") or {}).get("name") or {})
        assert_equal(name.get("displayValue"), "旧模板车辆", "legacy template display name")
        assert_equal((detail.get("carModelParam") or {}).get("gearBoxType"), "自动", "legacy template keeps old gearBoxType path")


def check_horizontal_admin_template_version_remains_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_horizontal_template_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(filename="horizontal-admin-template.xlsx", content=horizontal_admin_workbook_bytes())
        assert_equal(preview["ok"], True, "horizontal admin template preview should remain readable")
        assert_equal(preview["template_version"], "local_manual_vehicle_v2_excel_20260801_admin_fields", "horizontal admin template version should be reported")
        ProductMasterExcelImportService(store).confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_horizontal_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        assert_equal(((detail.get("baseCarInfo") or {}).get("name") or {}).get("displayValue"), "丰田 凯美瑞 2.0G", "horizontal template display name")
        assert_equal((detail.get("carModelParam") or {}).get("gearbox"), "自动", "horizontal template keeps admin gearbox path")


def check_path_vertical_template_version_remains_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_path_vertical_template_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(filename="path-vertical-template.xlsx", content=path_vertical_workbook_bytes())
        assert_equal(preview["ok"], True, "previous path-vertical template preview should remain readable")
        assert_equal(preview["template_version"], PATH_VERTICAL_TEMPLATE_VERSION, "path-vertical template version should be reported")
        ProductMasterExcelImportService(store).confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_path_vertical_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        assert_equal(((detail.get("baseCarInfo") or {}).get("name") or {}).get("displayValue"), "路径竖表兼容车", "path vertical template display name")


def check_old_simple_vertical_local_extension_fields_remain_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_old_simple_extension_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(
            filename="old-simple-local-extensions.xlsx",
            content=old_simple_vertical_workbook_bytes(),
        )
        assert_equal(preview["ok"], True, "old simple template with local extension rows should remain readable")
        assert_equal(preview["template_version"], SIMPLE_VERTICAL_TEMPLATE_VERSION, "old simple template version should be reported")
        ProductMasterExcelImportService(store).confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_old_simple_001", include_archived=True) or {}
        wechat = (((record.get("extensions") or {}).get("wechat_customer_service") or {}))
        annotations = wechat.get("customer_visible_annotations") or {}
        manual = wechat.get("manual_annotations") or {}
        assert_equal(annotations.get("category"), "used_car", "old simple local category row")
        assert_equal(manual.get("unit"), "台", "old simple manual unit row")
        templates = manual.get("reply_templates") or {}
        assert_equal(templates.get("default"), "旧 simple 默认回复", "old simple reply template row")


def check_simple_vertical_url_picture_sheet_remains_readable() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_simple_url_picture_compat_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(
            filename="simple-vertical-url-picture-template.xlsx",
            content=old_simple_vertical_workbook_bytes(include_picture_sheet=True),
        )
        assert_equal(preview["ok"], True, "previous simple vertical URL picture sheet remains readable")
        assert_equal((preview.get("summary") or {}).get("picture_count"), 1, "compat URL picture row should be counted")
        ProductMasterExcelImportService(store).confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_old_simple_001", include_archived=True) or {}
        pictures = ((((record.get("source_payloads") or {}).get("vehicle_pictures") or {}).get("payload")) or [])
        assert_equal(len(pictures), 1, "compat URL picture row should still import")
        assert_equal(pictures[0].get("pictureUrl"), "https://example.invalid/camry-front.jpg", "compat picture url")
        assert_true("assetFile" not in json.dumps(pictures, ensure_ascii=False), "Excel URL image import must not forge admin upload assetFile")


def check_unknown_column_and_missing_required_field_are_row_errors() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_unknown_column_check", root=Path(directory) / "pm")
        content = workbook_bytes(add_unknown_vehicle_column=True, blank_required_name=True)
        preview = ProductMasterExcelImportService(store).preview(filename="bad.xlsx", content=content)
        codes = {error.get("code") for error in preview.get("errors") or []}
        assert_true("unknown_column" in codes, "unknown columns must not be silently ignored")
        assert_true("required_missing" in codes or "missing_column" in codes, "missing required field must be reported")
        assert_equal(store.list_items(include_archived=True), [], "failed preview must not write product items")


def check_duplicate_header_is_rejected() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_duplicate_header_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(
            filename="duplicate-header.xlsx",
            content=workbook_bytes(add_duplicate_vehicle_header=True),
        )
        codes = {error.get("code") for error in preview.get("errors") or []}
        assert_true("duplicate_column" in codes, "duplicate headers must not be silently overwritten")
        assert_equal(store.list_items(include_archived=True), [], "duplicate header preview must not write")


def check_vehicle_and_picture_row_limits_are_errors() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_row_limit_check", root=Path(directory) / "pm")
        vehicle_preview = ProductMasterExcelImportService(store).preview(
            filename="too-many-vehicles.xlsx",
            content=workbook_bytes(extra_vehicle_rows=MAX_IMPORT_ROWS),
        )
        vehicle_codes = {error.get("code") for error in vehicle_preview.get("errors") or []}
        assert_true("too_many_rows" in vehicle_codes, "vehicle row overflow must be reported")
        assert_equal(store.list_items(include_archived=True), [], "vehicle overflow preview must not write")
        picture_preview = ProductMasterExcelImportService(store).preview(
            filename="too-many-pictures.xlsx",
            content=workbook_bytes(include_picture_sheet=True, extra_picture_rows=MAX_PICTURE_ROWS),
        )
        picture_codes = {error.get("code") for error in picture_preview.get("errors") or []}
        assert_true("too_many_rows" in picture_codes, "picture row overflow must be reported")
        assert_equal(store.list_items(include_archived=True), [], "picture overflow preview must not write")


def check_duplicate_field_item_and_picture_row_errors() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_duplicate_check", root=Path(directory) / "pm")
        preview = ProductMasterExcelImportService(store).preview(
            filename="duplicate.xlsx",
            content=workbook_bytes(add_duplicate_field_path=True, add_unknown_picture_local_id=True),
        )
        codes = {error.get("code") for error in preview.get("errors") or []}
        assert_true("duplicate_field_item" in codes, "duplicate vehicle id + field item must be rejected")
        assert_true("unknown_local_id" in codes, "picture rows must reference known local_id")
        assert_equal(store.list_items(include_archived=True), [], "failed preview must not write")


def check_repeat_import_preserves_existing_manual_fields() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_repeat_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        first = service.preview(filename="first.xlsx", content=workbook_bytes())
        service.confirm(preview_id=first["preview_id"])
        second = service.preview(filename="second.xlsx", content=workbook_bytes(blank_brand=True, sale_price=14.88))
        assert_equal((second.get("vehicles") or [])[0].get("action"), "update", "repeat import should update existing local_id")
        service.confirm(preview_id=second["preview_id"])
        record = store.get_item("local_camry_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        name = ((detail.get("baseCarInfo") or {}).get("name") or {})
        assert_equal(name.get("brandName"), "丰田", "blank Excel brand must not clear existing manual field")
        price = ((detail.get("carPriceInfo") or {}).get("salePrice"))
        assert_equal(price, 14.88, "nonblank Excel fields should update")


def check_confirm_rollback_prevents_half_batch_pollution() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_rollback_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="two.xlsx", content=workbook_bytes(add_second_vehicle=True))
        original_write_json = product_master_module.write_json

        def flaky_write_json(path: Path, payload: Any) -> None:
            if str(path).endswith("local_civic_002.json"):
                raise RuntimeError("injected file batch failure")
            original_write_json(path, payload)

        product_master_module.write_json = flaky_write_json  # type: ignore[assignment]
        try:
            try:
                service.confirm(preview_id=preview["preview_id"])
            except RuntimeError:
                pass
            else:
                raise AssertionError("injected save failure should bubble")
        finally:
            product_master_module.write_json = original_write_json  # type: ignore[assignment]
        assert_equal(store.list_items(include_archived=True), [], "failed confirm must roll back already written rows")


def check_confirm_uses_postgres_batch_transaction() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_postgres_batch_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes(add_second_vehicle=True))
        fake_db = FakePostgresStore()
        original_pg = product_master_module.postgres_store
        original_config = product_master_module.load_storage_config
        product_master_module.postgres_store = lambda tenant_id: fake_db  # type: ignore[assignment]
        product_master_module.load_storage_config = lambda: SimpleNamespace(mirror_files=False)  # type: ignore[assignment]
        try:
            result = service.confirm(preview_id=preview["preview_id"])
        finally:
            product_master_module.postgres_store = original_pg  # type: ignore[assignment]
            product_master_module.load_storage_config = original_config  # type: ignore[assignment]
        assert_equal(result.get("storage"), "postgres", "PostgreSQL batch storage mode")
        assert_equal(result.get("imported_count"), 2, "PostgreSQL batch import count")
        assert_equal(fake_db.atomic_calls, 1, "PostgreSQL import must use exactly one batch transaction")
        assert_equal(sorted(fake_db.items), ["local_camry_001", "local_civic_002"], "DB must contain the full batch")


def check_postgres_batch_failure_writes_zero_items() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_postgres_failure_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes(add_second_vehicle=True))
        fake_db = FakePostgresStore(fail_atomic=True)
        original_pg = product_master_module.postgres_store
        original_config = product_master_module.load_storage_config
        product_master_module.postgres_store = lambda tenant_id: fake_db  # type: ignore[assignment]
        product_master_module.load_storage_config = lambda: SimpleNamespace(mirror_files=False)  # type: ignore[assignment]
        try:
            try:
                service.confirm(preview_id=preview["preview_id"])
            except RuntimeError:
                pass
            else:
                raise AssertionError("PostgreSQL atomic failure should bubble")
        finally:
            product_master_module.postgres_store = original_pg  # type: ignore[assignment]
            product_master_module.load_storage_config = original_config  # type: ignore[assignment]
        assert_equal(fake_db.items, {}, "failed PostgreSQL transaction must leave DB empty")
        assert_equal(store.list_items(include_archived=True), [], "failed PostgreSQL transaction must not fall back to file writes")


def check_postgres_mirror_failure_warns_and_db_remains_canonical() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_postgres_mirror_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes(add_second_vehicle=True))
        fake_db = FakePostgresStore()
        original_pg = product_master_module.postgres_store
        original_config = product_master_module.load_storage_config
        original_write_json = product_master_module.write_json

        def failing_mirror_write_json(path: Path, payload: Any) -> None:
            if f"{Path('items')}" in str(path):
                raise OSError("injected mirror failure")
            original_write_json(path, payload)

        product_master_module.postgres_store = lambda tenant_id: fake_db  # type: ignore[assignment]
        product_master_module.load_storage_config = lambda: SimpleNamespace(mirror_files=True)  # type: ignore[assignment]
        product_master_module.write_json = failing_mirror_write_json  # type: ignore[assignment]
        try:
            result = service.confirm(preview_id=preview["preview_id"])
            db_list = store.list_items(include_archived=True)
        finally:
            product_master_module.write_json = original_write_json  # type: ignore[assignment]
            product_master_module.postgres_store = original_pg  # type: ignore[assignment]
            product_master_module.load_storage_config = original_config  # type: ignore[assignment]
        assert_equal(result.get("ok"), True, "mirror failure must not turn a committed DB batch into a failed confirm")
        warnings = result.get("warnings") or []
        assert_true(any(warning.get("code") == "mirror_files_failed" for warning in warnings), "mirror failure must be returned as a warning")
        mirror = result.get("mirror_files") or {}
        assert_equal(sorted(mirror.get("failed_ids") or []), ["local_camry_001", "local_civic_002"], "failed mirror ids")
        assert_equal(sorted(fake_db.items), ["local_camry_001", "local_civic_002"], "DB remains canonical after mirror failure")
        assert_equal(sorted(item.get("id") for item in db_list), ["local_camry_001", "local_civic_002"], "list_items must prefer DB records when DB is available")


def check_product_master_batch_rejects_duplicate_item_ids() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_duplicate_batch_check", root=Path(directory) / "pm")
        record = manual_vehicle_record("local_dup_001", "重复车辆")
        result = store._save_items_batch_atomic([record, deepcopy(record)])
        assert_equal(result.get("ok"), False, "batch duplicate item ids must be rejected before writing")
        assert_true(any("duplicate item id" in str(problem) for problem in result.get("problems") or []), "duplicate item id problem")
        assert_equal(store.list_items(include_archived=True), [], "duplicate batch must not write files")


def check_excel_manual_customer_evidence_filters_restricted_raw_payload_fields() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_evidence_filter_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes())
        service.confirm(preview_id=preview["preview_id"])
        record = store.get_item("local_camry_001", include_archived=True) or {}
        detail = (((record.get("source_payloads") or {}).get("vehicle_detail") or {}).get("payload") or {})
        detail.setdefault("baseCarInfo", {})["vinNumber"] = "TEST-EXCEL-VIN-MUST-NOT-LEAK"
        detail.setdefault("baseCarInfo", {})["plateNumber"] = "TEST-EXCEL-PLATE-MUST-NOT-LEAK"
        detail.setdefault("baseCarInfo", {})["vehicleNumber"] = "TEST-EXCEL-INTERNAL-IDENTITY-MUST-NOT-LEAK"
        detail.setdefault("baseCarInfo", {})["salesperson"] = "TEST-EXCEL-SALESPERSON-MUST-NOT-LEAK"
        detail.setdefault("carOwnerInfo", {})["phoneNumber"] = "TEST-EXCEL-OWNER-PHONE-MUST-NOT-LEAK"
        detail.setdefault("carOwnerInfo", {})["identify"] = "TEST-EXCEL-OWNER-ID-MUST-NOT-LEAK"
        detail.setdefault("carModelParam", {})["engineNumber"] = "TEST-EXCEL-ENGINE-MUST-NOT-LEAK"
        detail.setdefault("carPriceInfo", {})["purchasePrice"] = 5.11
        detail.setdefault("carPriceInfo", {})["salesPrice"] = 6.22
        detail.setdefault("carPriceInfo", {})["wholesalePrice"] = 7.33
        detail.setdefault("carPriceInfo", {})["managerPrice"] = 8.44
        saved = store.save_item(record)
        assert_equal(saved.get("ok"), True, "mutated manual V2 record should save")
        item = store.get_customer_evidence_item("local_camry_001")
        assert_true(isinstance(item, dict), "Excel manual V2 item should project through customer evidence seam")
        evidence = (item or {}).get("customer_evidence") or {}
        assert_equal(evidence.get("price"), 13.98, "only customer-visible salePrice may enter evidence")
        evidence_text = json.dumps(item, ensure_ascii=False)
        brain_compact_text = json.dumps(compact_product_item_for_brain_prompt(evidence, max_text_chars=240), ensure_ascii=False)
        combined = f"{evidence_text}\n{brain_compact_text}"
        for forbidden in (
            "source_payloads",
            "TEST-EXCEL-VIN-MUST-NOT-LEAK",
            "TEST-EXCEL-PLATE-MUST-NOT-LEAK",
            "TEST-EXCEL-INTERNAL-IDENTITY-MUST-NOT-LEAK",
            "TEST-EXCEL-SALESPERSON-MUST-NOT-LEAK",
            "TEST-EXCEL-OWNER-PHONE-MUST-NOT-LEAK",
            "TEST-EXCEL-OWNER-ID-MUST-NOT-LEAK",
            "TEST-EXCEL-ENGINE-MUST-NOT-LEAK",
            "5.11",
            "6.22",
            "7.33",
            "8.44",
        ):
            assert_true(forbidden not in combined, f"restricted Excel raw field leaked into customer evidence/Brain compact: {forbidden}")


def check_excel_manual_customer_evidence_tenant_and_shop_scope_isolated() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        store_a = ProductMasterStore(tenant_id="excel_scope_a", root=root / "tenant_a")
        store_b = ProductMasterStore(tenant_id="excel_scope_b", root=root / "tenant_b")
        service_a = ProductMasterExcelImportService(store_a)
        preview = service_a.preview(filename="vehicles.xlsx", content=workbook_bytes())
        service_a.confirm(preview_id=preview["preview_id"])
        assert_equal(store_b.list_customer_evidence_items(shop_code=None), [], "tenant B must not read tenant A Excel records")

        record = store_a.get_item("local_camry_001", include_archived=True) or {}
        record.setdefault("source", {})["binding"] = {"state": "bound", "shopCode": "SHOP-A"}
        saved = store_a.save_item(record)
        assert_equal(saved.get("ok"), True, "manual V2 shop binding test record should save")
        assert_equal(len(store_a.list_customer_evidence_items(shop_code="SHOP-A")), 1, "matching shop scope should see bound manual record")
        assert_equal(store_a.list_customer_evidence_items(shop_code="SHOP-B"), [], "wrong shop scope must not see bound manual record")
        assert_equal(store_a.list_customer_evidence_items(shop_code=None), [], "bound manual record requires explicit matching shop scope")


def check_product_console_excel_api_contract() -> None:
    with tempfile.TemporaryDirectory() as directory:
        console_service = ProductConsoleService()
        console_service.store.product_master = ProductMasterStore(tenant_id="excel_api_check", root=Path(directory) / "pm")  # type: ignore[assignment]
        original_service = product_console_api.service
        product_console_api.service = lambda: console_service  # type: ignore[assignment]
        try:
            app = FastAPI()
            app.include_router(product_console_api.router)
            client = TestClient(app)
            template = client.get("/api/product-console/local-vehicle-excel-template")
            assert_equal(template.status_code, 200, "template endpoint status")
            assert_true(template.content.startswith(b"PK"), "template endpoint must return xlsx bytes")
            preview = client.post(
                "/api/product-console/local-vehicle-excel-import/preview",
                files={"file": ("vehicles.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert_equal(preview.status_code, 200, "preview endpoint status")
            body = preview.json()
            assert_equal(body.get("ok"), True, "preview endpoint body")
            confirm = client.post("/api/product-console/local-vehicle-excel-import/confirm", json={"preview_id": body.get("preview_id")})
            assert_equal(confirm.status_code, 200, "confirm endpoint status")
            assert_equal(confirm.json().get("imported_count"), 1, "confirm endpoint count")
        finally:
            product_console_api.service = original_service  # type: ignore[assignment]


def check_unauthenticated_template_download_only_allows_template() -> None:
    with tempfile.TemporaryDirectory() as directory:
        console_service = ProductConsoleService()
        console_service.store.product_master = ProductMasterStore(tenant_id="excel_auth_check", root=Path(directory) / "pm")  # type: ignore[assignment]
        original_service = product_console_api.service
        product_console_api.service = lambda: console_service  # type: ignore[assignment]
        try:
            app = FastAPI()
            app.include_router(product_console_api.router)
            app.add_middleware(AuthTenantMiddleware)
            client = TestClient(app)

            template = client.get("/api/product-console/local-vehicle-excel-template")
            assert_equal(template.status_code, 200, "unauthenticated template download should be public")
            assert_true(template.content.startswith(b"PK"), "public template response must be xlsx bytes")
            assert_true(
                "spreadsheetml.sheet" in str(template.headers.get("content-type") or ""),
                "public template content type must be xlsx",
            )

            catalog = client.get("/api/product-console/catalog")
            assert_equal(catalog.status_code, 401, "unauthenticated catalog must remain protected")
            assert_equal((catalog.json() or {}).get("detail"), "authentication required", "catalog auth error")

            preview = client.post(
                "/api/product-console/local-vehicle-excel-import/preview",
                files={"file": ("vehicles.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
            assert_equal(preview.status_code, 401, "unauthenticated Excel preview must remain protected")

            confirm = client.post("/api/product-console/local-vehicle-excel-import/confirm", json={"preview_id": "preview_x"})
            assert_equal(confirm.status_code, 401, "unauthenticated Excel confirm must remain protected")
        finally:
            product_console_api.service = original_service  # type: ignore[assignment]


def check_preview_is_tenant_scoped() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        store_a = ProductMasterStore(tenant_id="tenant_a", root=root / "a")
        store_b = ProductMasterStore(tenant_id="tenant_b", root=root / "b")
        preview = ProductMasterExcelImportService(store_a).preview(filename="vehicles.xlsx", content=workbook_bytes())
        try:
            ProductMasterExcelImportService(store_b).confirm(preview_id=preview["preview_id"])
        except FileNotFoundError:
            pass
        else:
            raise AssertionError("tenant B must not confirm tenant A preview")
        assert_equal(store_b.list_items(include_archived=True), [], "tenant B must remain isolated")


def check_import_service_has_no_network_llm_or_dafengche_client_imports() -> None:
    source_path = APP_ROOT / "admin_backend" / "services" / "product_master_excel_import.py"
    tree = ast.parse(source_path.read_text(encoding="utf-8"))
    imports: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imports.extend(alias.name for alias in node.names)
        if isinstance(node, ast.ImportFrom):
            imports.append(node.module or "")
    forbidden = ("urllib", "requests", "llm_config", "knowledge_generator", "dafengche_product_master.client")
    for item in imports:
        assert_true(not any(item == bad or item.startswith(f"{bad}.") for bad in forbidden), f"Excel import must not import network/LLM/API client modules: {item}")


def check_customer_evidence_does_not_leak_raw_source_payloads() -> None:
    with tempfile.TemporaryDirectory() as directory:
        store = ProductMasterStore(tenant_id="excel_evidence_check", root=Path(directory) / "pm")
        service = ProductMasterExcelImportService(store)
        preview = service.preview(filename="vehicles.xlsx", content=workbook_bytes())
        service.confirm(preview_id=preview["preview_id"])
        evidence_items = store.list_customer_evidence_items()
        assert_equal(len(evidence_items), 1, "manual Excel vehicle should produce customer evidence")
        evidence_text = str(evidence_items[0])
        assert_true("source_payloads" not in evidence_text, "customer evidence must not leak raw source_payloads")
        assert_true("manual.dafengche_shaped_vehicle_detail" not in evidence_text, "customer evidence must not expose raw snapshot metadata")


def workbook_bytes(
    *,
    include_picture_sheet: bool = False,
    add_second_vehicle: bool = False,
    add_duplicate_field_path: bool = False,
    add_unknown_picture_local_id: bool = False,
    add_unknown_vehicle_column: bool = False,
    add_duplicate_vehicle_header: bool = False,
    extra_vehicle_rows: int = 0,
    extra_picture_rows: int = 0,
    blank_required_name: bool = False,
    blank_brand: bool = False,
    remove_field_sheet: bool = False,
    sale_price: float = 13.98,
    full_admin_values: bool = False,
) -> bytes:
    workbook = load_workbook(BytesIO(build_template_bytes()))
    vehicle = workbook[VEHICLE_SHEET]
    _fill_vertical_vehicle_group(vehicle, "local_camry_001", _sample_values_by_target())
    picture = None
    if include_picture_sheet or add_unknown_picture_local_id or extra_picture_rows:
        picture = workbook.create_sheet(PICTURE_SHEET)
        picture.append([spec.header for spec in SIMPLE_PICTURE_COLUMNS])
        picture.append(["local_camry_001", "https://example.invalid/camry-front.jpg", "", 1, "车头45度"])
    _set_vertical_value(vehicle, "local_camry_001", "detail.carPriceInfo.salePrice", sale_price)
    if full_admin_values:
        full_values = {
            "detail.baseCarInfo.name.displayValue": "丰田 凯美瑞 2.0G 豪华版",
            "detail.baseCarInfo.color": "白色备用",
            "detail.baseCarInfo.innerColor": "黑色内饰",
            "detail.baseCarInfo.area.cityName": "南京",
            "detail.baseCarInfo.area.provinceName": "江苏",
            "detail.baseCarInfo.area.displayValue": "江苏南京",
            "detail.baseCarInfo.registerArea.cityName": "苏州",
            "detail.baseCarInfo.registerArea.provinceName": "江苏",
            "detail.baseCarInfo.registerArea.displayValue": "江苏苏州",
            "detail.carModelParam.gearBoxType": "CVT",
            "detail.carModelParam.engineVolumeLiter": "2.0L",
            "detail.carLicenseInfo.keysCount": 2,
            "detail.carPriceInfo.purchasePrice": 10.88,
            "detail.carPriceInfo.dealPrice": 13.50,
            "detail.carPriceInfo.managerPrice": 13.30,
            "detail.carPriceInfo.wholesalePrice": 12.80,
        }
        for field_path, value in full_values.items():
            _set_vertical_value(vehicle, "local_camry_001", field_path, value)
    if blank_required_name:
        _set_vertical_value(vehicle, "local_camry_001", "detail.baseCarInfo.name.displayValue", "")
    if blank_brand:
        _set_vertical_value(vehicle, "local_camry_001", "detail.baseCarInfo.name.brandName", "")
    if add_unknown_vehicle_column:
        vehicle.cell(row=1, column=vehicle.max_column + 1).value = "未知字段"
        vehicle.cell(row=2, column=vehicle.max_column).value = "should fail"
    if add_duplicate_vehicle_header:
        vehicle.cell(row=1, column=vehicle.max_column + 1).value = "车辆编号"
        vehicle.cell(row=2, column=vehicle.max_column).value = "shadow_id"
    if add_second_vehicle:
        _append_vertical_vehicle_group(
            vehicle,
            source_local_id="local_camry_001",
            new_local_id="local_civic_002",
            replacements={
                "detail.baseCarInfo.name.displayValue": "本田 思域 1.5T",
                "detail.baseCarInfo.name.brandName": "本田",
                "detail.baseCarInfo.name.seriesName": "思域",
                "detail.baseCarInfo.carName": "思域",
            },
        )
    for index in range(extra_vehicle_rows):
        vehicle.append([f"local_limit_{index:04d}", _field_label_for_target("detail.baseCarInfo.name.displayValue"), f"测试车 {index:04d}", "limit row"])
    if add_duplicate_field_path:
        title_row = _vertical_row_for_path(vehicle, "detail.baseCarInfo.name.displayValue", local_id="local_camry_001")
        vehicle.append([vehicle.cell(row=title_row, column=column).value for column in range(1, 5)])
    if add_unknown_picture_local_id:
        if picture is None:
            picture = workbook.create_sheet(PICTURE_SHEET)
            picture.append([spec.header for spec in SIMPLE_PICTURE_COLUMNS])
        picture.append(["missing_vehicle", "https://example.invalid/missing.jpg", "", 1, "bad row"])
    for index in range(extra_picture_rows):
        if picture is None:
            picture = workbook.create_sheet(PICTURE_SHEET)
            picture.append([spec.header for spec in SIMPLE_PICTURE_COLUMNS])
        picture.append(["local_camry_001", f"https://example.invalid/camry-{index:04d}.jpg", "", index + 2, "limit row"])
    if remove_field_sheet:
        workbook.remove(workbook[FIELD_SHEET])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def old_simple_vertical_workbook_bytes(*, include_picture_sheet: bool = False) -> bytes:
    workbook = Workbook()
    vehicle = workbook.active
    vehicle.title = VEHICLE_SHEET
    vehicle.append(list(VERTICAL_VEHICLE_HEADERS))
    sample = _sample_values_by_target()
    sample.update(
        {
            "detail.baseCarInfo.name.displayValue": "旧 simple 兼容车",
            "annotations.category": "used_car",
            "manual.unit": "台",
            "manual.reply_templates.default": "旧 simple 默认回复",
        }
    )
    for spec in VEHICLE_COLUMNS:
        if spec.target == "local_id":
            continue
        vehicle.append(["local_old_simple_001", _simple_field_label(spec), sample.get(spec.target), _vertical_field_instruction_for_test(spec)])
    if include_picture_sheet:
        picture = workbook.create_sheet(PICTURE_SHEET)
        picture.append([spec.header for spec in SIMPLE_PICTURE_COLUMNS])
        picture.append(["local_old_simple_001", "https://example.invalid/camry-front.jpg", "", 1, "车头45度"])
    field = workbook.create_sheet(FIELD_SHEET)
    field.append(["模板版本", SIMPLE_VERTICAL_TEMPLATE_VERSION])
    field.append(["填写说明", "旧 simple 竖表兼容。"])
    field.append(["填写项", "是否必填", "填写说明"])
    field.append(["车辆编号", "是", "每组字段都填写同一个车辆编号，用于识别车辆。"])
    for spec in VEHICLE_COLUMNS:
        if spec.target == "local_id":
            continue
        field.append([_simple_field_label(spec), "是" if spec.required else "否", _vertical_field_instruction_for_test(spec)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def legacy_workbook_bytes() -> bytes:
    workbook = Workbook()
    vehicle = workbook.active
    vehicle.title = VEHICLE_SHEET
    vehicle.append([spec.header for spec in LEGACY_VEHICLE_COLUMNS])
    vehicle.append(
        [
            "local_legacy_001",
            "旧模板车辆",
            "测试品牌",
            "测试车系",
            "测试车型",
            "2020-01",
            "3.5万公里",
            "车况良好",
            "白色",
            "黑色",
            "自动",
            "2.0L",
            "汽油",
            "国六",
            5,
            12.34,
            18.88,
            "SALE",
            "旧车；旧模板",
            "旧模板卖点",
            "旧模板看车政策",
            "旧模板售后政策",
            "旧模板风险",
            "旧模板补充",
            "LEGACY-001",
            1,
            "旧模板默认回复",
        ]
    )
    picture = workbook.create_sheet(PICTURE_SHEET)
    picture.append([spec.header for spec in PICTURE_COLUMNS])
    picture.append(["local_legacy_001", "https://example.invalid/legacy.jpg", "", 1, "旧模板图片"])
    field = workbook.create_sheet(FIELD_SHEET)
    field.append(["template_version", LEGACY_TEMPLATE_VERSION])
    field.append([])
    field.append(["sheet", "column", "required", "target", "note"])
    for spec in LEGACY_VEHICLE_COLUMNS:
        field.append([VEHICLE_SHEET, spec.header, "yes" if spec.required else "no", spec.target, spec.note])
    for spec in PICTURE_COLUMNS:
        field.append([PICTURE_SHEET, spec.header, "yes" if spec.required else "no", spec.target, spec.note])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def horizontal_admin_workbook_bytes() -> bytes:
    workbook = Workbook()
    vehicle = workbook.active
    vehicle.title = VEHICLE_SHEET
    vehicle.append([spec.header for spec in VEHICLE_COLUMNS])
    sample = _sample_values_by_target()
    vehicle.append(["local_horizontal_001", *[sample.get(spec.target) for spec in VEHICLE_COLUMNS if spec.target != "local_id"]])
    picture = workbook.create_sheet(PICTURE_SHEET)
    picture.append([spec.header for spec in PICTURE_COLUMNS])
    picture.append(["local_horizontal_001", "https://example.invalid/horizontal.jpg", "", 1, "横向兼容图片"])
    field = workbook.create_sheet(FIELD_SHEET)
    field.append(["template_version", "local_manual_vehicle_v2_excel_20260801_admin_fields"])
    field.append([])
    field.append(["sheet", "column", "required", "target", "note"])
    for spec in VEHICLE_COLUMNS:
        field.append([VEHICLE_SHEET, spec.header, "yes" if spec.required else "no", spec.target, spec.note])
    for spec in PICTURE_COLUMNS:
        field.append([PICTURE_SHEET, spec.header, "yes" if spec.required else "no", spec.target, spec.note])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def path_vertical_workbook_bytes() -> bytes:
    workbook = Workbook()
    vehicle = workbook.active
    vehicle.title = VEHICLE_SHEET
    vehicle.append(list(PATH_VERTICAL_VEHICLE_HEADERS))
    sample = _sample_values_by_target()
    sample["detail.baseCarInfo.name.displayValue"] = "路径竖表兼容车"
    for spec in VEHICLE_COLUMNS:
        if spec.target == "local_id":
            continue
        vehicle.append(["local_path_vertical_001", spec.target, sample.get(spec.target), spec.note])
    picture = workbook.create_sheet(PICTURE_SHEET)
    picture.append([spec.header for spec in PICTURE_COLUMNS])
    picture.append(["local_path_vertical_001", "https://example.invalid/path-vertical.jpg", "", 1, "字段路径竖表兼容图片"])
    field = workbook.create_sheet(FIELD_SHEET)
    field.append(["template_version", PATH_VERTICAL_TEMPLATE_VERSION])
    field.append([])
    field.append(["sheet", "column", "required", "target", "note"])
    for spec in VEHICLE_COLUMNS:
        field.append([VEHICLE_SHEET, "字段路径", "yes" if spec.required else "no", spec.target, spec.note])
    for spec in PICTURE_COLUMNS:
        field.append([PICTURE_SHEET, spec.header, "yes" if spec.required else "no", spec.target, spec.note])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _vertical_row_for_path(sheet: Any, field_path: str, *, local_id: str | None = None) -> int:
    label = _field_label_for_target(field_path)
    for row in range(2, sheet.max_row + 1):
        if local_id is not None and str(sheet.cell(row=row, column=1).value or "") != local_id:
            continue
        if str(sheet.cell(row=row, column=2).value or "") == label:
            return row
    raise AssertionError(f"missing vertical field item: {field_path} / {label}")


def _set_vertical_value(sheet: Any, local_id: str, field_path: str, value: Any) -> None:
    row = _vertical_row_for_path(sheet, field_path, local_id=local_id)
    sheet.cell(row=row, column=3).value = value


def _fill_vertical_vehicle_group(sheet: Any, local_id: str, values_by_target: dict[str, Any]) -> None:
    for row in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row=row, column=2).value or "")
        if not label:
            continue
        target = _target_for_field_label(label)
        sheet.cell(row=row, column=1).value = local_id
        sheet.cell(row=row, column=3).value = values_by_target.get(target)


def _append_vertical_vehicle_group(sheet: Any, *, source_local_id: str, new_local_id: str, replacements: dict[str, Any]) -> None:
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value or "") != source_local_id:
            continue
        field_label = str(sheet.cell(row=row, column=2).value or "")
        field_path = _target_for_field_label(field_label)
        value = replacements.get(field_path, sheet.cell(row=row, column=3).value)
        sheet.append([new_local_id, field_label, value, sheet.cell(row=row, column=4).value])


def _field_label_for_target(target: str) -> str:
    for spec in (*RAW_VEHICLE_TEMPLATE_COLUMNS, *VEHICLE_COLUMNS):
        if spec.target == target:
            return _simple_field_label(spec)
    raise AssertionError(f"missing target in vehicle columns: {target}")


def _vertical_field_instruction_for_test(spec: Any) -> str:
    label = _simple_field_label(spec)
    parts = [label]
    if getattr(spec, "required", False):
        parts.append("必填")
    note = str(getattr(spec, "note", "") or "")
    if note:
        parts.append(note)
    return "；".join(parts)


def _target_for_field_label(label: str) -> str:
    for spec in (*RAW_VEHICLE_TEMPLATE_COLUMNS, *VEHICLE_COLUMNS):
        if _simple_field_label(spec) == label:
            return spec.target
    raise AssertionError(f"missing vehicle field label: {label}")


def _workbook_visible_text(workbook: Any) -> str:
    values: list[str] = []
    for sheet_name in (VEHICLE_SHEET, PICTURE_SHEET, FIELD_SHEET):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            if sheet.row_dimensions[row[0].row].hidden:
                continue
            for cell in row:
                if cell.value not in (None, ""):
                    values.append(str(cell.value))
    return "\n".join(values)


def _sample_values_by_target() -> dict[str, Any]:
    return {
        "detail.baseCarInfo.name.displayValue": "丰田 凯美瑞 2.0G",
        "detail.baseCarInfo.name.brandName": "丰田",
        "detail.baseCarInfo.name.seriesName": "凯美瑞",
        "detail.baseCarInfo.name.modelName": "2.0G",
        "detail.baseCarInfo.carName": "凯美瑞",
        "detail.baseCarInfo.firstLicensePlateDate": "2020-06",
        "detail.baseCarInfo.mileage": "4.2万公里",
        "detail.baseCarInfo.vehicleCondition": "车况精品",
        "detail.baseCarInfo.exteriorColor": "白色",
        "detail.baseCarInfo.color": "白色",
        "detail.baseCarInfo.interiorColor": "黑色",
        "detail.baseCarInfo.innerColor": "黑色",
        "detail.baseCarInfo.stockStatus": "在库",
        "detail.baseCarInfo.area.cityName": "南京",
        "detail.baseCarInfo.area.provinceName": "江苏",
        "detail.baseCarInfo.area.displayValue": "江苏南京",
        "detail.baseCarInfo.registerArea.cityName": "苏州",
        "detail.baseCarInfo.registerArea.provinceName": "江苏",
        "detail.baseCarInfo.registerArea.displayValue": "江苏苏州",
        "detail.baseCarInfo.productionDate": "2019-12",
        "detail.carModelParam.gearbox": "自动",
        "detail.carModelParam.gearBox": "AT",
        "detail.carModelParam.displacement": "2.0L",
        "detail.carModelParam.gearBoxType": "自动",
        "detail.carModelParam.engineVolumeLiter": "2.0L",
        "detail.carModelParam.fuelType": "汽油",
        "detail.carModelParam.emissionStandard": "国六",
        "detail.carModelParam.seatNumber": 5,
        "detail.carLicenseInfo.keysCount": 2,
        "detail.carLicenseInfo.transferTotal": 1,
        "detail.carLicenseInfo.licenseStatus": "手续齐全",
        "detail.carPriceInfo.salePrice": 13.98,
        "detail.carPriceInfo.purchasePrice": 10.88,
        "detail.carPriceInfo.dealPrice": 13.50,
        "detail.carPriceInfo.salesPrice": 13.50,
        "detail.carPriceInfo.managerPrice": 13.30,
        "detail.carPriceInfo.wholesalePrice": 12.80,
        "detail.carPriceInfo.newPrice": 19.98,
        "detail.carPriceInfo.exhibitionPrice": 14.20,
        "detail.operationPhase": "SALE",
        "annotations.category": "used_car",
        "annotations.aliases": "凯美瑞；丰田凯美瑞",
        "annotations.specs": "一手车，保养记录完整",
        "annotations.shipping_policy": "到店看车请提前预约",
        "annotations.warranty_policy": "以门店质保政策为准",
        "annotations.risk_rules": "不得承诺事故水泡火烧以外未核验事项",
        "annotations.additional_details": "适合家用代步",
        "manual.sku": "CAMRY-001",
        "manual.unit": "台",
        "manual.inventory": 1,
        "manual.reply_templates.default": "这台车目前在售，建议先预约到店看车。",
        "manual.reply_templates.quote": "这台目前公开报价 13.98 万，具体成交以到店确认为准。",
        "manual.reply_templates.discount_policy": "议价需要看客户付款方式和置换情况，不能提前承诺底价。",
        "manual.reply_templates.logistics": "看车前建议提前预约，方便安排车辆和销售接待。",
        "manual.reply_templates.after_sales": "售后以门店合同和实际质保政策为准。",
        "manual.reply_templates.notes": "内部备注仅供客服配置，不进入客户证据。",
    }


def manual_vehicle_record(record_id: str, display_name: str) -> dict[str, Any]:
    return create_manual_vehicle(
        record_id=record_id,
        vehicle_detail_payload={
            "baseCarInfo": {
                "name": {
                    "displayValue": display_name,
                    "brandName": "",
                    "seriesName": "",
                    "modelName": "",
                }
            }
        },
        pictures_payload=[],
        observed_at="2026-08-01T00:00:00+00:00",
    )


class FakePostgresStore:
    def __init__(self, *, fail_atomic: bool = False) -> None:
        self.fail_atomic = fail_atomic
        self.items: dict[str, dict[str, Any]] = {}
        self.atomic_calls = 0

    def upsert_knowledge_items_atomic(
        self,
        tenant_id: str,
        layer: str,
        category_id: str,
        items: list[dict[str, Any]],
        *,
        product_id: str = "",
    ) -> None:
        self.atomic_calls += 1
        if self.fail_atomic:
            raise RuntimeError("injected PostgreSQL transaction failure")
        staged = {str(item.get("id") or ""): deepcopy(item) for item in items}
        self.items.update(staged)

    def list_knowledge_items(
        self,
        tenant_id: str,
        *,
        layer: str | None = None,
        category_id: str | None = None,
        product_id: str | None = None,
        include_archived: bool = False,
    ) -> list[dict[str, Any]]:
        values = list(self.items.values())
        if not include_archived:
            values = [item for item in values if str(item.get("status") or "active") != "archived"]
        return [deepcopy(item) for item in values]


def assert_equal(actual: Any, expected: Any, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: expected {expected!r}, got {actual!r}")


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _fill_rgb(cell: Any) -> str:
    rgb = str(getattr(getattr(cell, "fill", None), "fgColor", None).rgb or "")
    return rgb[-6:].upper()


if __name__ == "__main__":
    raise SystemExit(main())
