"""Focused checks for the VPS admin control plane and VPS-LOCAL protocol."""

from __future__ import annotations

import json
import os
import shutil
import sys
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from fastapi.testclient import TestClient
from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parents[1]
TEST_ROOT = PROJECT_ROOT / "runtime" / "apps" / "wechat_ai_customer_service" / "test_artifacts" / "vps_admin_control_plane"
TEST_RUN_ID = uuid.uuid4().hex[:8]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from apps.wechat_ai_customer_service.vps_admin.app import create_app  # noqa: E402
from apps.wechat_ai_customer_service.knowledge_paths import tenant_knowledge_base_root  # noqa: E402


_ADMIN_TOKEN_CACHE = ""


def test_email(username: str) -> str:
    return f"{username}-{TEST_RUN_ID}@example.local"


def main() -> int:
    cleanup_test_root()
    old_env = snapshot_env()
    try:
        os.environ["WECHAT_VPS_ADMIN_USERNAME"] = "admin"
        os.environ["WECHAT_VPS_ADMIN_PASSWORD"] = "1234.abcd"
        os.environ["WECHAT_VPS_ADMIN_EMAIL"] = "admin@example.local"
        os.environ["WECHAT_VPS_ADMIN_USER_ID"] = "platform-admin"
        os.environ["WECHAT_VPS_NODE_ENROLLMENT_TOKEN"] = "enroll-test"
        os.environ["WECHAT_SHARED_PATCH_SECRET"] = "shared-secret-test"
        os.environ["WECHAT_EMAIL_OTP_REQUIRED"] = "1"
        os.environ["WECHAT_EMAIL_OTP_DEBUG"] = "1"
        os.environ["WECHAT_EMAIL_OUTBOX_PATH"] = str(TEST_ROOT / "email_outbox.jsonl")
        client = TestClient(create_app(state_path=TEST_ROOT / "state.json"))
        checks: list[Callable[[TestClient], None]] = [
            check_admin_login_hidden_and_reserved,
            check_local_client_accounts_are_mirrored,
            check_tenant_customer_guest_flow,
            check_local_node_and_command_roundtrip,
            check_shared_knowledge_review_flow,
            check_restore_release_and_latest_update,
            check_customer_data_shared_sync_and_full_backup_entries,
            check_vps_console_chinese_shell,
        ]
        results = []
        for check in checks:
            try:
                check(client)
                results.append({"name": check.__name__, "ok": True})
            except Exception as exc:
                results.append({"name": check.__name__, "ok": False, "error": repr(exc)})
                break
    finally:
        restore_env(old_env)
    failures = [item for item in results if not item.get("ok")]
    print(json.dumps({"ok": not failures, "count": len(results), "failures": failures, "results": results}, ensure_ascii=False, indent=2))
    return 0 if not failures else 1


def check_admin_login_hidden_and_reserved(client: TestClient) -> None:
    token = admin_token(client)
    users = client.get("/v1/admin/users", headers=auth_headers(token))
    assert_status(users, 200, "list users")
    listed_users = users.json().get("users", [])
    assert_true(
        all(item.get("role") != "admin" and item.get("username") != "admin" for item in listed_users),
        "admin is not listed as stored user",
    )

    forbidden = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "shadow-admin", "password": "x", "role": "admin", "tenant_ids": ["tenant_a"]},
    )
    assert_equal(forbidden.status_code, 400, "stored admin is forbidden")


def check_local_client_accounts_are_mirrored(client: TestClient) -> None:
    token = admin_token(client)
    local_accounts_path = TEST_ROOT / "local_accounts.json"
    old_path = os.environ.get("WECHAT_LOCAL_ACCOUNTS_STATE_PATH")
    local_accounts_path.parent.mkdir(parents=True, exist_ok=True)
    local_accounts_path.write_text(
        json.dumps(
            {
                "accounts": {
                    "local_usedcar_customer": {
                        "user_id": "local_usedcar_customer",
                        "username": "local_usedcar_customer",
                        "display_name": "Local Usedcar Customer",
                        "role": "customer",
                        "tenant_ids": ["local_usedcar_tenant"],
                        "status": "active",
                    },
                    "local_trade_customer": {
                        "user_id": "local_trade_customer",
                        "username": "local_trade_customer",
                        "display_name": "Local Trade Customer",
                        "role": "customer",
                        "tenant_ids": ["local_trade_tenant"],
                        "status": "active",
                    },
                }
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    os.environ["WECHAT_LOCAL_ACCOUNTS_STATE_PATH"] = str(local_accounts_path)
    try:
        users = client.get("/v1/admin/users", headers=auth_headers(token))
        assert_status(users, 200, "list users with local mirrors")
        usernames = {item.get("username") for item in users.json().get("users", [])}
        assert_true({"local_usedcar_customer", "local_trade_customer"}.issubset(usernames), "VPS should show Local customer accounts")
        tenants = client.get("/v1/admin/tenants", headers=auth_headers(token))
        assert_status(tenants, 200, "list tenants with local mirrors")
        tenant_ids = {item.get("tenant_id") for item in tenants.json().get("tenants", [])}
        assert_true({"local_usedcar_tenant", "local_trade_tenant"}.issubset(tenant_ids), "VPS should mirror Local customer data spaces")
    finally:
        if old_path is None:
            os.environ.pop("WECHAT_LOCAL_ACCOUNTS_STATE_PATH", None)
        else:
            os.environ["WECHAT_LOCAL_ACCOUNTS_STATE_PATH"] = old_path


def check_tenant_customer_guest_flow(client: TestClient) -> None:
    token = admin_token(client)
    auto_customer = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "customer_auto", "password": "customer-pass", "role": "customer"},
    )
    assert_status(auto_customer, 200, "create customer with account-scoped data")
    assert_equal(auto_customer.json()["user"]["tenant_ids"], ["customer_auto"], "customer owns same-name data scope")
    assert_equal(auto_customer.json()["user"]["authorized_customers"], ["customer_auto"], "customer readable access label")

    blocked_guest = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "guest_missing", "password": "guest-pass", "role": "guest", "authorized_customer": "missing_customer"},
    )
    assert_equal(blocked_guest.status_code, 404, "guest must be assigned to an existing customer")

    scoped_guest = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "guest_auto", "password": "guest-pass", "role": "guest", "authorized_customer": "customer_auto"},
    )
    assert_status(scoped_guest, 200, "create guest for customer")
    assert_equal(scoped_guest.json()["user"]["tenant_ids"], ["customer_auto"], "guest inherits authorized customer data scope")
    assert_equal(scoped_guest.json()["user"]["authorized_customers"], ["customer_auto"], "guest readable access label")

    customer_auto_token = account_token(
        client,
        username="customer_auto",
        password="customer-pass",
        tenant_id="customer_auto",
        email=test_email("customer_auto"),
        new_password="customer-auto.5678",
    )
    assert_true(bool(customer_auto_token), "same-name customer login")
    guest_auto_token = account_token(
        client,
        username="guest_auto",
        password="guest-pass",
        tenant_id="customer_auto",
        email=test_email("guest_auto"),
        new_password="guest-auto.5678",
    )
    assert_true(bool(guest_auto_token), "admin-assigned guest login")

    orphan_tenant = client.post("/v1/admin/tenants", headers=auth_headers(token), json={"tenant_id": "orphan_scope", "display_name": "Orphan Scope"})
    assert_status(orphan_tenant, 200, "create orphan tenant scope")
    blocked_raw_guest = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "guest_raw_scope", "password": "guest-pass", "role": "guest", "tenant_ids": ["orphan_scope"]},
    )
    assert_equal(blocked_raw_guest.status_code, 400, "guest raw tenant scope must map to an existing customer")

    tenant = client.post("/v1/admin/tenants", headers=auth_headers(token), json={"tenant_id": "tenant_a", "display_name": "Tenant A"})
    assert_status(tenant, 200, "create tenant")

    customer = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "customer_a", "password": "customer-pass", "role": "customer", "tenant_ids": ["tenant_a"]},
    )
    assert_status(customer, 200, "create customer")

    guest = client.post(
        "/v1/admin/users",
        headers=auth_headers(token),
        json={"username": "guest_a", "password": "guest-pass", "role": "guest", "tenant_ids": ["tenant_a"]},
    )
    assert_status(guest, 200, "create guest")

    customer_token = account_token(
        client,
        username="customer_a",
        password="customer-pass",
        tenant_id="tenant_a",
        email=test_email("customer_a"),
        new_password="customer-a.5678",
    )
    customer_me = client.get("/v1/auth/me", headers=auth_headers(customer_token))
    assert_status(customer_me, 200, "customer login")
    assert_equal(customer_me.json()["session"]["user"]["role"], "customer", "customer role")

    guest_token = account_token(
        client,
        username="guest_a",
        password="guest-pass",
        tenant_id="tenant_a",
        email=test_email("guest_a"),
        new_password="guest-a.5678",
    )
    guest_me = client.get("/v1/auth/me", headers=auth_headers(guest_token))
    assert_status(guest_me, 200, "guest login")
    assert_equal(guest_me.json()["session"]["user"]["role"], "guest", "guest role")


def check_local_node_and_command_roundtrip(client: TestClient) -> None:
    token = admin_token(client)
    registered = client.post(
        "/v1/local/nodes/register",
        headers={"X-Enrollment-Token": "enroll-test"},
        json={
            "node_id": "local_tenant_a_01",
            "display_name": "Local Tenant A 01",
            "tenant_ids": ["tenant_a"],
            "version": "0.1.0-test",
            "capabilities": ["backup_tenant", "backup_all", "check_update"],
        },
    )
    assert_status(registered, 200, "register node")
    node_token = registered.json()["node"]["node_token"]

    heartbeat = client.post(
        "/v1/local/nodes/local_tenant_a_01/heartbeat",
        headers={"X-Node-Token": node_token},
        json={"status": "online", "version": "0.1.1-test", "metrics": {"pending_jobs": 0}},
    )
    assert_status(heartbeat, 200, "node heartbeat")

    backup = client.post(
        "/v1/admin/backups",
        headers=auth_headers(token),
        json={"scope": "tenant", "tenant_id": "tenant_a", "node_id": "local_tenant_a_01"},
    )
    assert_status(backup, 200, "request backup")
    command_id = backup.json()["command"]["command_id"]

    poll = client.get(
        "/v1/local/commands?tenant_id=tenant_a&node_id=local_tenant_a_01",
        headers={"X-Node-Token": node_token},
    )
    assert_status(poll, 200, "poll commands")
    commands = poll.json()["commands"]
    assert_equal(len(commands), 1, "one pending command")
    assert_equal(commands[0]["command_id"], command_id, "command id")
    assert_equal(commands[0]["type"], "backup_tenant", "command type")

    result = client.post(
        f"/v1/local/commands/{command_id}/result",
        headers={"X-Node-Token": node_token},
        json={"command_id": command_id, "accepted": True, "result": {"ok": True, "backup_id": "backup_test"}},
    )
    assert_status(result, 200, "submit command result")

    commands_after = client.get("/v1/admin/commands", headers=auth_headers(token)).json()["commands"]
    matched = next(item for item in commands_after if item["command_id"] == command_id)
    assert_equal(matched["status"], "succeeded", "command succeeded")


def check_shared_knowledge_review_flow(client: TestClient) -> None:
    token = admin_token(client)
    patch_node = client.post(
        "/v1/local/nodes/register",
        headers={"X-Enrollment-Token": "enroll-test"},
        json={
            "node_id": "local_tenant_a_patch_01",
            "display_name": "Local Tenant A Patch 01",
            "tenant_ids": ["tenant_a"],
            "version": "0.1.0-test",
            "capabilities": ["pull_shared_patch"],
        },
    )
    assert_status(patch_node, 200, "register shared patch node")
    patch_node_token = patch_node.json()["node"]["node_token"]

    manual = client.post(
        "/v1/admin/shared/library",
        headers=auth_headers(token),
        json={
            "item_id": "manual_console_rule",
            "category_id": "global_guidelines",
            "title": "Manual Rule",
            "content": "Manual shared content",
            "keywords": ["manual", "rule"],
            "applies_to": "Shared rule review",
            "notes": "Created through human form fields",
        },
    )
    assert_status(manual, 200, "create shared library item")
    assert_equal(manual.json()["item"]["keywords"], ["manual", "rule"], "shared item keywords are stored")
    assert_equal(manual.json()["item"]["applies_to"], "Shared rule review", "shared item applies_to stored")
    updated = client.patch(
        "/v1/admin/shared/library/manual_console_rule",
        headers=auth_headers(token),
        json={"title": "Manual Rule Updated", "content": "Updated shared content", "status": "active", "keywords": ["updated"]},
    )
    assert_status(updated, 200, "update shared library item")
    assert_equal(updated.json()["item"]["title"], "Manual Rule Updated", "shared library item updated")
    assert_equal(updated.json()["item"]["keywords"], ["updated"], "shared item keywords updated")
    fetched = client.get("/v1/admin/shared/library/manual_console_rule", headers=auth_headers(token))
    assert_status(fetched, 200, "get shared library item")
    assert_true(not str(fetched.json()["item"]["content"]).lstrip().startswith("{"), "shared detail is human-readable")
    deleted = client.delete("/v1/admin/shared/library/manual_console_rule", headers=auth_headers(token))
    assert_status(deleted, 200, "delete shared library item")

    probe_dir = tenant_knowledge_base_root("default") / "policies" / "items"
    universal_probe = probe_dir / "vps_shared_universal_probe.json"
    private_probe = probe_dir / "vps_shared_private_probe.json"
    probe_dir.mkdir(parents=True, exist_ok=True)
    universal_probe.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "id": "vps_shared_universal_probe",
                "category_id": "policies",
                "status": "active",
                "data": {
                    "title": "通用人工转接说明",
                    "answer": "当客户明确要求人工服务时，应说明已经转接人工客服并请稍等片刻。",
                    "keywords": ["人工客服", "转接人工"],
                    "applicability_scope": "global",
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    private_probe.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "id": "vps_shared_private_probe",
                "category_id": "policies",
                "status": "active",
                "data": {
                    "title": "江苏车金客户跟进规则",
                    "answer": "江苏车金南京门店客户张三手机号13800138000，咨询二手车试驾和过户，必须由门店销售跟进。",
                    "keywords": ["江苏车金", "南京门店", "二手车"],
                    "applicability_scope": "global",
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    try:
        generated = client.post(
            "/v1/admin/shared/proposals/generate-from-formal",
            headers=auth_headers(token),
            json={"tenant_id": "default", "use_llm": False, "limit": 50},
        )
        assert_status(generated, 200, "generate shared candidates from formal knowledge")
        created = generated.json().get("created", [])
        assert_true(created, "formal knowledge extraction should create candidate shared proposals")
        assert_true(
            all("江苏车金" not in json.dumps(item, ensure_ascii=False) for item in created),
            "customer-private formal knowledge must not become shared candidates",
        )
        generated_proposal = created[0]
        assert_equal(generated_proposal.get("status"), "pending_review", "generated shared proposal must wait for admin review")
        assert_true(
            str(generated_proposal.get("source") or "").startswith("formal_knowledge_universal"),
            "generated shared proposal source should identify formal-knowledge extraction",
        )
        assert_true(
            generated_proposal.get("source_meta", {}).get("llm_used") is False,
            "use_llm false should record that LLM was not used",
        )
        assert_true(
            generated_proposal.get("review_assist", {}).get("recommendation") in {"accept", "reject", "revise"},
            "generated shared proposal should include admin review assist",
        )
        rejected_generated = client.post(
            f"/v1/admin/shared/proposals/{generated_proposal['proposal_id']}/review",
            headers=auth_headers(token),
            json={"action": "reject", "note": "test cleanup"},
        )
        assert_status(rejected_generated, 200, "reject generated shared candidate")
        assert_equal(rejected_generated.json()["proposal"]["status"], "rejected", "generated candidate can be rejected by admin")
        generated_again = client.post(
            "/v1/admin/shared/proposals/generate-from-formal",
            headers=auth_headers(token),
            json={"tenant_id": "default", "use_llm": False, "limit": 50, "only_unscanned": True},
        )
        assert_status(generated_again, 200, "generate shared candidates only unscanned")
        assert_equal(generated_again.json().get("created", []), [], "already checked formal knowledge should not generate duplicates")
        assert_true(generated_again.json().get("scan", {}).get("scan_state_count", 0) >= 1, "formal scan state should be persisted")
    finally:
        for path in (universal_probe, private_probe):
            if path.exists():
                path.unlink()

    proposal = client.post(
        "/v1/shared/proposals",
        json={
            "tenant_id": "tenant_a",
            "title": "Shared after-sale policy",
            "summary": "Common rule extracted from tenant A",
            "operations": [
                {
                    "op": "upsert_json",
                    "path": "global_guidelines/items/after_sale_policy.json",
                    "content": {"schema_version": 1, "id": "after_sale_policy", "data": {"title": "After Sale"}},
                }
            ],
        },
    )
    assert_status(proposal, 200, "submit shared proposal")
    proposal_id = proposal.json()["proposal"]["proposal_id"]
    refreshed_assist = client.post(
        f"/v1/admin/shared/proposals/{proposal_id}/review-assist",
        headers=auth_headers(token),
        json={"use_llm": False},
    )
    assert_status(refreshed_assist, 200, "refresh proposal review assist")
    assert_true(
        refreshed_assist.json()["review_assist"].get("recommendation") in {"accept", "reject", "revise"},
        "proposal review assist has a recommendation",
    )

    reviewed = client.post(
        f"/v1/admin/shared/proposals/{proposal_id}/review",
        headers=auth_headers(token),
        json={"action": "accept", "version": "shared-test.1"},
    )
    assert_status(reviewed, 200, "accept proposal")
    assert_equal(reviewed.json()["proposal"]["status"], "accepted", "proposal accepted")
    assert_equal(reviewed.json()["patch"]["status"], "published", "patch published")
    assert_true(bool(reviewed.json()["patch"].get("signature")), "published patch is signed when secret is configured")
    assert_true(reviewed.json()["library_items"], "accepted proposal writes official library")
    patch_id = reviewed.json()["patch"]["patch_id"]
    patches_before_push = client.get("/v1/admin/shared/patches", headers=auth_headers(token))
    assert_status(patches_before_push, 200, "list shared patches before push")
    listed_patch = next(item for item in patches_before_push.json()["patches"] if item["patch_id"] == patch_id)
    assert_equal(listed_patch["delivery"]["overall_status"], "not_pushed", "new patch starts without delivery")

    pushed = client.post(
        f"/v1/admin/shared/patches/{patch_id}/push",
        headers=auth_headers(token),
        json={"tenant_id": "tenant_a", "node_id": "local_tenant_a_patch_01"},
    )
    assert_status(pushed, 200, "push shared patch to tenant nodes")
    assert_true(pushed.json()["commands"], "patch push creates commands")
    assert_equal(pushed.json()["commands"][0]["type"], "pull_shared_patch", "patch push command type")
    assert_equal(pushed.json()["commands"][0]["payload"]["patch"]["patch_id"], patch_id, "patch command carries full patch")
    assert_equal(pushed.json()["delivery"]["counts"]["total"], 1, "patch delivery tracks one target")
    assert_equal(pushed.json()["delivery"]["overall_status"], "pending", "patch push starts pending delivery")
    command_id = pushed.json()["commands"][0]["command_id"]

    patch_poll = client.get(
        "/v1/local/commands?tenant_id=tenant_a&node_id=local_tenant_a_patch_01",
        headers={"X-Node-Token": patch_node_token},
    )
    assert_status(patch_poll, 200, "patch node polls shared patch command")
    assert_equal(patch_poll.json()["commands"][0]["command_id"], command_id, "patch command delivered to target node")
    patches_after_poll = client.get("/v1/admin/shared/patches", headers=auth_headers(token))
    assert_status(patches_after_poll, 200, "list shared patches after poll")
    polled_patch = next(item for item in patches_after_poll.json()["patches"] if item["patch_id"] == patch_id)
    assert_equal(polled_patch["delivery"]["counts"]["sent"], 1, "patch delivery shows sent before apply result")

    patch_result = client.post(
        f"/v1/local/commands/{command_id}/result",
        headers={"X-Node-Token": patch_node_token},
        json={"command_id": command_id, "accepted": True, "result": {"ok": True, "applied": {"ok": True}}},
    )
    assert_status(patch_result, 200, "patch node reports apply result")
    patches_after_apply = client.get("/v1/admin/shared/patches", headers=auth_headers(token))
    assert_status(patches_after_apply, 200, "list shared patches after apply")
    applied_patch = next(item for item in patches_after_apply.json()["patches"] if item["patch_id"] == patch_id)
    assert_equal(applied_patch["delivery"]["overall_status"], "applied", "patch delivery becomes applied")
    assert_equal(applied_patch["delivery"]["targets"][0]["node_id"], "local_tenant_a_patch_01", "patch delivery includes target node")

    library = client.get("/v1/admin/shared/library/after_sale_policy", headers=auth_headers(token))
    assert_status(library, 200, "accepted proposal library item")
    assert_equal(library.json()["item"]["title"], "After Sale", "accepted proposal title")
    cloud_snapshot = client.get("/v1/shared/knowledge?tenant_id=tenant_a", headers=auth_headers(token))
    assert_status(cloud_snapshot, 200, "cloud official shared snapshot")
    snapshot = cloud_snapshot.json()["snapshot"]
    assert_true(any(item.get("item_id") == "after_sale_policy" for item in snapshot.get("items", [])), "cloud snapshot includes accepted library item")
    not_modified = client.get(f"/v1/shared/knowledge?tenant_id=tenant_a&since_version={snapshot['version']}", headers=auth_headers(token))
    assert_status(not_modified, 200, "cloud official shared snapshot not modified")
    assert_equal(not_modified.json()["snapshot"].get("not_modified"), True, "cloud snapshot supports version short circuit")

    unsafe = client.post(
        "/v1/shared/proposals",
        json={
            "tenant_id": "tenant_a",
            "title": "Unsafe shared proposal",
            "summary": "Should be rejected at review time",
            "operations": [
                {
                    "op": "upsert_json",
                    "path": "../escape.json",
                    "content": {"schema_version": 1, "id": "escape", "data": {"title": "bad"}},
                }
            ],
        },
    )
    assert_status(unsafe, 200, "submit unsafe proposal")
    unsafe_review = client.post(
        f"/v1/admin/shared/proposals/{unsafe.json()['proposal']['proposal_id']}/review",
        headers=auth_headers(token),
        json={"action": "accept"},
    )
    assert_equal(unsafe_review.status_code, 400, "unsafe proposal cannot be accepted")


def check_restore_release_and_latest_update(client: TestClient) -> None:
    token = admin_token(client)
    restore = client.post(
        "/v1/admin/restores",
        headers=auth_headers(token),
        json={"tenant_id": "tenant_a", "node_id": "local_tenant_a_01", "backup_id": "backup_test", "dry_run": True},
    )
    assert_status(restore, 200, "request restore dry-run")
    assert_equal(restore.json()["command"]["type"], "restore_backup", "restore command type")

    release = client.post(
        "/v1/admin/releases",
        headers=auth_headers(token),
        json={
            "version": "0.2.0-test",
            "channel": "stable",
            "title": "Test release",
            "artifact_url": "https://example.invalid/release.zip",
            "sha256": "0" * 64,
            "signature": "sig-test",
            "notes": "Test release notes",
        },
    )
    assert_status(release, 200, "create release")
    assert_equal(release.json()["release"]["sha256"], "0" * 64, "release sha256 stored")
    release_id = release.json()["release"]["release_id"]
    pushed_release = client.post(
        f"/v1/admin/releases/{release_id}/push",
        headers=auth_headers(token),
        json={"tenant_id": "tenant_a", "mode": "check_update"},
    )
    assert_status(pushed_release, 200, "push release check command")
    assert_true(pushed_release.json()["commands"], "release push creates commands")
    assert_equal(pushed_release.json()["commands"][0]["type"], "check_update", "release push command type")

    latest = client.get("/v1/updates/latest?channel=stable")
    assert_status(latest, 200, "latest update")
    assert_equal(latest.json()["update"]["version"], "0.2.0-test", "latest version")
    assert_equal(latest.json()["update"]["signature"], "sig-test", "latest signature")


def check_customer_data_shared_sync_and_full_backup_entries(client: TestClient) -> None:
    token = admin_token(client)
    bootstrapped = client.post(
        "/v1/admin/customer-data/bootstrap-test01",
        headers=auth_headers(token),
        json={"tenant_id": "default"},
    )
    assert_status(bootstrapped, 200, "bootstrap test01")
    assert_equal(bootstrapped.json()["user"]["username"], "test01", "test01 username")
    assert_true(bootstrapped.json()["package"]["summary"]["formal_knowledge"]["item_count"] >= 1, "test01 package has formal knowledge")
    test01_package_id = bootstrapped.json()["package"]["package_id"]
    test01_readable = client.get(f"/v1/admin/customer-data/{test01_package_id}/readable-download", headers=auth_headers(token))
    assert_status(test01_readable, 200, "test01 readable Excel download")
    test01_workbook = load_workbook(BytesIO(test01_readable.content), read_only=True)
    assert_true("正式-商品资料" in test01_workbook.sheetnames, "test01 readable export splits product formal knowledge")
    assert_true("正式-政策规则" in test01_workbook.sheetnames, "test01 readable export splits policy formal knowledge")
    assert_true("原始聊天" in test01_workbook.sheetnames, "readable export includes raw message sheet")
    assert_true("待确认知识" in test01_workbook.sheetnames, "readable export includes pending candidate sheet")
    assert_true("导入资料" in test01_workbook.sheetnames, "readable export includes upload sheet")
    assert_true(test01_workbook["正式-政策规则"].max_row >= 2, "test01 policy knowledge sheet includes rows")

    customer_data = client.get("/v1/admin/customer-data", headers=auth_headers(token))
    assert_status(customer_data, 200, "list customer data")
    assert_true(any(item.get("account_username") == "test01" for item in customer_data.json()["packages"]), "test01 package listed")

    selected_package = client.post(
        "/v1/admin/customer-data/package-customer",
        headers=auth_headers(token),
        json={"account_username": "customer_a", "tenant_id": "tenant_a"},
    )
    assert_status(selected_package, 200, "package selected customer")
    package_id = selected_package.json()["package"]["package_id"]
    detail = client.get(f"/v1/admin/customer-data/{package_id}", headers=auth_headers(token))
    assert_status(detail, 200, "customer package detail")
    assert_equal(detail.json()["package"]["account_username"], "customer_a", "detail account username")
    download = client.get(f"/v1/admin/customer-data/{package_id}/download", headers=auth_headers(token))
    assert_status(download, 200, "customer package download")
    assert_true(bool(download.content), "download returns package bytes")
    readable = client.get(f"/v1/admin/customer-data/{package_id}/readable-download", headers=auth_headers(token))
    assert_status(readable, 200, "customer readable Excel download")
    workbook = load_workbook(BytesIO(readable.content), read_only=True)
    assert_true(any(name.startswith("正式") for name in workbook.sheetnames), "readable export has formal category sheets")
    assert_true(any(name.startswith("商品专属") for name in workbook.sheetnames), "readable export has product knowledge sheets")
    assert_true({"原始聊天", "待确认知识", "导入资料"}.issubset(set(workbook.sheetnames)), "readable export matches latest client data structure")
    deleted = client.delete(f"/v1/admin/customer-data/{package_id}", headers=auth_headers(token))
    assert_status(deleted, 200, "customer package delete")
    assert_equal(deleted.json()["package"]["package_id"], package_id, "deleted package id")

    shared_overview = client.get("/v1/admin/shared/overview", headers=auth_headers(token))
    assert_status(shared_overview, 200, "shared overview")
    assert_equal(shared_overview.json()["official"]["source"], "cloud_official_shared_library", "shared overview uses cloud official source")
    assert_true(shared_overview.json()["official"]["items"], "cloud official shared items are visible")
    assert_true(shared_overview.json()["local_legacy"]["structure"]["separated"] is True, "legacy local shared structure remains inspectable")

    shared_sync = client.post("/v1/admin/shared/sync-local", headers=auth_headers(token), json={})
    assert_status(shared_sync, 200, "sync shared snapshot")
    assert_true(shared_sync.json()["snapshot"]["summary"]["item_count"] >= 1, "shared snapshot item count")

    local_backup = client.post("/v1/admin/backups/local-now", headers=auth_headers(token), json={"scope": "all", "tenant_id": "default"})
    assert_status(local_backup, 200, "local full backup")
    assert_equal(local_backup.json()["request"]["status"], "succeeded", "local backup succeeded")
    backup_request_id = local_backup.json()["request"]["request_id"]

    restore_latest = client.post("/v1/admin/restores/latest", headers=auth_headers(token), json={"scope": "all", "tenant_id": "default", "dry_run": True})
    assert_status(restore_latest, 200, "restore latest dry-run")
    assert_equal(restore_latest.json()["request"]["dry_run"], True, "restore is dry-run")
    assert_equal(restore_latest.json()["command"]["type"], "restore_backup", "latest restore command type")

    deleted_backup = client.delete(f"/v1/admin/backups/{backup_request_id}", headers=auth_headers(token))
    assert_status(deleted_backup, 200, "delete backup record")
    backup_list = client.get("/v1/admin/backups", headers=auth_headers(token))
    assert_status(backup_list, 200, "list backups after delete")
    assert_true(
        all(item.get("request_id") != backup_request_id for item in backup_list.json()["items"]),
        "deleted backup is removed from admin list",
    )


def check_vps_console_chinese_shell(client: TestClient) -> None:
    response = client.get("/")
    assert_status(response, 200, "console index")
    html = response.text
    js = client.get("/static/app.js")
    assert_status(js, 200, "console app js")
    assert_true('id="login-screen"' in html, "VPS login uses a separate screen")
    assert_true('id="logout-button"' in html, "VPS console has logout entry")
    assert_true("登录服务端" in html, "console login title is Chinese")
    assert_true("访客可查看客户" in html, "guest access selector is user-readable")
    assert_true("打包所选客户数据" in html, "selected customer package button visible")
    assert_true("客户电脑连接" in html, "local node wording is user-readable")
    assert_true("共享公共知识" in html, "shared knowledge navigation visible")
    assert_true("候选共享公共知识库" not in html, "shared candidate review is merged into shared knowledge")
    assert_true("AI 提炼候选共享知识" in html, "shared candidate generation action is visible")
    assert_true("AI复核建议" in js.text and "review-assist" in js.text, "shared candidate review assist action is visible")
    assert_true("pendingProposals" in js.text and "archivedProposals" in js.text, "candidate list only shows pending items with archive access")
    assert_true('id="sync-shared"' in html and 'id="shared-patch-list"' in html, "shared snapshot and patch list controls are visible")
    assert_true("slice(0, 10)" in js.text and "patchDeliveryHtml" in js.text, "published shared patches are capped and show delivery status")
    assert_true("sharedPatchPushMessage" in js.text, "shared patch push action reports target delivery feedback")
    assert_true("已拒绝 / 已作废归档" in js.text, "voided or rejected candidates are archived behind a collapsed entry")
    assert_true("标准操作方法" in html, "release update operation method visible")
    assert_true("客户租户" not in html, "tenant jargon is hidden from console UI")
    assert_true("生成 test01 测试客户" not in html, "test01 bootstrap is hidden from console UI")
    return
    assert_true("VPS 管理控制台" in html, "console title is Chinese")
    assert_true("共享公共知识" in html, "shared knowledge navigation visible")
    assert_true("打包所选客户数据" in html, "selected customer package button visible")
    assert_true("客户电脑连接" in html, "local node wording is user-readable")
    assert_true("标准操作方法" in html, "release update operation method visible")
    assert_true("生成 test01 测试客户" not in html, "test01 bootstrap is hidden from console UI")


def admin_token(client: TestClient) -> str:
    global _ADMIN_TOKEN_CACHE
    if _ADMIN_TOKEN_CACHE:
        me = client.get("/v1/auth/me", headers=auth_headers(_ADMIN_TOKEN_CACHE))
        if me.status_code == 200:
            return _ADMIN_TOKEN_CACHE
        _ADMIN_TOKEN_CACHE = ""

    for password in ("admin.5678", "1234.abcd"):
        response = client.post(
            "/v1/auth/login/start",
            json={"username": "admin", "password": password, "tenant_id": "default", "device_id": "vps-test-admin"},
        )
        if response.status_code == 401:
            continue
        assert_status(response, 200, "admin login start")
        body = response.json()
        if body.get("requires_initialization"):
            initialize_vps_account(
                client,
                body,
                email="admin@example.local",
                new_password="admin.5678",
                smtp_config={"otp_required": True, "from_email": "admin@example.local"},
            )
            return admin_token(client)
        token = token_from_started_login(client, body, trust_device=True)
        _ADMIN_TOKEN_CACHE = token
        return token
    raise AssertionError("admin login could not start")


def account_token(
    client: TestClient,
    *,
    username: str,
    password: str,
    tenant_id: str,
    email: str,
    new_password: str,
) -> str:
    for candidate_password in (new_password, password):
        response = client.post(
            "/v1/auth/login/start",
            json={"username": username, "password": candidate_password, "tenant_id": tenant_id, "device_id": f"vps-test-{username}"},
        )
        if response.status_code == 401:
            continue
        assert_status(response, 200, f"{username} login start")
        body = response.json()
        if body.get("requires_initialization"):
            initialize_vps_account(client, body, email=email, new_password=new_password)
            return account_token(
                client,
                username=username,
                password=password,
                tenant_id=tenant_id,
                email=email,
                new_password=new_password,
            )
        return token_from_started_login(client, body)
    raise AssertionError(f"{username} login could not start")


def initialize_vps_account(
    client: TestClient,
    body: dict[str, Any],
    *,
    email: str,
    new_password: str,
    smtp_config: dict[str, Any] | None = None,
) -> None:
    started = client.post(
        "/v1/auth/initialize/start",
        json={
            "challenge_id": body["challenge_id"],
            "email": email,
            "new_password": new_password,
            "smtp_config": smtp_config or {},
        },
    )
    assert_status(started, 200, "start account initialization")
    payload = started.json()
    assert_true(bool(payload.get("debug_code")), "initialization debug code exists")
    verified = client.post(
        "/v1/auth/initialize/verify",
        json={"challenge_id": payload["challenge_id"], "code": payload["debug_code"]},
    )
    assert_status(verified, 200, "verify account initialization")


def token_from_started_login(client: TestClient, body: dict[str, Any], *, trust_device: bool = False) -> str:
    if body.get("session"):
        token = body.get("session", {}).get("token")
        assert_true(bool(token), "session token exists")
        return str(token)
    assert_true(bool(body.get("requires_verification")), "email verification required")
    assert_true(bool(body.get("debug_code")), "login debug code exists")
    verified = client.post(
        "/v1/auth/login/verify",
        json={"challenge_id": body["challenge_id"], "code": body["debug_code"], "trust_device": trust_device},
    )
    assert_status(verified, 200, "verify login")
    token = verified.json().get("session", {}).get("token")
    assert_true(bool(token), "verified token exists")
    return str(token)


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


ENV_KEYS = (
    "WECHAT_VPS_ADMIN_USERNAME",
    "WECHAT_VPS_ADMIN_PASSWORD",
    "WECHAT_VPS_ADMIN_EMAIL",
    "WECHAT_VPS_ADMIN_USER_ID",
    "WECHAT_VPS_NODE_ENROLLMENT_TOKEN",
    "WECHAT_SHARED_PATCH_SECRET",
    "WECHAT_VPS_ADMIN_STATE_PATH",
    "WECHAT_EMAIL_OTP_REQUIRED",
    "WECHAT_EMAIL_OTP_DEBUG",
    "WECHAT_EMAIL_OUTBOX_PATH",
)


def snapshot_env() -> dict[str, str | None]:
    return {key: os.environ.get(key) for key in ENV_KEYS}


def restore_env(values: dict[str, str | None]) -> None:
    for key, value in values.items():
        if value is None:
            os.environ.pop(key, None)
        else:
            os.environ[key] = value


def cleanup_test_root() -> None:
    resolved = TEST_ROOT.resolve()
    expected_parent = (PROJECT_ROOT / "runtime" / "apps" / "wechat_ai_customer_service" / "test_artifacts").resolve()
    if expected_parent not in resolved.parents and resolved != expected_parent:
        raise RuntimeError(f"unsafe test cleanup path: {resolved}")
    if resolved.exists():
        shutil.rmtree(resolved)
    resolved.mkdir(parents=True, exist_ok=True)


def assert_status(response: Any, expected: int, message: str) -> None:
    if response.status_code != expected:
        raise AssertionError(f"{message}: expected status {expected}, got {response.status_code}, body={response.text}")


def assert_equal(actual: Any, expected: Any, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: expected {expected!r}, got {actual!r}")


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


if __name__ == "__main__":
    raise SystemExit(main())
