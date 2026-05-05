"""Jiangsu Chejin used-car customer-service and recorder checks.

The checks create an isolated customer tenant, load used-car test materials,
verify upload learning/RAG/candidates, exercise the guarded customer-service
reply workflow, and optionally run a live WeChat smoke through File Transfer
Assistant plus the selected test group.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

os.environ.setdefault("WECHAT_STORAGE_BACKEND", "file")
os.environ["WECHAT_VPS_BASE_URL"] = ""
os.environ["WECHAT_VPS_AUTH_REQUIRED"] = "0"
os.environ["WECHAT_VPS_AUTO_DISCOVER"] = "0"

from fastapi.testclient import TestClient
from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_ROOT.parents[1]
WORKFLOWS_ROOT = APP_ROOT / "workflows"
ADAPTERS_ROOT = APP_ROOT / "adapters"
for path in (PROJECT_ROOT, WORKFLOWS_ROOT, APP_ROOT, ADAPTERS_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from apps.wechat_ai_customer_service.admin_backend.api.tenants import initialize_tenant_knowledge_base  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.app import create_app  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.formal_review_state import mark_item_new  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.knowledge_base_store import KnowledgeBaseStore  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.raw_message_learning_service import RawMessageLearningService  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.raw_message_store import RawMessageStore  # noqa: E402
from apps.wechat_ai_customer_service.admin_backend.services.recorder_service import RecorderService  # noqa: E402
from apps.wechat_ai_customer_service.auth.passwords import hash_password  # noqa: E402
from apps.wechat_ai_customer_service.auth.session import AuthService  # noqa: E402
from apps.wechat_ai_customer_service.knowledge_paths import (  # noqa: E402
    tenant_context,
    tenant_metadata_path,
    tenant_rag_index_root,
    tenant_review_candidates_root,
    tenant_root,
    tenant_runtime_root,
)
from apps.wechat_ai_customer_service.workflows.rag_layer import RagService  # noqa: E402
from listen_and_reply import TargetConfig, configured_reply_prefix, process_target  # noqa: E402
from wechat_connector import FILE_TRANSFER_ASSISTANT, WeChatConnector  # noqa: E402


TENANT_ID = "jiangsu_chejin_usedcar_customer_20260501"
PASSWORD = "chejin.20260501"
EMAIL = "jiangsu-chejin-usedcar@example.local"
DISPLAY_NAME = "江苏车金二手车测试客户 2026-05-01"
BASE_ARTIFACT_ROOT = PROJECT_ROOT / "runtime" / "apps" / "wechat_ai_customer_service" / "test_artifacts" / "jiangsu_chejin_used_car"


@dataclass
class FakeConnector:
    messages: list[dict[str, Any]]

    def __post_init__(self) -> None:
        self.sent_texts: list[str] = []

    def get_messages(self, target: str, exact: bool = True) -> dict[str, Any]:
        return {"ok": True, "target": target, "exact": exact, "messages": self.messages}

    def send_text_and_verify(self, target: str, text: str, exact: bool = True) -> dict[str, Any]:
        self.sent_texts.append(text)
        return {"ok": True, "verified": True, "target": target, "exact": exact, "text": text}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--live-wechat", action="store_true", help="Also send live WeChat messages and verify capture/reply.")
    parser.add_argument("--group-name", default="偷数据测试")
    args = parser.parse_args()

    token = "CHEJIN_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    results: list[dict[str, Any]] = []
    with tenant_context(TENANT_ID):
        ensure_customer_account()
        client, headers = login_client()
        material_result = check_material_upload_learning(client, headers, token)
        results.append(material_result)
        seed_result = seed_formal_used_car_knowledge(token)
        results.append(seed_result)
        results.append(check_customer_service_matrix(token))
        results.append(check_recorder_offline_matrix(token))
        results.append(check_tenant_scoped_backend(client, headers, token))
        if args.live_wechat:
            results.append(check_live_wechat_matrix(token, group_name=args.group_name))

    failures = [item for item in results if not item.get("ok")]
    payload = {
        "ok": not failures,
        "tenant_id": TENANT_ID,
        "username": TENANT_ID,
        "password": PASSWORD,
        "batch_token": token,
        "failures": failures,
        "results": results,
        "artifact_root": str(BASE_ARTIFACT_ROOT),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if payload["ok"] else 1


def ensure_customer_account() -> None:
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    root = tenant_root(TENANT_ID)
    for child in ("knowledge_bases", "product_item_knowledge", "rag_sources", "rag_experience", "review_candidates"):
        (root / child).mkdir(parents=True, exist_ok=True)
    initialize_tenant_knowledge_base(TENANT_ID)
    metadata_path = tenant_metadata_path(TENANT_ID)
    metadata = {
        "schema_version": 1,
        "tenant_id": TENANT_ID,
        "display_name": DISPLAY_NAME,
        "knowledge_base_root": "knowledge_bases",
        "product_item_knowledge_root": "product_item_knowledge",
        "created_at": now,
        "sync": {"private_backup": {"enabled": False, "schedule": "manual"}},
    }
    if metadata_path.exists():
        existing = json.loads(metadata_path.read_text(encoding="utf-8"))
        metadata = {**existing, **metadata, "updated_at": now}
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    service = AuthService()
    accounts = service.read_local_account_overrides()
    existing = accounts.get(TENANT_ID, {})
    accounts[TENANT_ID] = {
        **existing,
        "user_id": TENANT_ID,
        "username": TENANT_ID,
        "display_name": DISPLAY_NAME,
        "role": "customer",
        "email": EMAIL,
        "password_hash": existing.get("password_hash") or hash_password(PASSWORD),
        "tenant_ids": [TENANT_ID],
        "active_tenant_id": TENANT_ID,
        "resource_scopes": ["*"],
        "initialized_at": existing.get("initialized_at") or now,
        "updated_at": now,
    }
    service.write_local_account_overrides(accounts)
    service.local_login(username=TENANT_ID, password=PASSWORD, tenant_id=TENANT_ID)


def login_client() -> tuple[TestClient, dict[str, str]]:
    client = TestClient(create_app())
    response = client.post(
        "/api/auth/login/start",
        json={"username": TENANT_ID, "password": PASSWORD, "tenant_id": TENANT_ID},
    )
    assert_status(response.status_code, 200, "customer login")
    token = response.json()["session"]["token"]
    return client, {"Authorization": f"Bearer {token}"}


def check_material_upload_learning(client: TestClient, headers: dict[str, str], token: str) -> dict[str, Any]:
    materials = write_materials(token)
    uploads = []
    jobs = []
    for kind, path in materials.items():
        with path.open("rb") as handle:
            upload = client.post(
                "/api/uploads",
                headers=headers,
                data={"kind": kind},
                files={"file": (path.name, handle.read(), "text/plain")},
            )
        assert_status(upload.status_code, 200, f"upload {kind}")
        upload_item = upload.json()["item"]
        uploads.append(upload_item)
        job = client.post("/api/learning/jobs", headers=headers, json={"upload_ids": [upload_item["upload_id"]], "use_llm": False})
        assert_status(job.status_code, 200, f"learning {kind}")
        jobs.append(job.json()["job"])

    rag_terms = [
        f"凯美瑞 自动挡 省油 {token}",
        f"宝马 320Li 首付三成 {token}",
        f"沉睡客户 唤醒 新车源 {token}",
    ]
    rag_checks = [rag_search_check(term) for term in rag_terms]
    candidates = list_candidate_payloads()
    matching_candidates = [candidate_summary(item) for item in candidates if token in json.dumps(item, ensure_ascii=False)]
    skipped_duplicates = [
        {
            "job_id": job.get("job_id"),
            "candidate_id": item.get("candidate_id"),
            "reason": (item.get("duplicate") or {}).get("reason"),
            "existing_item_id": (item.get("duplicate") or {}).get("existing_item_id"),
            "source": (item.get("duplicate") or {}).get("source"),
        }
        for job in jobs
        for item in job.get("skipped_duplicates", []) or []
    ]
    generated_or_deduped_count = len(matching_candidates) + len(skipped_duplicates)
    assert_true(all(item["found"] for item in rag_checks), "uploaded materials should be searchable in RAG")
    assert_true(
        generated_or_deduped_count >= 2 or all(int(job.get("candidate_count") or 0) == 0 for job in jobs),
        "uploaded materials should either dedupe old candidates or stay in RAG-only learning under the unified promotion chain",
    )
    return {
        "name": "material_upload_learning",
        "ok": True,
        "uploads": [{"upload_id": item.get("upload_id"), "kind": item.get("kind"), "path": item.get("path")} for item in uploads],
        "jobs": [
            {
                "job_id": item.get("job_id"),
                "candidate_count": item.get("candidate_count"),
                "candidate_ids": item.get("candidate_ids"),
                "skipped_duplicate_count": item.get("skipped_duplicate_count"),
            }
            for item in jobs
        ],
        "promotion_policy": "uploads first enter RAG; candidate creation requires explicit RAG-to-candidate promotion",
        "rag_checks": rag_checks,
        "matching_candidates": matching_candidates,
        "skipped_duplicates": skipped_duplicates,
    }


def write_materials(token: str) -> dict[str, Path]:
    root = BASE_ARTIFACT_ROOT / "materials" / token
    root.mkdir(parents=True, exist_ok=True)
    vehicles = root / f"chejin_vehicles_{token}.txt"
    policies = root / f"chejin_policies_{token}.txt"
    chats = root / f"chejin_chats_{token}.txt"
    vehicles.write_text(vehicle_material(token), encoding="utf-8")
    policies.write_text(policy_material(token), encoding="utf-8")
    chats.write_text(chat_material(token), encoding="utf-8")
    return {"products": vehicles, "policies": policies, "chats": chats}


def vehicle_material(token: str) -> str:
    return f"""
商品资料：江苏车金二手车车源 {token}
测试批次：{token}
商品名称：2021款丰田凯美瑞2.0G豪华版
型号：CHEJIN-CAMRY-2021G
商品类目：二手车/中级轿车
别名关键词：凯美瑞,丰田凯美瑞,8万预算,自动挡,省油,通勤
规格参数：2021年上牌，表显4.8万公里，2.0L自动挡，南京现车，一手车，支持第三方检测。
价格：8.98万
单位：台
库存：1
发货：南京门店可看车，异地客户需人工确认物流与提档。
售后：非重大事故、非水泡、非火烧以合同和检测报告为准。
标准回复：这台凯美瑞适合8到10万预算、家用通勤和省油需求，建议先确认到店时间、是否置换、是否贷款。

商品资料：江苏车金二手车车源 {token}
测试批次：{token}
商品名称：2020款本田思域220TURBO劲动版
型号：CHEJIN-CIVIC-2020T
商品类目：二手车/紧凑型轿车
别名关键词：思域,本田思域,年轻客户,按揭,首付三成
规格参数：2020年上牌，表显6.1万公里，1.5T自动挡，外观轻微补漆，支持按揭初审。
价格：7.58万
单位：台
库存：1
发货：门店看车后可协助过户，金融结果以资方审批为准。
售后：车况以检测报告和合同约定为准，不承诺贷款必过。

商品资料：江苏车金二手车车源 {token}
测试批次：{token}
商品名称：2019款宝马320Li M运动套装
型号：CHEJIN-BMW320-2019M
商品类目：二手车/豪华轿车
别名关键词：宝马320,宝马3系,320Li,试驾,精品车况
规格参数：2019年上牌，表显5.6万公里，2.0T自动挡，一手车源，支持到店试驾。
价格：12.80万
单位：台
库存：1
发货：高意向试驾需转人工确认门店和时间。
售后：精品车况需以检测报告为准，事故、水泡、火烧承诺必须人工确认。

商品资料：江苏车金二手车车源 {token}
测试批次：{token}
商品名称：2022款比亚迪秦PLUS DM-i 55KM
型号：CHEJIN-QINPLUS-2022DMI
商品类目：二手车/新能源轿车
别名关键词：秦PLUS,比亚迪秦,新能源,绿牌,混动,网约车
规格参数：2022年上牌，表显3.2万公里，插混，低油耗，适合通勤和网约车。
价格：8.68万
单位：台
库存：1
发货：新能源车需人工确认电池检测和当地迁入政策。
售后：电池和三电政策以厂家与检测报告为准。
""".strip() + "\n"


def policy_material(token: str) -> str:
    return f"""
政策规则：江苏车金高意向转人工 {token}
测试批次：{token}
规则名称：试驾到店与订金转人工
规则类型：manual_required
触发关键词：试驾,到店,订金,定金,今天看车,明天看车,销售联系,企业微信群
答案：客户明确要求试驾、到店、订金、销售联系时，AI只做确认和记录，必须转人工或企业微信群承接，不能自行承诺车源保留和最终价格。
允许自动回复：false
必须转人工：true
提醒人工客服：true
风险等级：high

政策规则：江苏车金金融与车况合规 {token}
测试批次：{token}
规则名称：金融审批与车况承诺边界
规则类型：contract
触发关键词：贷款,按揭,首付,月供,征信,保证无事故,水泡,火烧,赔偿,退款
答案：金融方案只能做初步说明，审批结果、利率、月供以资方审核为准；车况承诺必须以检测报告和合同为准，涉及赔偿、退款、纠纷必须转人工。
允许自动回复：false
必须转人工：true
提醒人工客服：true
风险等级：high

政策规则：江苏车金沉睡客户唤醒 {token}
测试批次：{token}
规则名称：沉睡客户递减唤醒策略
规则类型：other
触发关键词：沉睡客户,唤醒,老线索,新车源,还在看车
答案：首次唤醒只做轻提醒；3天未回复推同预算新车源；7天未回复询问需求是否变化；客户回复后停止自动唤醒并转入正常接待。
允许自动回复：true
必须转人工：false
风险等级：normal
""".strip() + "\n"


def chat_material(token: str) -> str:
    return f"""
聊天记录：江苏车金预算推荐话术 {token}
测试批次：{token}
客户：我从抖音直播来的，8万左右想买自动挡省油代步车，有什么推荐？
客服：您这个预算可以先看凯美瑞、思域和秦PLUS DM-i。凯美瑞偏家用稳、省心；思域更年轻；秦PLUS更省油。您是在南京看车吗？是否考虑贷款或置换？
意图标签：预算推荐,车源推荐,抖音线索

聊天记录：江苏车金置换话术 {token}
测试批次：{token}
客户：我有一台老朗逸想置换。
客服：可以的，麻烦补充上牌年份、公里数、车况、所在城市和是否有贷款，我先帮您记录，评估价需要人工复核。
意图标签：置换,线索采集

聊天记录：江苏车金睡客唤醒话术 {token}
测试批次：{token}
客户：之前那台没买，现在还有新车源吗？
客服：有的，我先按您之前的预算和用途筛一遍；如果预算或车型变了，直接告诉我，我会按新需求推荐更合适的车。
意图标签：沉睡客户,唤醒,二次转化
""".strip() + "\n"


def seed_formal_used_car_knowledge(token: str) -> dict[str, Any]:
    store = KnowledgeBaseStore()
    items = [
        (
            "products",
            {
                "schema_version": 1,
                "category_id": "products",
                "id": "chejin_camry_2021_20g",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "name": "2021款丰田凯美瑞2.0G豪华版",
                    "sku": "CHEJIN-CAMRY-2021G",
                    "category": "二手车/中级轿车",
                    "aliases": ["凯美瑞", "丰田凯美瑞", "8万预算", "自动挡省油", token],
                    "specs": "2021年上牌，表显4.8万公里，2.0L自动挡，南京现车。",
                    "price": 8.98,
                    "unit": "台",
                    "inventory": 1,
                    "shipping_policy": "南京门店可看车，异地提档和物流需人工确认。",
                    "warranty_policy": "车况以检测报告和合同为准，不承诺口头赔付。",
                    "reply_templates": {"recommendation": "适合8到10万预算、家用通勤和省油需求。"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "policies",
            {
                "schema_version": 1,
                "category_id": "policies",
                "id": "chejin_handoff_high_intent",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "title": "江苏车金试驾到店转人工",
                    "policy_type": "manual_required",
                    "keywords": ["试驾", "到店", "订金", "定金", "保证无事故", "包过"],
                    "answer": "客户明确要求试驾、到店、订金或事故赔付承诺时，AI只能记录并转人工确认。",
                    "allow_auto_reply": False,
                    "requires_handoff": True,
                    "handoff_reason": "used_car_high_intent_or_risk",
                    "operator_alert": True,
                    "risk_level": "high",
                },
                "runtime": {"allow_auto_reply": False, "requires_handoff": True, "risk_level": "high", "operator_alert": True},
            },
        ),
        (
            "policies",
            {
                "schema_version": 1,
                "category_id": "policies",
                "id": "chejin_douyin_lead_first_reply",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "title": "江苏车金抖音线索首轮接待",
                    "policy_type": "company",
                    "keywords": ["抖音", "直播", "刚加", "新来的"],
                    "answer": "抖音直播或新加微信线索可以由 AI 先问预算、用途、想看的车型和所在城市，并提示后续由销售确认到店或成交细节。",
                    "allow_auto_reply": True,
                    "requires_handoff": False,
                    "operator_alert": False,
                    "risk_level": "normal",
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "policies",
            {
                "schema_version": 1,
                "category_id": "policies",
                "id": "chejin_trade_in_info_collect",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "title": "江苏车金置换资料收集",
                    "policy_type": "company",
                    "keywords": ["置换", "卖车", "评估", "旧车"],
                    "answer": "客户咨询置换时，AI 可以先收集旧车品牌车型、上牌年份、公里数、车况、所在城市和是否有贷款；估价、收车价和最终成交必须人工复核。",
                    "allow_auto_reply": True,
                    "requires_handoff": False,
                    "operator_alert": False,
                    "risk_level": "normal",
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "chats",
            {
                "schema_version": 1,
                "category_id": "chats",
                "id": "chejin_budget_chat_commuter",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "customer_message": "8万预算自动挡省油代步",
                    "service_reply": "可以优先看凯美瑞、思域、秦PLUS DM-i，先确认城市、贷款、置换和到店时间。",
                    "intent_tags": ["scene_product", "catalog", "quote"],
                    "tone_tags": ["专业", "克制"],
                    "usable_as_template": True,
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "product_faq",
            {
                "schema_version": 1,
                "category_id": "product_faq",
                "id": "chejin_camry_commute_faq",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "product_id": "chejin_camry_2021_20g",
                    "title": "凯美瑞适合家用通勤吗",
                    "keywords": ["凯美瑞", "家用", "通勤", "省油"],
                    "question": "这台凯美瑞适合家用通勤吗？",
                    "answer": "这台凯美瑞偏家用稳、省心，适合8到10万预算和日常通勤；仍建议到店看车并以检测报告为准。",
                    "additional_details": {"适用商品": "2021款丰田凯美瑞2.0G豪华版"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "product_rules",
            {
                "schema_version": 1,
                "category_id": "product_rules",
                "id": "chejin_qinplus_battery_handoff",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "product_id": "chejin_qinplus_2022_dmi55",
                    "title": "秦PLUS电池检测必须人工确认",
                    "keywords": ["秦PLUS", "电池检测", "三电", "迁入政策"],
                    "answer": "秦PLUS 的电池检测、三电政策和当地迁入政策必须由人工结合检测报告确认，AI 不做最终承诺。",
                    "allow_auto_reply": False,
                    "requires_handoff": True,
                    "handoff_reason": "new_energy_battery_and_policy_check",
                    "additional_details": {"适用商品": "2022款比亚迪秦PLUS DM-i 55KM"},
                },
                "runtime": {"allow_auto_reply": False, "requires_handoff": True, "risk_level": "high", "operator_alert": True},
            },
        ),
        (
            "product_explanations",
            {
                "schema_version": 1,
                "category_id": "product_explanations",
                "id": "chejin_gl8_business_scene",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "product_id": "chejin_gl8_2020_es653t",
                    "title": "GL8适合商务接待和七座刚需",
                    "keywords": ["GL8", "商务接待", "七座", "试乘"],
                    "content": "这台 GL8 更适合商务接待、多人家庭和七座刚需。涉及试乘、付款方式、内饰磨损和保养记录时，应转人工结合实车说明。",
                    "additional_details": {"适用商品": "2020款别克GL8 ES陆尊653T豪华型"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
    ]
    items.extend(diverse_formal_fixture_items(token))
    saved = []
    for category, item in items:
        item = mark_item_new(item, {"source_module": "jiangsu_chejin_test", "target_category": category, "item_id": item["id"]})
        result = store.save_item(category, item)
        assert_true(result.get("ok"), f"formal fixture save failed: {category} {result}")
        saved.append({"category": category, "id": item["id"]})
    return {"name": "seed_formal_used_car_knowledge", "ok": True, "items": saved}


def diverse_formal_fixture_items(token: str) -> list[tuple[str, dict[str, Any]]]:
    return [
        (
            "products",
            {
                "schema_version": 1,
                "category_id": "products",
                "id": "chejin_civic_2020_220turbo",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "name": "2020款本田思域220TURBO劲动版",
                    "sku": "CHEJIN-CIVIC-2020T",
                    "category": "二手车/紧凑型轿车",
                    "aliases": ["思域", "本田思域", "年轻客户", "首付三成", token],
                    "specs": "2020年上牌，表显6.1万公里，1.5T自动挡，外观轻微补漆。",
                    "price": 7.58,
                    "unit": "台",
                    "inventory": 1,
                    "shipping_policy": "门店看车后可协助过户，金融结果以资方审批为准。",
                    "warranty_policy": "车况以检测报告和合同为准，不承诺贷款必过。",
                    "reply_templates": {"recommendation": "适合年轻客户、预算7到8万、想要动力和外观的需求。"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "products",
            {
                "schema_version": 1,
                "category_id": "products",
                "id": "chejin_bmw320_2019_m",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "name": "2019款宝马320Li M运动套装",
                    "sku": "CHEJIN-BMW320-2019M",
                    "category": "二手车/豪华轿车",
                    "aliases": ["宝马320", "宝马3系", "320Li", "试驾", token],
                    "specs": "2019年上牌，表显5.6万公里，2.0T自动挡，一手车源。",
                    "price": 12.80,
                    "unit": "台",
                    "inventory": 1,
                    "shipping_policy": "高意向试驾需转人工确认门店和时间。",
                    "warranty_policy": "精品车况需以检测报告为准，事故、水泡、火烧承诺必须人工确认。",
                    "reply_templates": {"recommendation": "适合12到14万预算、关注品牌和驾驶感的客户。"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "products",
            {
                "schema_version": 1,
                "category_id": "products",
                "id": "chejin_qinplus_2022_dmi55",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "name": "2022款比亚迪秦PLUS DM-i 55KM",
                    "sku": "CHEJIN-QINPLUS-2022DMI",
                    "category": "二手车/新能源轿车",
                    "aliases": ["秦PLUS", "比亚迪秦", "新能源", "绿牌", "混动", token],
                    "specs": "2022年上牌，表显3.2万公里，插混，低油耗。",
                    "price": 8.68,
                    "unit": "台",
                    "inventory": 1,
                    "shipping_policy": "新能源车需人工确认电池检测和当地迁入政策。",
                    "warranty_policy": "电池和三电政策以厂家与检测报告为准。",
                    "reply_templates": {"recommendation": "适合通勤、省油、绿牌或网约车用途。"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "products",
            {
                "schema_version": 1,
                "category_id": "products",
                "id": "chejin_gl8_2020_es653t",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "name": "2020款别克GL8 ES陆尊653T豪华型",
                    "sku": "CHEJIN-GL8-2020ES653T",
                    "category": "二手车/MPV",
                    "aliases": ["GL8", "别克GL8", "商务接待", "七座", token],
                    "specs": "2020年上牌，表显7.4万公里，2.0T自动挡，七座商务MPV。",
                    "price": 17.60,
                    "unit": "台",
                    "inventory": 1,
                    "shipping_policy": "商务客户看车、试乘、付款方式均需人工确认。",
                    "warranty_policy": "内饰磨损和保养记录以门店实车与检测报告为准。",
                    "reply_templates": {"recommendation": "适合商务接待、多人家庭和七座刚需客户。"},
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "chats",
            {
                "schema_version": 1,
                "category_id": "chats",
                "id": "chejin_trade_in_chat_collect",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "customer_message": "我有一台老朗逸想置换",
                    "service_reply": "可以置换。麻烦补充旧车品牌车型、上牌年份、公里数、车况、所在城市和是否有贷款，估价需要人工复核。",
                    "intent_tags": ["trade_in", "lead_capture"],
                    "tone_tags": ["专业", "克制"],
                    "usable_as_template": True,
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
        (
            "chats",
            {
                "schema_version": 1,
                "category_id": "chats",
                "id": "chejin_sleep_wakeup_chat_new_stock",
                "status": "active",
                "source": {"type": "test_fixture", "batch_token": token},
                "data": {
                    "customer_message": "之前那台没买，现在还有新车源吗",
                    "service_reply": "有的，我先按您之前的预算和用途筛；如果预算或车型变了，直接告诉我，我会重新匹配。",
                    "intent_tags": ["sleep_wakeup", "reconversion"],
                    "tone_tags": ["自然", "克制"],
                    "usable_as_template": True,
                },
                "runtime": {"allow_auto_reply": True, "requires_handoff": False, "risk_level": "normal"},
            },
        ),
    ]


def check_customer_service_matrix(token: str) -> dict[str, Any]:
    config = used_car_service_config(token)
    rules = used_car_rules(token)
    target = TargetConfig(name="车金测试私聊", enabled=True, exact=True, allow_self_for_test=True, max_batch_messages=3)
    state: dict[str, Any] = {"version": 1, "targets": {}}
    cases = []

    cases.append(run_service_case(config, rules, target, state, "cs-01", "我从抖音直播来的，8万左右想看看自动挡", expect_action="sent", expect_contains="预算"))
    cases.append(run_service_case(config, rules, target, state, "cs-02", f"8万预算，自动挡省油，通勤代步 {token}", expect_action="sent", expect_contains="凯美瑞"))
    cases.append(run_service_case(config, rules, target, state, "cs-03", f"宝马320Li那台今天能到店试驾吗 {token}", expect_action="handoff_sent", expect_handoff=True))
    cases.append(run_service_case(config, rules, target, state, "cs-04", "思域首付三成月供大概多少，贷款能不能包过？", expect_action="handoff_sent", expect_handoff=True))
    cases.append(run_service_case(config, rules, target, state, "cs-05", "我有一台老朗逸想置换", expect_action="sent", expect_contains="上牌年份"))
    cases.append(run_service_case(config, rules, target, state, "cs-08", "你保证无事故吗，不对就赔我？", expect_action="handoff_sent", expect_handoff=True))
    cases.append(run_service_case(config, rules, target, state, "cs-11", "今天天气怎么样，顺便讲个笑话", expect_action="handoff_sent", expect_handoff=True))

    data_state: dict[str, Any] = {"version": 1, "targets": {}}
    incomplete = run_service_case(
        config,
        rules,
        target,
        data_state,
        "cs-07",
        "客户资料\n电话：13900001111\n预算：8万\n车型：自动挡省油代步",
        expect_action="sent",
        expect_contains="姓名",
    )
    complete = run_service_case(
        config,
        rules,
        target,
        data_state,
        "cs-06",
        "姓名：张先生\n电话：13900001111\n预算：8万\n车型：凯美瑞或思域\n到店时间：明天下午",
        expect_action="handoff_sent",
        expect_contains="记",
        expect_handoff=True,
        write_data=True,
    )
    cases.extend([incomplete, complete])
    workbook_path = Path(config["data_capture"]["workbook_path"])
    assert_true(workbook_path.exists(), "complete lead should create Excel workbook")
    workbook = load_workbook(workbook_path)
    assert_true(workbook[config["data_capture"]["sheet_name"]].max_row >= 2, "lead workbook should contain data row")

    bot_prefix = configured_reply_prefix(config)
    connector = FakeConnector(
        [
            {"id": "bot-skip", "type": "text", "sender": "self", "content": bot_prefix + "上一轮客服回复，不应重复处理"},
            {"id": "real-after-bot", "type": "text", "sender": "self", "content": "还有哪些车源可以看看？"},
        ]
    )
    event = process_target(
        connector=connector,
        target=target,
        config=config,
        rules=rules,
        state={"version": 1, "targets": {}},
        send=True,
        write_data=False,
        allow_fallback_send=False,
        mark_dry_run=False,
    )
    assert_true(event.get("message_ids") == ["real-after-bot"], "bot prefix should be skipped")
    cases.append({"case": "cs-12", "ok": True, "action": event.get("action"), "message_ids": event.get("message_ids")})

    return {"name": "customer_service_matrix", "ok": True, "cases": cases, "lead_workbook": str(workbook_path)}


def used_car_service_config(token: str) -> dict[str, Any]:
    root = BASE_ARTIFACT_ROOT / "customer_service" / token
    root.mkdir(parents=True, exist_ok=True)
    return {
        "version": 1,
        "reply": {"prefix": "[车金AI] ", "allow_fallback_send": False},
        "rate_limits": {"min_seconds_between_replies": 0, "max_replies_per_10_minutes": 50, "max_replies_per_hour": 200, "notice_customer": False},
        "raw_messages": {"enabled": True, "learning_enabled": True, "auto_learn": True, "use_llm": False, "notify_enabled": False},
        "data_capture": {
            "enabled": True,
            "workbook_path": str(root / "chejin_leads.xlsx"),
            "sheet_name": "二手车线索",
            "required_fields": ["name", "phone"],
            "write_on_send_only": False,
            "success_reply": "客户资料已记录，我会让销售尽快跟进。",
            "incomplete_reply": "我看到了客户资料，但还缺少：{missing_fields}。请补充后我再记录。",
        },
        "product_knowledge": {"enabled": False},
        "handoff": {
            "enabled": True,
            "auto_acknowledge": True,
            "acknowledgement_reply": "这个问题需要销售人工确认，我先帮您记录并提醒同事跟进。",
        },
        "operator_alert": {"enabled": True, "alert_log_path": str(root / "operator_alerts.jsonl")},
        "service_profile": {
            "role": "二手车微信销冠客服",
            "tone": "简短、真实、克制、不夸大车况",
            "answer_boundary": "不承诺最低价、贷款必过、绝对无事故或赔偿；高意向和高风险转人工。",
        },
        "intent_assist": {"enabled": True, "mode": "heuristic", "advisory_only": True, "llm_advisory": {"enabled": False}},
        "rag_response": {
            "enabled": True,
            "apply_to_unmatched": True,
            "apply_to_matched_product": False,
            "apply_to_small_talk": False,
            "skip_llm_after_apply": True,
            "min_hit_score": 0.12,
            "max_reply_chars": 240,
            "max_snippet_chars": 140,
        },
    }


def used_car_rules(token: str) -> dict[str, Any]:
    return {
        "default_reply": "收到，我先记录一下，涉及车况、价格、金融或到店安排会请销售人工确认。",
        "rules": [
            {
                "name": "douyin_lead_greeting",
                "priority": 100,
                "keywords": ["抖音", "直播", "刚加", "新来的"],
                "reply": "您好，我是江苏车金AI助手。您可以直接告诉我预算、用途、想看的车型和所在城市，我先帮您筛车源。",
            },
            {
                "name": "budget_recommendation",
                "priority": 90,
                "keywords": ["8万预算", "自动挡", "省油", "代步", token],
                "reply": "8万左右自动挡省油代步，可以优先看凯美瑞、思域和秦PLUS DM-i。凯美瑞偏家用稳，思域更年轻，秦PLUS更省油。您是在南京看车吗？是否考虑贷款或置换？",
            },
            {
                "name": "trade_in",
                "priority": 80,
                "keywords": ["置换", "卖车", "评估"],
                "reply": "可以置换。麻烦补充旧车品牌车型、上牌年份、公里数、车况、所在城市和是否有贷款，我先帮您登记，估价需要人工复核。",
            },
            {
                "name": "finance_guarded",
                "priority": 70,
                "keywords": ["首付", "月供", "贷款", "按揭"],
                "reply": "金融可以先做方案测算，但首付、月供、利率和能否通过都要以资方审批为准。我可以先记录预算和车型，再让销售确认。",
            },
            {
                "name": "sleep_wakeup_reply",
                "priority": 60,
                "keywords": ["之前", "新车源", "还在看车", "现在还有"],
                "reply": "有新车源可以继续筛。您之前的预算和用途如果没变，我按原需求推荐；如果预算或车型变了，直接告诉我我重新匹配。",
            },
        ],
    }


def run_service_case(
    config: dict[str, Any],
    rules: dict[str, Any],
    target: TargetConfig,
    state: dict[str, Any],
    case_id: str,
    content: str,
    *,
    expect_action: str,
    expect_contains: str = "",
    expect_handoff: bool = False,
    write_data: bool = False,
) -> dict[str, Any]:
    message_id = f"{case_id}-{datetime.now().timestamp()}"
    connector = FakeConnector([{"id": message_id, "type": "text", "sender": "self", "content": content}])
    event = process_target(
        connector=connector,
        target=target,
        config=config,
        rules=rules,
        state=state,
        send=True,
        write_data=write_data,
        allow_fallback_send=False,
        mark_dry_run=False,
    )
    assert_true(event.get("action") == expect_action, f"{case_id} expected {expect_action}, got {event.get('action')} {event}")
    reply_text = str((event.get("decision") or {}).get("reply_text") or (connector.sent_texts[-1] if connector.sent_texts else ""))
    if expect_contains:
        assert_true(expect_contains in reply_text, f"{case_id} reply should contain {expect_contains}: {reply_text}")
    if expect_handoff:
        assert_true(bool((event.get("decision") or {}).get("need_handoff")) or event.get("action") in {"handoff", "handoff_sent"}, f"{case_id} should hand off")
    return {
        "case": case_id,
        "ok": True,
        "action": event.get("action"),
        "rule": (event.get("decision") or {}).get("rule_name"),
        "rag_applied": bool((event.get("rag_reply") or {}).get("applied")),
        "reply_text": reply_text,
    }


def check_recorder_offline_matrix(token: str) -> dict[str, Any]:
    store = RawMessageStore()
    group_conversation = {
        "target_name": "离线回归-记录员群聊",
        "display_name": "离线回归-记录员群聊",
        "conversation_type": "group",
        "selected_by_user": False,
        "learning_enabled": True,
        "notify_enabled": False,
        "source": {"type": "jiangsu_chejin_offline_recorder", "batch_token": token},
    }
    file_conversation = {
        "target_name": FILE_TRANSFER_ASSISTANT,
        "display_name": FILE_TRANSFER_ASSISTANT,
        "conversation_type": "file_transfer",
        "selected_by_user": True,
        "learning_enabled": True,
        "notify_enabled": False,
        "source": {"type": "jiangsu_chejin_offline_recorder", "batch_token": token},
    }
    group_messages = [
        {
            "id": f"{token}-rec-product",
            "type": "text",
            "sender": "销售A",
            "content": f"商品资料：江苏车金记录员完整车源 {token}\n测试批次：{token}\n商品名称：2020款别克GL8 ES陆尊653T豪华型\n型号：CHEJIN-REC-GL8-2020ES\n商品类目：二手车/MPV\n价格：17.60万\n单位：台\n库存：1\n发货：南京门店可看车，商务客户试乘需人工确认\n售后：车况以检测报告为准",
            "time": "2026-05-01 16:00:00",
        },
        {
            "id": f"{token}-rec-risk",
            "type": "text",
            "sender": "销售B",
            "content": f"政策规则：江苏车金记录员风控 {token}\n测试批次：{token}\n规则名称：保证无事故转人工\n规则类型：contract\n触发关键词：保证无事故,赔偿,{token}\n答案：此类问题必须让销售人工按检测报告和合同回复。",
            "time": "2026-05-01 16:01:00",
        },
        {
            "id": f"{token}-rec-noise",
            "type": "text",
            "sender": "销售C",
            "content": f"边界噪音：{token}_NOISE 今天午饭吃什么，这不是车源也不是客服知识。",
            "time": "2026-05-01 16:02:00",
        },
    ]
    file_messages = [
        {
            "id": f"{token}-rec-chat",
            "type": "text",
            "sender": "self",
            "content": f"聊天记录：江苏车金记录员话术 {token}\n客户：想买新能源通勤，能看秦PLUS吗？\n客服：可以先看秦PLUS DM-i，低油耗适合通勤；但电池检测、当地迁入和金融方案都需要人工确认。",
            "time": "2026-05-01 16:03:00",
        }
    ]
    first = store.upsert_messages(group_conversation, group_messages, source_module="jiangsu_chejin_recorder_test", batch_reason="offline_recorder")
    second = store.upsert_messages(file_conversation, file_messages, source_module="jiangsu_chejin_recorder_test", batch_reason="offline_recorder")
    duplicate = store.upsert_messages(group_conversation, group_messages, source_module="jiangsu_chejin_recorder_test", batch_reason="offline_recorder")
    assert_true(duplicate["inserted_count"] == 0, "recorder duplicate messages should not insert twice")
    learning_results = []
    for result in (first, second):
        batch = result.get("batch")
        if batch:
            learning_results.append(RawMessageLearningService().process_batch(str(batch.get("batch_id") or ""), use_llm=False))
    candidate_matches = [candidate_summary(item) for item in list_candidate_payloads() if token in json.dumps(item, ensure_ascii=False)]
    noise_candidates = [item for item in candidate_matches if f"{token}_NOISE" in json.dumps(item, ensure_ascii=False)]
    raw_matches = [item for item in store.list_messages(limit=500) if token in str(item.get("content") or "")]
    rag_checks = [rag_search_check(f"别克GL8 ES陆尊 {token}"), rag_search_check(f"秦PLUS DM-i 电池检测 {token}")]
    first_learning = learning_results[0] if learning_results else {}
    second_learning = learning_results[1] if len(learning_results) > 1 else {}
    group_candidate_or_deduped_count = int(first_learning.get("candidate_count") or 0) + int(first_learning.get("skipped_duplicate_count") or 0)
    chat_candidate_or_deduped_count = int(second_learning.get("candidate_count") or 0) + int(second_learning.get("skipped_duplicate_count") or 0)
    assert_true(len(raw_matches) >= 4, "recorder raw messages should be stored")
    assert_true(all(item["found"] for item in rag_checks), "recorder learned batches should enter RAG")
    assert_true(
        group_candidate_or_deduped_count >= 0 and chat_candidate_or_deduped_count >= 0,
        "recorder learning should report candidate counters even when unified chain keeps new data in RAG only",
    )
    assert_true(not noise_candidates, "noise should not produce candidates")
    return {
        "name": "recorder_offline_matrix",
        "ok": True,
        "raw_message_ids": [item.get("raw_message_id") for item in raw_matches],
        "learning_results": [
            {"candidate_count": item.get("candidate_count"), "candidate_ids": item.get("candidate_ids"), "skipped_duplicate_count": item.get("skipped_duplicate_count")}
            for item in learning_results
        ],
        "candidate_matches": candidate_matches,
        "promotion_policy": "recorder messages first enter RAG; candidate creation requires explicit RAG-to-candidate promotion",
        "rag_checks": rag_checks,
    }


def check_tenant_scoped_backend(client: TestClient, headers: dict[str, str], token: str) -> dict[str, Any]:
    raw = client.get("/api/raw-messages/messages", headers=headers, params={"query": token, "limit": 200})
    assert_status(raw.status_code, 200, "raw message query")
    candidates = client.get("/api/candidates", headers=headers, params={"status": "pending"})
    assert_status(candidates.status_code, 200, "candidate list")
    rag = client.post("/api/rag/search", headers=headers, json={"query": f"凯美瑞 自动挡 {token}", "limit": 10})
    assert_status(rag.status_code, 200, "rag search")
    matching_candidates = [candidate_summary(item) for item in candidates.json().get("items", []) if token in json.dumps(item, ensure_ascii=False)]
    return {
        "name": "tenant_scoped_backend",
        "ok": True,
        "raw_count": len(raw.json().get("items", [])),
        "pending_candidate_count": len(candidates.json().get("items", [])),
        "matching_candidates": matching_candidates,
        "rag_found": any(token in str(hit.get("text") or "") for hit in rag.json().get("hits", [])),
        "paths": {
            "tenant_root": str(tenant_root(TENANT_ID)),
            "raw_messages": str(tenant_runtime_root(TENANT_ID) / "raw_messages" / "messages.json"),
            "review_candidates": str(tenant_review_candidates_root(TENANT_ID) / "pending"),
            "rag_index": str(tenant_rag_index_root(TENANT_ID) / "index.json"),
        },
    }


def check_live_wechat_matrix(token: str, *, group_name: str) -> dict[str, Any]:
    connector = WeChatConnector()
    status = connector.status()
    if not status.get("ok") or not status.get("online"):
        raise AssertionError(f"WeChat is not online: {status}")
    service = RecorderService()
    raw_store = RawMessageStore()
    service.save_settings(
        {
            "group_recording_enabled": True,
            "private_recording_enabled": True,
            "file_transfer_recording_enabled": True,
            "auto_learn": True,
            "use_llm": False,
            "notify_on_collect": False,
        }
    )
    group = service.ensure_conversation(
        {
            "target_name": group_name,
            "display_name": group_name,
            "conversation_type": "group",
            "selected_by_user": True,
            "status": "active",
            "exact": True,
            "learning_enabled": True,
            "notify_enabled": False,
            "source": {"type": "jiangsu_chejin_live_test", "batch_token": token},
        }
    )
    file_target = service.ensure_conversation(
        {
            "target_name": FILE_TRANSFER_ASSISTANT,
            "display_name": FILE_TRANSFER_ASSISTANT,
            "conversation_type": "file_transfer",
            "selected_by_user": True,
            "status": "active",
            "exact": True,
            "learning_enabled": True,
            "notify_enabled": False,
            "source": {"type": "jiangsu_chejin_live_test", "batch_token": token},
        }
    )
    live_messages = [
        (
            group_name,
            f"商品资料：江苏车金真机记录员车源 {token}\n测试批次：{token}\n商品名称：2020款别克GL8 ES陆尊653T豪华型\n型号：CHEJIN-LIVE-GL8-2020ES\n商品类目：二手车/MPV\n价格：17.66万\n单位：台\n库存：1\n发货：南京门店可看车，商务客户试乘需人工确认\n售后：车况以检测报告为准",
        ),
        (
            group_name,
            f"政策规则：江苏车金真机金融边界 {token}\n测试批次：{token}\n规则名称：新能源电池与金融审批转人工\n规则类型：contract\n触发关键词：电池检测,首付,月供,贷款包过,{token}\n答案：涉及新能源电池检测、首付月供、贷款通过率时必须转人工确认，不允许AI承诺结果。",
        ),
        (group_name, f"边界噪音：{token}_LIVE_NOISE 今天群里测试心跳，不是车源知识。"),
        (
            FILE_TRANSFER_ASSISTANT,
            f"我从抖音直播来的，想看秦PLUS或者思域，预算8到9万，最好省油，能不能贷款包过？ {token}",
        ),
    ]
    sent = []
    for target, text in live_messages:
        send = connector.send_text(target, text, exact=True)
        sent.append({"target": target, "ok": bool(send.get("ok")), "text": text[:80]})
        assert_true(bool(send.get("ok")), f"live send failed: {send}")
        time.sleep(1.8)
    capture = retry_capture(service, raw_store, token)
    assert_true(capture.get("ok"), f"live capture did not find token: {capture}")

    config = used_car_service_config(token)
    rules = used_car_rules(token)
    target_config = TargetConfig(name=FILE_TRANSFER_ASSISTANT, enabled=True, exact=True, allow_self_for_test=True, max_batch_messages=2)
    service_event = process_target(
        connector=connector,
        target=target_config,
        config=config,
        rules=rules,
        state={"version": 1, "targets": {}},
        send=True,
        write_data=False,
        allow_fallback_send=False,
        mark_dry_run=False,
    )
    assert_true(service_event.get("action") in {"sent", "handoff_sent"}, f"live service reply should send: {service_event}")
    idempotency = service.capture_selected_once(send_notifications=False)
    return {
        "name": "live_wechat_matrix",
        "ok": True,
        "status_user": (status.get("my_info") or {}).get("display_name"),
        "targets": {"group": group, "file_transfer": file_target},
        "sent": sent,
        "capture": capture,
        "service_event": {
            "action": service_event.get("action"),
            "reply_text": (service_event.get("decision") or {}).get("reply_text"),
            "rag_applied": bool((service_event.get("rag_reply") or {}).get("applied")),
        },
        "idempotency_capture": {"inserted_count": idempotency.get("inserted_count"), "conversation_count": idempotency.get("conversation_count")},
    }


def retry_capture(service: RecorderService, raw_store: RawMessageStore, token: str) -> dict[str, Any]:
    captures = []
    for _ in range(5):
        capture = service.capture_selected_once(send_notifications=False)
        captures.append(capture)
        messages = [item for item in raw_store.list_messages(limit=500) if token in str(item.get("content") or "")]
        if len(messages) >= 3:
            return {"ok": True, "captures": captures, "raw_message_ids": [item.get("raw_message_id") for item in messages]}
        time.sleep(1.8)
    return {"ok": False, "captures": captures, "raw_message_ids": []}


def rag_search_check(query: str) -> dict[str, Any]:
    search = RagService().search(query, limit=30)
    return {
        "query": query,
        "found": any(any(part and part in str(hit.get("text") or "") for part in query.split()[:3]) for hit in search.get("hits", []) or []),
        "hit_count": len(search.get("hits", []) or []),
        "top_hits": [
            {"source_id": hit.get("source_id"), "score": hit.get("score"), "text": str(hit.get("text") or "")[:160]}
            for hit in (search.get("hits", []) or [])[:3]
        ],
    }


def list_candidate_payloads() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for status in ("pending", "approved", "rejected"):
        root = tenant_review_candidates_root() / status
        if not root.exists():
            continue
        for path in root.glob("*.json"):
            try:
                items.append(json.loads(path.read_text(encoding="utf-8")))
            except (OSError, json.JSONDecodeError):
                continue
    return items


def candidate_summary(candidate: dict[str, Any]) -> dict[str, Any]:
    patch = ((candidate.get("proposal") or {}).get("formal_patch") or {})
    return {
        "candidate_id": candidate.get("candidate_id"),
        "target_category": patch.get("target_category"),
        "review_status": (candidate.get("review") or {}).get("status", "pending"),
        "completeness_status": (candidate.get("review") or {}).get("completeness_status") or (candidate.get("intake") or {}).get("status"),
        "source_type": (candidate.get("source") or {}).get("type"),
    }


def assert_status(actual: int, expected: int, message: str) -> None:
    if actual != expected:
        raise AssertionError(f"{message}: expected {expected}, got {actual}")


def assert_true(value: Any, message: str) -> None:
    if not value:
        raise AssertionError(message)


if __name__ == "__main__":
    raise SystemExit(main())
