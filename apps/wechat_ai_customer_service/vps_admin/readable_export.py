"""Human-readable export helpers for VPS customer data packages."""

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from apps.wechat_ai_customer_service.knowledge_paths import runtime_app_root


EXPORT_ROOT = runtime_app_root() / "vps_admin" / "readable_exports"
MAX_CELL_TEXT = 32000
FORMAL_CATEGORY_LABELS = {
    "products": "商品资料",
    "chats": "聊天记录与话术",
    "policies": "政策规则",
    "erp_exports": "ERP导出",
}
PRODUCT_SECTION_LABELS = {
    "rules": "规则",
    "faq": "FAQ",
    "explanations": "说明",
}


def build_customer_readable_workbook(package: dict[str, Any], package_path: Path) -> Path:
    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
    package_id = safe_filename(str(package.get("package_id") or package_path.stem))
    output_path = EXPORT_ROOT / f"{package_id}_readable.xlsx"
    workbook = Workbook()
    default = workbook.active
    default.title = "说明"

    data = collect_package_rows(package_path)
    write_intro(default, package, data)
    write_grouped_sheets(
        workbook,
        "正式",
        ["条目ID", "状态", "标题/名称", "正文/回复", "关键词", "自动回复", "需人工", "风险", "来源文件"],
        data["formal_by_category"],
        FORMAL_CATEGORY_LABELS,
    )
    write_grouped_sheets(
        workbook,
        "商品专属",
        ["商品ID", "条目ID", "状态", "标题/名称", "正文/回复", "关键词", "自动回复", "需人工", "来源文件"],
        data["product_by_section"],
        PRODUCT_SECTION_LABELS,
    )
    write_sheet(
        workbook,
        "RAG资料",
        ["资料ID", "分类", "商品ID", "标题", "正文/摘要", "来源文件"],
        data["rag_sources"],
    )
    write_sheet(
        workbook,
        "RAG经验",
        ["经验ID", "状态", "客户问法", "回复要点", "命中资料", "质量", "使用次数", "来源文件"],
        data["rag_experience"],
    )
    write_sheet(
        workbook,
        "技术文件清单",
        ["文件路径", "大小", "SHA256"],
        data["manifest"],
    )
    workbook.save(output_path)
    return output_path


def collect_package_rows(package_path: Path) -> dict[str, Any]:
    rows: dict[str, Any] = {
        "formal": [],
        "formal_by_category": {},
        "product": [],
        "product_by_section": {},
        "rag_sources": [],
        "rag_experience": [],
        "manifest": [],
    }
    with zipfile.ZipFile(package_path) as archive:
        manifest = read_archive_json(archive, "manifest.json", default={})
        for item in manifest.get("files", []) if isinstance(manifest, dict) else []:
            rows["manifest"].append([item.get("path", ""), item.get("bytes", 0), item.get("sha256", "")])
        for name in sorted(archive.namelist()):
            if not name.endswith(".json") or name == "manifest.json":
                continue
            payload = read_archive_json(archive, name, default=None)
            if payload is None:
                continue
            normalized = normalize_payload(payload)
            if "/knowledge_bases/" in name and "/items/" in name:
                category_id = str(normalized.get("category_id") or category_from_formal_path(name) or "unknown")
                row = formal_row(name, normalized)
                rows["formal"].append(row)
                rows["formal_by_category"].setdefault(category_id, []).append(row[1:])
            elif "/product_item_knowledge/" in name and "/items/" not in name:
                product_id, section = product_path_parts(name)
                section_id = section or "other"
                row = product_row(name, normalized)
                rows["product"].append(row)
                rows["product_by_section"].setdefault(section_id, []).append([row[0], *row[2:]])
            elif name.endswith("/rag_sources/sources.json"):
                rows["rag_sources"].extend(rag_source_rows(name, payload))
            elif name.endswith("/rag_experience/experiences.json"):
                rows["rag_experience"].extend(rag_experience_rows(name, payload))
    return rows


def formal_row(path: str, item: dict[str, Any]) -> list[Any]:
    data = item_data(item)
    runtime = item.get("runtime") if isinstance(item.get("runtime"), dict) else {}
    return [
        item.get("category_id") or category_from_formal_path(path),
        item.get("id") or Path(path).stem,
        item.get("status") or "",
        human_title(item),
        human_content(item),
        join_value(data.get("keywords") or data.get("intent_tags") or data.get("aliases")),
        bool_label(data.get("allow_auto_reply", runtime.get("allow_auto_reply"))),
        bool_label(data.get("requires_handoff", runtime.get("requires_handoff"))),
        runtime.get("risk_level") or data.get("risk_level") or "",
        path,
    ]


def product_row(path: str, item: dict[str, Any]) -> list[Any]:
    data = item_data(item)
    runtime = item.get("runtime") if isinstance(item.get("runtime"), dict) else {}
    product_id, section = product_path_parts(path)
    return [
        data.get("product_id") or product_id,
        section,
        item.get("id") or Path(path).stem,
        item.get("status") or "",
        human_title(item),
        human_content(item),
        join_value(data.get("keywords") or data.get("intent_tags") or data.get("aliases")),
        bool_label(data.get("allow_auto_reply", runtime.get("allow_auto_reply"))),
        bool_label(data.get("requires_handoff", runtime.get("requires_handoff"))),
        path,
    ]


def rag_source_rows(path: str, payload: Any) -> list[list[Any]]:
    values = payload if isinstance(payload, list) else payload.get("sources", []) if isinstance(payload, dict) else []
    rows = []
    for item in values:
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                item.get("source_id") or item.get("id") or "",
                item.get("category") or item.get("source_type") or "",
                item.get("product_id") or "",
                item.get("title") or Path(str(item.get("source_path") or "")).name,
                item.get("text") or item.get("summary") or item.get("content") or "",
                item.get("source_path") or path,
            ]
        )
    return rows


def rag_experience_rows(path: str, payload: Any) -> list[list[Any]]:
    values = payload if isinstance(payload, list) else payload.get("experiences", []) if isinstance(payload, dict) else []
    rows = []
    for item in values:
        if not isinstance(item, dict):
            continue
        hit = item.get("rag_hit") if isinstance(item.get("rag_hit"), dict) else {}
        usage = item.get("usage") if isinstance(item.get("usage"), dict) else {}
        quality = item.get("quality") if isinstance(item.get("quality"), dict) else {}
        rows.append(
            [
                item.get("experience_id") or item.get("id") or "",
                item.get("status") or "",
                item.get("question") or "",
                item.get("reply_text") or item.get("summary") or "",
                hit.get("text") or "",
                quality.get("band") or quality.get("score") or "",
                usage.get("reply_count") or "",
                path,
            ]
        )
    return rows


def write_intro(sheet: Worksheet, package: dict[str, Any], data: dict[str, list[list[Any]]]) -> None:
    rows = [
        ["文件用途", "给管理员阅读客户知识内容；不是备份还原包。"],
        ["客户账号", package.get("account_username") or package.get("tenant_id") or ""],
        ["数据包编号", package.get("package_id") or ""],
        ["备份编号", package.get("backup_id") or ""],
        ["正式知识条数", len(data["formal"])],
        ["商品专属知识条数", len(data["product"])],
        ["RAG资料条数", len(data["rag_sources"])],
        ["RAG经验条数", len(data["rag_experience"])],
        ["技术文件数", len(data["manifest"])],
    ]
    write_sheet_rows(sheet, ["项目", "内容"], rows)


def write_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title)
    write_sheet_rows(sheet, headers, rows)


def write_grouped_sheets(
    workbook: Workbook,
    prefix: str,
    headers: list[str],
    grouped_rows: dict[str, list[list[Any]]],
    labels: dict[str, str],
) -> None:
    ordered_keys = [key for key in labels if key in grouped_rows]
    ordered_keys.extend(sorted(key for key in grouped_rows if key not in labels))
    if not ordered_keys:
        write_sheet(workbook, prefix, headers, [])
        return
    used_names: set[str] = set(workbook.sheetnames)
    for key in ordered_keys:
        label = labels.get(key) or key
        title = unique_sheet_name(f"{prefix}-{label}", used_names)
        write_sheet(workbook, title, headers, grouped_rows.get(key, []))


def write_sheet_rows(sheet: Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in rows:
        sheet.append([cell_text(value) for value in row])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, header in enumerate(headers, start=1):
        width = 16
        if header in {"正文/回复", "正文/摘要", "回复要点", "命中资料", "来源文件", "文件路径"}:
            width = 48
        if header in {"标题/名称", "客户问法"}:
            width = 28
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width
    sheet.freeze_panes = "A2"


def unique_sheet_name(title: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]+", "-", title).strip()[:31] or "Sheet"
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"-{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def read_archive_json(archive: zipfile.ZipFile, name: str, *, default: Any) -> Any:
    try:
        return json.loads(archive.read(name).decode("utf-8"))
    except Exception:
        return default


def normalize_payload(payload: Any) -> dict[str, Any]:
    return payload if isinstance(payload, dict) else {"data": payload}


def item_data(item: dict[str, Any]) -> dict[str, Any]:
    data = item.get("data") if isinstance(item.get("data"), dict) else {}
    nested = data.get("data") if isinstance(data.get("data"), dict) else None
    return nested or data


def human_title(item: dict[str, Any]) -> str:
    data = item_data(item)
    for key in ("title", "name", "sku", "customer_message", "question", "summary"):
        if data.get(key):
            return str(data[key])
    return str(item.get("id") or item.get("item_id") or "")


def human_content(item: dict[str, Any]) -> str:
    data = item_data(item)
    parts = []
    for key in ("answer", "service_reply", "reply_text", "guideline_text", "content", "body", "specs", "shipping_policy", "warranty_policy"):
        if data.get(key):
            parts.append(f"{field_label(key)}：{join_value(data[key])}")
    if data.get("reply_templates"):
        parts.append(f"回复模板：{join_value(data['reply_templates'])}")
    if data.get("additional_details"):
        parts.append(f"补充信息：{join_value(data['additional_details'])}")
    return "\n".join(parts) or join_value(data)


def field_label(key: str) -> str:
    return {
        "answer": "答案",
        "service_reply": "客服回复",
        "reply_text": "回复要点",
        "guideline_text": "共享规则",
        "content": "正文",
        "body": "正文",
        "specs": "规格",
        "shipping_policy": "发货规则",
        "warranty_policy": "质保规则",
    }.get(key, key)


def product_path_parts(path: str) -> tuple[str, str]:
    parts = path.split("/")
    try:
        idx = parts.index("product_item_knowledge")
        return parts[idx + 1], parts[idx + 2] if len(parts) > idx + 2 else ""
    except (ValueError, IndexError):
        return "", ""


def category_from_formal_path(path: str) -> str:
    parts = path.split("/")
    try:
        idx = parts.index("knowledge_bases")
        return parts[idx + 1]
    except (ValueError, IndexError):
        return ""


def join_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    if isinstance(value, list):
        return "；".join(join_value(item) for item in value if item is not None)
    if isinstance(value, dict):
        pairs = []
        for key, item in value.items():
            pairs.append(f"{key}: {join_value(item)}")
        return "；".join(pairs)
    return str(value)


def bool_label(value: Any) -> str:
    if value is True:
        return "是"
    if value is False:
        return "否"
    return ""


def cell_text(value: Any) -> Any:
    text = join_value(value)
    return text[:MAX_CELL_TEXT] if isinstance(text, str) else text


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "customer_data"
