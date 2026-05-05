"""Local Excel exports for collected formal and reference knowledge."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .knowledge_base_store import KnowledgeBaseStore, product_scoped_category_records
from .knowledge_registry import KnowledgeRegistry
from apps.wechat_ai_customer_service.knowledge_paths import active_tenant_id, runtime_app_root
from apps.wechat_ai_customer_service.sync import BackupService
from apps.wechat_ai_customer_service.exports.readable_export import build_customer_readable_workbook


EXPORT_ROOT = runtime_app_root() / "admin" / "knowledge_exports"


class KnowledgeExportService:
    def __init__(self) -> None:
        self.registry = KnowledgeRegistry()
        self.base_store = KnowledgeBaseStore(self.registry)

    def build_export(self, *, sort_by: str = "type") -> dict[str, Any]:
        sort_by = "time" if sort_by == "time" else "type"
        if sort_by == "time":
            path = self.build_time_sorted_workbook()
            return {"ok": True, "sort_by": sort_by, "path": str(path), "filename": path.name}
        backup = BackupService(output_root=EXPORT_ROOT / "packages").build_backup(scope="tenant", tenant_id=active_tenant_id())
        package = {
            "package_id": f"local_data_pkg_{active_tenant_id()}_{backup['backup_id']}",
            "account_username": active_tenant_id(),
            "tenant_id": active_tenant_id(),
            "backup_id": backup["backup_id"],
            "manifest": backup["manifest"],
            "package_path": backup["package_path"],
        }
        path = build_customer_readable_workbook(package, Path(str(backup["package_path"])), output_root=EXPORT_ROOT / "readable")
        return {"ok": True, "sort_by": sort_by, "path": str(path), "filename": path.name, "backup": backup}

    def build_time_sorted_workbook(self) -> Path:
        EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
        path = EXPORT_ROOT / f"knowledge_time_sorted_{active_tenant_id()}_{timestamp_id()}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "正式知识-时间排序"
        rows = self.formal_knowledge_rows()
        rows.sort(key=lambda row: str(row[7] or row[6] or ""), reverse=True)
        write_rows(
            sheet,
            ["类型", "条目ID", "状态", "标题/名称", "正文/回复", "关键词", "创建时间", "更新时间", "新加入"],
            rows,
        )
        workbook.save(path)
        return path

    def formal_knowledge_rows(self) -> list[list[Any]]:
        category_ids = [str(item.get("id") or "") for item in self.registry.list_categories(enabled_only=True)]
        category_ids.extend(str(item.get("id") or "") for item in product_scoped_category_records())
        rows = []
        for category_id in sorted(set(item for item in category_ids if item)):
            try:
                items = self.base_store.list_items(category_id, include_archived=False)
            except FileNotFoundError:
                continue
            for item in items:
                data = item.get("data") if isinstance(item.get("data"), dict) else {}
                metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
                review_state = item.get("review_state") if isinstance(item.get("review_state"), dict) else {}
                rows.append(
                    [
                        category_id,
                        item.get("id") or "",
                        item.get("status") or "",
                        human_title(data, item),
                        human_content(data),
                        join_value(data.get("keywords") or data.get("aliases") or data.get("intent_tags")),
                        metadata.get("created_at") or "",
                        metadata.get("updated_at") or "",
                        "是" if review_state.get("is_new") else "",
                    ]
                )
        return rows


def write_rows(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF3")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column, value=value)
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + column)].width = 22


def human_title(data: dict[str, Any], item: dict[str, Any]) -> str:
    for key in ("name", "title", "question", "customer_message", "external_id"):
        value = data.get(key)
        if value:
            return str(value)
    return str(item.get("id") or "")


def human_content(data: dict[str, Any]) -> str:
    for key in ("answer", "content", "service_reply", "description", "specs"):
        value = data.get(key)
        if value:
            return join_value(value)
    return join_value(data.get("additional_details") or "")


def join_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "、".join(join_value(item) for item in value if join_value(item))
    if isinstance(value, dict):
        return "；".join(f"{key}: {join_value(inner)}" for key, inner in value.items() if join_value(inner))
    return str(value)


def timestamp_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")
