"""Raw upload persistence."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any

from .audit_log import append_audit
from apps.wechat_ai_customer_service.knowledge_paths import active_tenant_id, tenant_admin_upload_index_path, tenant_raw_inbox_root
from apps.wechat_ai_customer_service.storage import get_postgres_store, load_storage_config


APP_ROOT = Path(__file__).resolve().parents[2]
WORKFLOWS_ROOT = APP_ROOT / "workflows"
if str(WORKFLOWS_ROOT) not in sys.path:
    sys.path.insert(0, str(WORKFLOWS_ROOT))

from rag_layer import RagService  # noqa: E402
ALLOWED_KINDS = {"products", "chats", "policies", "erp_exports"}
ALLOWED_SUFFIXES = {".txt", ".md", ".json", ".csv", ".xlsx"}
SPREADSHEET_SUFFIXES = {".xlsx"}


class UploadStore:
    def __init__(self, *, tenant_id: str | None = None) -> None:
        self.tenant_id = active_tenant_id(tenant_id)
        self.raw_inbox_root = tenant_raw_inbox_root(self.tenant_id)
        self.index_path = tenant_admin_upload_index_path(self.tenant_id)

    def save_upload(self, filename: str, content: bytes, kind: str) -> dict[str, Any]:
        if kind not in ALLOWED_KINDS:
            return {"ok": False, "message": f"unsupported kind: {kind}"}
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_SUFFIXES:
            return {"ok": False, "message": f"unsupported suffix: {suffix}"}
        if not content.strip():
            return {"ok": False, "message": "empty upload"}

        stored_content = content
        stored_suffix = suffix
        normalized_from = ""
        if suffix in SPREADSHEET_SUFFIXES:
            converted = spreadsheet_to_text(content)
            if not converted.strip():
                return {"ok": False, "message": "spreadsheet has no readable cells"}
            stored_content = converted.encode("utf-8")
            stored_suffix = ".txt"
            normalized_from = suffix

        digest = hashlib.sha256(content).hexdigest()
        upload_id = "upload_" + digest[:16]
        safe_name = re.sub(r"[^A-Za-z0-9_.\-\u4e00-\u9fff]", "_", Path(filename).name)
        if stored_suffix != suffix:
            safe_name = safe_name + stored_suffix
        target_dir = self.raw_inbox_root / kind
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / f"{upload_id}_{safe_name}"
        target_path.write_bytes(stored_content)
        record = {
            "upload_id": upload_id,
            "filename": filename,
            "kind": kind,
            "path": str(target_path),
            "original_suffix": suffix,
            "stored_suffix": stored_suffix,
            "normalized_from": normalized_from,
            "sha256": digest,
            "size": len(stored_content),
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            "learned": False,
        }
        records = [item for item in self.list_uploads() if item.get("upload_id") != upload_id]
        records.append(record)
        db = postgres_store()
        config = load_storage_config()
        if db:
            db.upsert_upload(active_tenant_id(self.tenant_id), record)
            if not config.mirror_files:
                append_audit("upload_created", {"upload_id": upload_id, "kind": kind, "path": str(target_path)})
                return {"ok": True, "item": record}
        self.write_index(records)
        append_audit("upload_created", {"upload_id": upload_id, "kind": kind, "path": str(target_path)})
        return {"ok": True, "item": record}

    def list_uploads(self) -> list[dict[str, Any]]:
        db = postgres_store()
        if db:
            records = db.list_uploads(active_tenant_id(self.tenant_id))
            if records:
                return records
        if not self.index_path.exists():
            return []
        return json.loads(self.index_path.read_text(encoding="utf-8"))

    def get_upload(self, upload_id: str) -> dict[str, Any] | None:
        for item in self.list_uploads():
            if item.get("upload_id") == upload_id:
                return item
        return None

    def delete_upload(self, upload_id: str) -> dict[str, Any]:
        records = self.list_uploads()
        target = next((item for item in records if item.get("upload_id") == upload_id), None)
        if not target:
            return {"ok": False, "message": f"upload not found: {upload_id}"}
        remaining = [item for item in records if item.get("upload_id") != upload_id]
        deleted_file = False
        file_path = Path(str(target.get("path") or ""))
        try:
            resolved_file = file_path.resolve()
            raw_root = self.raw_inbox_root.resolve()
            if raw_root in resolved_file.parents and resolved_file.exists():
                RagService(tenant_id=self.tenant_id).delete_source_by_path(file_path)
                resolved_file.unlink()
                deleted_file = True
        except OSError:
            deleted_file = False
        db = postgres_store()
        config = load_storage_config()
        if db:
            db.delete_upload(active_tenant_id(self.tenant_id), upload_id)
            if not config.mirror_files:
                append_audit("upload_deleted", {"upload_id": upload_id, "path": str(file_path), "deleted_file": deleted_file})
                return {"ok": True, "item": target, "deleted_file": deleted_file}
        self.write_index(remaining)
        append_audit("upload_deleted", {"upload_id": upload_id, "path": str(file_path), "deleted_file": deleted_file})
        return {"ok": True, "item": target, "deleted_file": deleted_file}

    def mark_learned(self, upload_ids: list[str], candidate_ids: list[str]) -> None:
        records = self.list_uploads()
        for item in records:
            if item.get("upload_id") in upload_ids:
                item["learned"] = True
                item["candidate_ids"] = sorted(set([*item.get("candidate_ids", []), *candidate_ids]))
                item["learned_at"] = datetime.now().isoformat(timespec="seconds")
                db = postgres_store()
                if db:
                    db.upsert_upload(active_tenant_id(self.tenant_id), item)
        if postgres_store() and not load_storage_config().mirror_files:
            return
        self.write_index(records)

    def write_index(self, records: list[dict[str, Any]]) -> None:
        self.index_path.parent.mkdir(parents=True, exist_ok=True)
        self.index_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def spreadsheet_to_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for xlsx uploads") from exc
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [cell_to_text(value) for value in row]
            if any(values):
                lines.append(",".join(values))
    return "\n".join(lines).strip() + "\n"


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def postgres_store():
    config = load_storage_config()
    if not config.use_postgres or not config.postgres_configured:
        return None
    store = get_postgres_store(config=config)
    return store if store.available() else None
