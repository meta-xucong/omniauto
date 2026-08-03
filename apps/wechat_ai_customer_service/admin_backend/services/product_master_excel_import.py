"""Deterministic Excel import for local/manual V2 vehicle records.

This module is intentionally local-only: it does not call Dafengche APIs,
network transports, LLM helpers, schedulers, Brain code, or RPA adapters.
Excel rows are converted into the same manual V2 vehicle envelope used by the
existing admin edit path.
"""

from __future__ import annotations

import copy
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from apps.wechat_ai_customer_service.product_master import (
    SAFE_PRODUCT_ID_RE,
    ProductMasterStore,
    normalize_product_item,
    validate_product_item,
)
from packages.dafengche_product_master import apply_admin_vehicle_update, create_manual_vehicle
from packages.dafengche_product_master.contract import VEHICLE_DETAIL_FIELD_GROUP_SPECS


LEGACY_TEMPLATE_VERSION = "local_manual_vehicle_v2_excel_20260801"
HORIZONTAL_TEMPLATE_VERSION = "local_manual_vehicle_v2_excel_20260801_admin_fields"
PATH_VERTICAL_TEMPLATE_VERSION = "local_manual_vehicle_v2_excel_20260801_vertical"
SIMPLE_VERTICAL_TEMPLATE_VERSION = "local_manual_vehicle_v2_excel_20260801_simple_vertical"
TEMPLATE_VERSION = "local_manual_vehicle_v2_excel_20260802_simple_raw_fields"
SUPPORTED_TEMPLATE_VERSIONS = {LEGACY_TEMPLATE_VERSION, HORIZONTAL_TEMPLATE_VERSION, PATH_VERTICAL_TEMPLATE_VERSION, SIMPLE_VERTICAL_TEMPLATE_VERSION, TEMPLATE_VERSION}
VEHICLE_SHEET = "车辆信息"
PICTURE_SHEET = "车辆图片"
FIELD_SHEET = "字段说明"
MAX_IMPORT_ROWS = 2000
MAX_PICTURE_ROWS = 5000
MAX_EXCEL_BYTES = 8 * 1024 * 1024
PATH_VERTICAL_VEHICLE_HEADERS = ("local_id", "字段路径", "字段值", "填写说明")
VERTICAL_VEHICLE_ID_HEADER = "车辆编号 *"
VERTICAL_VEHICLE_ID_LEGACY_HEADER = "车辆编号"
VERTICAL_VEHICLE_HEADERS = (VERTICAL_VEHICLE_ID_HEADER, "填写项", "填写内容", "填写说明")
VERTICAL_VEHICLE_HEADER_ALIASES = {VERTICAL_VEHICLE_ID_LEGACY_HEADER: VERTICAL_VEHICLE_ID_HEADER}


@dataclass(frozen=True)
class ColumnSpec:
    header: str
    target: str
    required: bool = False
    value_type: str = "text"
    note: str = ""


VEHICLE_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("local_id", "local_id", True, "id", "本地车辆唯一编号；不会伪造成 shopCode/carId。"),
    ColumnSpec("车辆标题 baseCarInfo.name.displayValue", "detail.baseCarInfo.name.displayValue", True, "text", "客户和管理台看到的车辆标题。"),
    ColumnSpec("品牌 baseCarInfo.name.brandName", "detail.baseCarInfo.name.brandName"),
    ColumnSpec("车系 baseCarInfo.name.seriesName", "detail.baseCarInfo.name.seriesName"),
    ColumnSpec("车型 baseCarInfo.name.modelName", "detail.baseCarInfo.name.modelName"),
    ColumnSpec("车辆名称 baseCarInfo.carName", "detail.baseCarInfo.carName"),
    ColumnSpec("上牌时间 baseCarInfo.firstLicensePlateDate", "detail.baseCarInfo.firstLicensePlateDate"),
    ColumnSpec("表显里程 baseCarInfo.mileage", "detail.baseCarInfo.mileage"),
    ColumnSpec("车况 baseCarInfo.vehicleCondition", "detail.baseCarInfo.vehicleCondition"),
    ColumnSpec("外观颜色 baseCarInfo.exteriorColor", "detail.baseCarInfo.exteriorColor"),
    ColumnSpec("车身颜色备用 baseCarInfo.color", "detail.baseCarInfo.color"),
    ColumnSpec("内饰颜色 baseCarInfo.interiorColor", "detail.baseCarInfo.interiorColor"),
    ColumnSpec("变速箱 carModelParam.gearbox", "detail.carModelParam.gearbox"),
    ColumnSpec("变速箱备用 carModelParam.gearBox", "detail.carModelParam.gearBox"),
    ColumnSpec("排量 carModelParam.displacement", "detail.carModelParam.displacement"),
    ColumnSpec("燃料 carModelParam.fuelType", "detail.carModelParam.fuelType"),
    ColumnSpec("排放 carModelParam.emissionStandard", "detail.carModelParam.emissionStandard"),
    ColumnSpec("座位数 carModelParam.seatNumber", "detail.carModelParam.seatNumber", False, "number"),
    ColumnSpec("手续状态 carLicenseInfo.licenseStatus", "detail.carLicenseInfo.licenseStatus"),
    ColumnSpec("公开售价 carPriceInfo.salePrice", "detail.carPriceInfo.salePrice", False, "number"),
    ColumnSpec("收购价 carPriceInfo.purchasePrice", "detail.carPriceInfo.purchasePrice", False, "number"),
    ColumnSpec("成交价 carPriceInfo.salesPrice", "detail.carPriceInfo.salesPrice", False, "number"),
    ColumnSpec("经理价 carPriceInfo.managerPrice", "detail.carPriceInfo.managerPrice", False, "number"),
    ColumnSpec("批发价 carPriceInfo.wholesalePrice", "detail.carPriceInfo.wholesalePrice", False, "number"),
    ColumnSpec("新车指导价 carPriceInfo.newPrice", "detail.carPriceInfo.newPrice", False, "number"),
    ColumnSpec("销售状态 operationPhase", "detail.operationPhase"),
    ColumnSpec("本地类目 annotations.category", "annotations.category"),
    ColumnSpec("客户常用叫法 annotations.aliases", "annotations.aliases", False, "list"),
    ColumnSpec("客户可见卖点 annotations.specs", "annotations.specs"),
    ColumnSpec("物流/看车政策 annotations.shipping_policy", "annotations.shipping_policy"),
    ColumnSpec("售后/质保 annotations.warranty_policy", "annotations.warranty_policy"),
    ColumnSpec("风险/禁用承诺 annotations.risk_rules", "annotations.risk_rules", False, "list"),
    ColumnSpec("补充信息 annotations.additional_details", "annotations.additional_details"),
    ColumnSpec("内部SKU manual.sku", "manual.sku"),
    ColumnSpec("本地单位 manual.unit", "manual.unit"),
    ColumnSpec("库存数量 manual.inventory", "manual.inventory", False, "integer"),
    ColumnSpec("默认回复 manual.reply_templates.default", "manual.reply_templates.default"),
    ColumnSpec("报价回复 manual.reply_templates.quote", "manual.reply_templates.quote"),
    ColumnSpec("议价政策 manual.reply_templates.discount_policy", "manual.reply_templates.discount_policy"),
    ColumnSpec("物流回复 manual.reply_templates.logistics", "manual.reply_templates.logistics"),
    ColumnSpec("售后回复 manual.reply_templates.after_sales", "manual.reply_templates.after_sales"),
    ColumnSpec("内部备注 manual.reply_templates.notes", "manual.reply_templates.notes"),
)

_RAW_TEMPLATE_SYSTEM_PATHS = frozenset({"carId", "orgId", "shopCode", "owner", "creator"})
_RAW_TEMPLATE_OBJECT_PATHS = frozenset(
    spec.path
    for _group_id, _group_label, specs in VEHICLE_DETAIL_FIELD_GROUP_SPECS
    for spec in specs
    if any(
        other.path.startswith(f"{spec.path}.")
        for _other_group_id, _other_group_label, other_specs in VEHICLE_DETAIL_FIELD_GROUP_SPECS
        for other in other_specs
    )
)
_RAW_TEMPLATE_TECHNICAL_CODE_PATHS = frozenset(
    {
        "baseCarInfo.name.brandCode",
        "baseCarInfo.name.seriesCode",
        "baseCarInfo.name.modelCode",
        "baseCarInfo.area.cityCode",
        "baseCarInfo.area.provinceCode",
        "baseCarInfo.registerArea.cityCode",
        "baseCarInfo.registerArea.provinceCode",
    }
)


def _raw_vehicle_template_columns() -> tuple[ColumnSpec, ...]:
    columns = [ColumnSpec("local_id", "local_id", True, "id", "本地车辆唯一编号；不会伪造成 shopCode/carId。")]
    for _group_id, group_label, specs in VEHICLE_DETAIL_FIELD_GROUP_SPECS:
        for spec in specs:
            if spec.path in _RAW_TEMPLATE_SYSTEM_PATHS or spec.path in _RAW_TEMPLATE_OBJECT_PATHS or spec.path in _RAW_TEMPLATE_TECHNICAL_CODE_PATHS:
                continue
            columns.append(
                ColumnSpec(
                    f"{group_label}｜{spec.label}",
                    f"detail.{spec.path}",
                    spec.path == "baseCarInfo.name.displayValue",
                    _raw_template_value_type(spec.path),
                    _raw_template_field_note(spec.path),
                )
            )
    return tuple(columns)

def _raw_template_value_type(path: str) -> str:
    text = str(path or "")
    if text.startswith("carPriceInfo."):
        return "number"
    if text in {
        "carModelParam.seatNumber",
        "carLicenseInfo.keysCount",
        "carLicenseInfo.transferTotal",
        "carLicenseInfo.xiancheshangyexianjine",
    }:
        return "number"
    return "text"


def _raw_template_field_note(path: str) -> str:
    if path == "baseCarInfo.name.displayValue":
        return "必填；管理台和客服证据使用的车辆展示名称。"
    if path.startswith("carPriceInfo."):
        return "填写数字；内部底价类字段会保存在本地原始记录中，但不会进入客户可见证据。"
    if path.startswith("carOwnerInfo.") or path in {"baseCarInfo.vinNumber", "baseCarInfo.plateNumber", "carModelParam.engineNumber"}:
        return "原始受限字段；可维护在本地记录中，默认不会进入客户可见证据。"
    return "大风车原始车辆字段；留空不会覆盖已有值。"


RAW_VEHICLE_TEMPLATE_COLUMNS: tuple[ColumnSpec, ...] = _raw_vehicle_template_columns()

LEGACY_VEHICLE_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("local_id", "local_id", True, "id", "本地车辆唯一编号；不会伪造成 shopCode/carId。"),
    ColumnSpec("车辆标题 baseCarInfo.name.displayValue", "detail.baseCarInfo.name.displayValue", True, "text", "客户和管理台看到的车辆标题。"),
    ColumnSpec("品牌 baseCarInfo.name.brandName", "detail.baseCarInfo.name.brandName"),
    ColumnSpec("车系 baseCarInfo.name.seriesName", "detail.baseCarInfo.name.seriesName"),
    ColumnSpec("车型 baseCarInfo.name.modelName", "detail.baseCarInfo.name.modelName"),
    ColumnSpec("上牌时间 baseCarInfo.firstLicensePlateDate", "detail.baseCarInfo.firstLicensePlateDate"),
    ColumnSpec("表显里程 baseCarInfo.mileage", "detail.baseCarInfo.mileage"),
    ColumnSpec("车况 baseCarInfo.vehicleCondition", "detail.baseCarInfo.vehicleCondition"),
    ColumnSpec("外观颜色 baseCarInfo.color", "detail.baseCarInfo.color"),
    ColumnSpec("内饰颜色 baseCarInfo.innerColor", "detail.baseCarInfo.innerColor"),
    ColumnSpec("变速箱 carModelParam.gearBoxType", "detail.carModelParam.gearBoxType"),
    ColumnSpec("排量 carModelParam.engineVolumeLiter", "detail.carModelParam.engineVolumeLiter"),
    ColumnSpec("燃料 carModelParam.fuelType", "detail.carModelParam.fuelType"),
    ColumnSpec("排放 carModelParam.emissionStandard", "detail.carModelParam.emissionStandard"),
    ColumnSpec("座位数 carModelParam.seatNumber", "detail.carModelParam.seatNumber", False, "number"),
    ColumnSpec("公开售价 carPriceInfo.salePrice", "detail.carPriceInfo.salePrice", False, "number"),
    ColumnSpec("新车指导价 carPriceInfo.newPrice", "detail.carPriceInfo.newPrice", False, "number"),
    ColumnSpec("销售状态 operationPhase", "detail.operationPhase"),
    ColumnSpec("客户常用叫法 annotations.aliases", "annotations.aliases", False, "list"),
    ColumnSpec("客户可见卖点 annotations.specs", "annotations.specs"),
    ColumnSpec("物流/看车政策 annotations.shipping_policy", "annotations.shipping_policy"),
    ColumnSpec("售后/质保 annotations.warranty_policy", "annotations.warranty_policy"),
    ColumnSpec("风险/禁用承诺 annotations.risk_rules", "annotations.risk_rules", False, "list"),
    ColumnSpec("补充信息 annotations.additional_details", "annotations.additional_details"),
    ColumnSpec("内部SKU manual.sku", "manual.sku"),
    ColumnSpec("库存数量 manual.inventory", "manual.inventory", False, "integer"),
    ColumnSpec("默认回复 manual.reply_templates.default", "manual.reply_templates.default"),
)

PICTURE_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("local_id", "local_id", True, "id", "必须匹配车辆信息表中的 local_id。"),
    ColumnSpec("图片URL pictureUrl", "pictureUrl", False, "text", "Excel 只导入远程 URL；本地图片文件请使用管理台图片上传入口。"),
    ColumnSpec("大图URL bigPictureUrl", "bigPictureUrl"),
    ColumnSpec("排序 sort", "sort", False, "integer"),
    ColumnSpec("图片说明 description", "description"),
)

SIMPLE_PICTURE_COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("车辆编号", "local_id", True, "id", "必须匹配车辆信息表中的车辆编号。"),
    ColumnSpec("图片地址", "pictureUrl", False, "text", "Excel 只填写图片网址；本地图片文件请使用管理台图片上传入口。"),
    ColumnSpec("大图地址", "bigPictureUrl"),
    ColumnSpec("排序", "sort", False, "integer"),
    ColumnSpec("图片说明", "description"),
)

_SIMPLE_FIELD_LABELS_BY_TARGET: dict[str, str] = {
    "detail.baseCarInfo.name.displayValue": "车辆标题",
    "detail.baseCarInfo.name.brandName": "品牌",
    "detail.baseCarInfo.name.seriesName": "车系",
    "detail.baseCarInfo.name.modelName": "车型",
    "detail.baseCarInfo.carName": "车辆名称",
    "detail.baseCarInfo.firstLicensePlateDate": "首次上牌",
    "detail.baseCarInfo.mileage": "表显里程",
    "detail.baseCarInfo.vehicleCondition": "车况",
    "detail.baseCarInfo.exteriorColor": "外观颜色",
    "detail.baseCarInfo.color": "车身颜色",
    "detail.baseCarInfo.interiorColor": "内饰颜色",
    "detail.carModelParam.gearbox": "变速箱",
    "detail.carModelParam.gearBox": "变速箱备用",
    "detail.carModelParam.displacement": "排量",
    "detail.carModelParam.fuelType": "燃料",
    "detail.carModelParam.emissionStandard": "排放标准",
    "detail.carModelParam.seatNumber": "座位数",
    "detail.carLicenseInfo.licenseStatus": "手续状态",
    "detail.carPriceInfo.salePrice": "公开售价",
    "detail.carPriceInfo.purchasePrice": "收购价",
    "detail.carPriceInfo.salesPrice": "成交价",
    "detail.carPriceInfo.managerPrice": "经理价",
    "detail.carPriceInfo.wholesalePrice": "批发价",
    "detail.carPriceInfo.newPrice": "新车指导价",
    "detail.operationPhase": "业务阶段",
    "annotations.category": "本地类目",
    "annotations.aliases": "客户常用叫法",
    "annotations.specs": "客户可见卖点",
    "annotations.shipping_policy": "看车/交付说明",
    "annotations.warranty_policy": "售后/合同口径",
    "annotations.risk_rules": "风险/禁用承诺",
    "annotations.additional_details": "补充信息",
    "manual.sku": "内部编号",
    "manual.unit": "本地单位",
    "manual.inventory": "库存数量",
    "manual.reply_templates.default": "默认回复",
    "manual.reply_templates.quote": "报价回复",
    "manual.reply_templates.discount_policy": "议价回复",
    "manual.reply_templates.logistics": "物流回复",
    "manual.reply_templates.after_sales": "售后回复",
    "manual.reply_templates.notes": "内部备注",
}


class ProductMasterExcelImportService:
    """Excel template, preview and confirmed import facade."""

    def __init__(self, store: ProductMasterStore) -> None:
        self.store = store

    @property
    def preview_root(self) -> Path:
        return self.store.root / ".excel_import_previews"

    def template_bytes(self) -> bytes:
        return build_template_bytes()

    def preview(self, *, filename: str, content: bytes) -> dict[str, Any]:
        parsed = parse_vehicle_workbook(content, store=self.store)
        preview_id = f"xlsx_{uuid4().hex}"
        artifact = {
            "preview_id": preview_id,
            "filename": str(filename or "vehicles.xlsx"),
            "tenant_id": self.store.tenant_id,
            "template_version": parsed["template_version"],
            "ok": parsed["ok"],
            "created_at": _now_iso(),
            "summary": parsed["summary"],
            "vehicles": parsed["vehicles"],
            "errors": parsed["errors"],
            "records": parsed["records"],
        }
        self.preview_root.mkdir(parents=True, exist_ok=True)
        _write_json(self.preview_root / f"{preview_id}.json", artifact)
        return _public_preview(artifact)

    def confirm(self, *, preview_id: str) -> dict[str, Any]:
        clean_preview_id = str(preview_id or "").strip()
        if not re.fullmatch(r"xlsx_[a-f0-9]{32}", clean_preview_id):
            raise ValueError("invalid preview_id")
        artifact = _read_json(self.preview_root / f"{clean_preview_id}.json")
        if not isinstance(artifact, dict) or artifact.get("tenant_id") != self.store.tenant_id:
            raise FileNotFoundError(clean_preview_id)
        if not artifact.get("ok"):
            raise ValueError("cannot confirm an Excel import preview with errors")
        records = artifact.get("records") if isinstance(artifact.get("records"), list) else []
        schema = self.store.load_schema()
        validation_errors: list[dict[str, Any]] = []
        normalized_records: list[dict[str, Any]] = []
        for index, record in enumerate(records, start=1):
            normalized = normalize_product_item(record if isinstance(record, dict) else {})
            validation = validate_product_item(normalized, schema)
            if not validation.get("ok"):
                validation_errors.append(
                    {
                        "row_number": _record_row_number(record),
                        "local_id": str((record or {}).get("id") or ""),
                        "field": "record",
                        "code": "record_validation_failed",
                        "message": "；".join(str(item) for item in validation.get("problems") or []) or f"record #{index} failed validation",
                    }
                )
            normalized_records.append(normalized)
        if validation_errors:
            raise ValueError(json.dumps(validation_errors, ensure_ascii=False))

        batch_result = self.store._save_items_batch_atomic(normalized_records)
        if not batch_result.get("ok"):
            raise ValueError(json.dumps(batch_result.get("problems") or batch_result, ensure_ascii=False))
        saved_ids = [str(item.get("id") or "") for item in batch_result.get("items") or normalized_records]
        self.store.write_manifest(extra={"last_manual_excel_import_at": _now_iso(), "last_manual_excel_import_count": len(saved_ids)})
        result = {
            "ok": True,
            "preview_id": clean_preview_id,
            "imported_count": len(saved_ids),
            "saved_ids": saved_ids,
            "mode": "local_manual_vehicle_excel",
            "storage": str(batch_result.get("storage") or ""),
        }
        mirror = batch_result.get("mirror_files") if isinstance(batch_result.get("mirror_files"), dict) else {}
        if mirror and mirror.get("ok") is False:
            result["warnings"] = [
                {
                    "code": "mirror_files_failed",
                    "message": "PostgreSQL import succeeded; file mirror rebuild is required.",
                    "failed_ids": list(mirror.get("failed_ids") or []),
                }
            ]
            result["mirror_files"] = copy.deepcopy(mirror)
        return result


def build_template_bytes() -> bytes:
    workbook = Workbook()
    vehicle_sheet = workbook.active
    vehicle_sheet.title = VEHICLE_SHEET
    vehicle_sheet.append(list(VERTICAL_VEHICLE_HEADERS))
    for spec in RAW_VEHICLE_TEMPLATE_COLUMNS:
        if spec.target == "local_id":
            continue
        vehicle_sheet.append(["", _simple_field_label(spec), "", _vertical_field_instruction(spec)])
    field_sheet = workbook.create_sheet(FIELD_SHEET)
    field_sheet.append(["模板版本", TEMPLATE_VERSION])
    field_sheet.append(
        [
            "填写说明",
            "车辆信息表为竖向填写：第 1 行四个表头请勿改名；按车辆编号填写整组字段行；新车可复制整组字段行后填写新的车辆编号。填写内容留空不会覆盖既有手工字段。Excel 只维护原始车辆字段；本地客服扩展请在管理台高级选项中维护。车辆图片请在管理台车辆编辑页上传，可一次选择多张。",
            "",
        ]
    )
    field_sheet.append(["填写项", "是否必填", "填写说明"])
    field_sheet.append(["车辆编号", "是", "每组字段都填写同一个车辆编号，用于识别车辆。"])
    for spec in RAW_VEHICLE_TEMPLATE_COLUMNS:
        if spec.target == "local_id":
            continue
        field_sheet.append([_simple_field_label(spec), "是" if spec.required else "否", _vertical_field_instruction(spec)])
    _style_template_workbook(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BODY_FONT = Font(color="1F2937", size=10)
_MUTED_FONT = Font(color="4B5563", size=10)
_GROUP_HEADER_FILLS = {
    "identity": "374151",
    "base": "1F4E78",
    "model": "376092",
    "price": "7F6000",
    "state": "44546A",
    "annotations": "548235",
    "manual": "7030A0",
    "picture_identity": "374151",
    "picture_url": "1F4E78",
    "picture_order": "7F6000",
    "picture_note": "548235",
}
_GROUP_SAMPLE_FILLS = {
    "identity": "F3F4F6",
    "base": "EAF3F8",
    "model": "EEF3FA",
    "price": "FFF2CC",
    "state": "F2F4F7",
    "annotations": "EAF4E2",
    "manual": "F3EAF8",
    "picture_identity": "F3F4F6",
    "picture_url": "EAF3F8",
    "picture_order": "FFF2CC",
    "picture_note": "EAF4E2",
}


def _style_template_workbook(workbook: Workbook) -> None:
    _style_vehicle_sheet(workbook[VEHICLE_SHEET])
    if PICTURE_SHEET in workbook.sheetnames:
        _style_picture_sheet(workbook[PICTURE_SHEET])
    _style_field_sheet(workbook[FIELD_SHEET])


def _style_vehicle_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.row_dimensions[1].height = 46
    for column_index, cell in enumerate(sheet[1], start=1):
        _apply_header_style(cell, "identity" if column_index == 1 else "base")
    widths = {1: 20, 2: 30, 3: 46, 4: 74}
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    path_specs = _vehicle_field_specs_by_target(RAW_VEHICLE_TEMPLATE_COLUMNS)
    label_specs = _vehicle_field_specs_by_label(RAW_VEHICLE_TEMPLATE_COLUMNS)
    simple_vertical = str(sheet.cell(row=1, column=2).value or "").strip() == "填写项"
    for row_index in range(2, sheet.max_row + 1):
        field_key = str(sheet.cell(row=row_index, column=2).value or "")
        spec = label_specs.get(field_key) if simple_vertical else path_specs.get(field_key)
        group = _vehicle_style_group(spec.target if spec else field_key)
        sheet.row_dimensions[row_index].height = 32 if row_index > 2 else 38
        for column_index, cell in enumerate(sheet[row_index], start=1):
            _apply_sample_style(cell, group, spec if column_index == 3 else None)
            if column_index == 2:
                cell.font = Font(color="111827", bold=True, size=10)
            if column_index == 3:
                cell.fill = PatternFill("solid", fgColor="FFF8E1")
            if column_index == 4:
                cell.font = _MUTED_FONT


def _style_picture_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 46
    sheet.row_dimensions[2].height = 46
    header_map = {spec.header: spec for spec in (*PICTURE_COLUMNS, *SIMPLE_PICTURE_COLUMNS)}
    for column_index, cell in enumerate(sheet[1], start=1):
        spec = header_map.get(str(cell.value or ""))
        group = _picture_style_group(spec.target if spec else "")
        _apply_header_style(cell, group)
        sheet.column_dimensions[get_column_letter(column_index)].width = _picture_column_width(spec.target if spec else "")
    for column_index, cell in enumerate(sheet[2], start=1):
        spec = header_map.get(str(sheet.cell(row=1, column=column_index).value or ""))
        group = _picture_style_group(spec.target if spec else "")
        _apply_sample_style(cell, group, spec)


def _style_field_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[1].hidden = True
    sheet.row_dimensions[2].height = 90
    sheet.row_dimensions[3].height = 34
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 78
    sheet.print_title_rows = "3:3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    sheet["A1"].comment = Comment("模板版本由系统读取，请不要修改 A1/B1。", "OmniAuto")
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.font = _MUTED_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = _THIN_BORDER
    for cell in sheet[3]:
        cell.fill = PatternFill("solid", fgColor="374151")
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    for row_index in range(4, sheet.max_row + 1):
        fill = "FFFFFF" if row_index % 2 else "F9FAFB"
        sheet.row_dimensions[row_index].height = 30
        for cell in sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER


def _apply_header_style(cell: Any, group: str) -> None:
    cell.fill = PatternFill("solid", fgColor=_GROUP_HEADER_FILLS.get(group, "374151"))
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _apply_sample_style(cell: Any, group: str, spec: ColumnSpec | None) -> None:
    cell.fill = PatternFill("solid", fgColor=_GROUP_SAMPLE_FILLS.get(group, "F9FAFB"))
    cell.font = _BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = _THIN_BORDER
    if spec and spec.value_type == "integer":
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    elif spec and spec.value_type == "number":
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)


def _vehicle_style_group(target: str) -> str:
    if target == "local_id" or not target:
        return "identity"
    if target.startswith("detail.baseCarInfo."):
        return "base"
    if target.startswith("detail.carModelParam."):
        return "model"
    if target.startswith("detail.carPriceInfo."):
        return "price"
    if target.startswith("detail.carLicenseInfo.") or target == "detail.operationPhase":
        return "state"
    if target.startswith("annotations."):
        return "annotations"
    if target.startswith("manual."):
        return "manual"
    return "identity"


def _picture_style_group(target: str) -> str:
    if target == "local_id":
        return "picture_identity"
    if target in {"pictureUrl", "bigPictureUrl"}:
        return "picture_url"
    if target == "sort":
        return "picture_order"
    return "picture_note"


def _vehicle_column_width(target: str, header: str) -> float:
    if target == "local_id":
        return 18
    if target in {"detail.baseCarInfo.name.displayValue", "annotations.additional_details"}:
        return 30
    if target.startswith("manual.reply_templates."):
        return 34
    if target in {"annotations.specs", "annotations.shipping_policy", "annotations.warranty_policy", "annotations.risk_rules"}:
        return 28
    if target.startswith("detail.carPriceInfo."):
        return 15
    return min(24, max(14, len(header) * 0.9))


def _picture_column_width(target: str) -> float:
    if target == "local_id":
        return 18
    if target in {"pictureUrl", "bigPictureUrl"}:
        return 44
    if target == "sort":
        return 12
    return 28


def _vertical_sample_values() -> dict[str, Any]:
    return {
        "detail.baseCarInfo.name.displayValue": "丰田 凯美瑞 2.0G",
        "detail.baseCarInfo.name.brandName": "丰田",
        "detail.baseCarInfo.name.seriesName": "凯美瑞",
        "detail.baseCarInfo.name.modelName": "2.0G",
        "detail.baseCarInfo.carName": "凯美瑞",
        "detail.baseCarInfo.firstLicensePlateDate": "2020-06",
        "detail.baseCarInfo.mileage": "4.2万公里",
        "detail.baseCarInfo.vehicleCondition": "车况精品",
        "detail.baseCarInfo.exteriorColor": "白色",
        "detail.baseCarInfo.color": "白色",
        "detail.baseCarInfo.interiorColor": "黑色",
        "detail.carModelParam.gearbox": "自动",
        "detail.carModelParam.gearBox": "AT",
        "detail.carModelParam.displacement": "2.0L",
        "detail.carModelParam.fuelType": "汽油",
        "detail.carModelParam.emissionStandard": "国六",
        "detail.carModelParam.seatNumber": 5,
        "detail.carLicenseInfo.licenseStatus": "手续齐全",
        "detail.carPriceInfo.salePrice": 13.98,
        "detail.carPriceInfo.purchasePrice": 10.88,
        "detail.carPriceInfo.salesPrice": 13.50,
        "detail.carPriceInfo.managerPrice": 13.30,
        "detail.carPriceInfo.wholesalePrice": 12.80,
        "detail.carPriceInfo.newPrice": 19.98,
        "detail.operationPhase": "SALE",
        "annotations.category": "used_car",
        "annotations.aliases": "凯美瑞；丰田凯美瑞",
        "annotations.specs": "一手车，保养记录完整",
        "annotations.shipping_policy": "到店看车请提前预约",
        "annotations.warranty_policy": "以门店质保政策为准",
        "annotations.risk_rules": "不得承诺事故水泡火烧以外未核验事项",
        "annotations.additional_details": "适合家用代步",
        "manual.sku": "CAMRY-001",
        "manual.unit": "台",
        "manual.inventory": 1,
        "manual.reply_templates.default": "这台车目前在售，建议先预约到店看车。",
        "manual.reply_templates.quote": "这台目前公开报价 13.98 万，具体成交以到店确认为准。",
        "manual.reply_templates.discount_policy": "议价需要看客户付款方式和置换情况，不能提前承诺底价。",
        "manual.reply_templates.logistics": "看车前建议提前预约，方便安排车辆和销售接待。",
        "manual.reply_templates.after_sales": "售后以门店合同和实际质保政策为准。",
        "manual.reply_templates.notes": "内部备注仅供客服配置，不进入客户证据。",
    }


def _vertical_field_instruction(spec: ColumnSpec) -> str:
    parts: list[str] = []
    if spec.required:
        parts.append("必填")
    if spec.value_type == "list":
        parts.append("可用分号、逗号或换行分隔多项")
    elif spec.value_type == "number":
        parts.append("填写数字")
    elif spec.value_type == "integer":
        parts.append("填写整数")
    note = str(spec.note or "").strip()
    if spec.required and note.startswith("必填；"):
        note = note[len("必填；") :]
    if spec.value_type in {"number", "integer"} and note.startswith("填写数字；"):
        note = note[len("填写数字；") :]
    if spec.value_type == "integer" and note.startswith("填写整数；"):
        note = note[len("填写整数；") :]
    if note:
        parts.append(note)
    return "；".join(parts) or "留空不会覆盖已有值。"


def parse_vehicle_workbook(content: bytes, *, store: ProductMasterStore) -> dict[str, Any]:
    if not content:
        raise ValueError("Excel file is empty")
    if len(content) > MAX_EXCEL_BYTES:
        raise ValueError(f"Excel file is too large; max {MAX_EXCEL_BYTES // 1024 // 1024} MiB")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("invalid Excel workbook; please upload an .xlsx file generated from the current template") from exc
    if VEHICLE_SHEET not in workbook.sheetnames:
        raise ValueError(f"missing required sheet: {VEHICLE_SHEET}")
    if FIELD_SHEET not in workbook.sheetnames:
        raise ValueError(f"missing required sheet: {FIELD_SHEET}")
    template_version = _template_version(workbook)
    if template_version not in SUPPORTED_TEMPLATE_VERSIONS:
        raise ValueError(f"unsupported Excel template version: {template_version or 'missing'}")
    vehicle_sheet = workbook[VEHICLE_SHEET]
    if template_version == TEMPLATE_VERSION:
        vehicle_rows, vehicle_errors = _parse_vehicle_rows_simple_vertical(vehicle_sheet, columns=RAW_VEHICLE_TEMPLATE_COLUMNS)
        picture_columns = SIMPLE_PICTURE_COLUMNS
    elif template_version == SIMPLE_VERTICAL_TEMPLATE_VERSION:
        vehicle_rows, vehicle_errors = _parse_vehicle_rows_simple_vertical(vehicle_sheet, columns=VEHICLE_COLUMNS)
        picture_columns = SIMPLE_PICTURE_COLUMNS
    elif template_version == PATH_VERTICAL_TEMPLATE_VERSION:
        vehicle_rows, vehicle_errors = _parse_vehicle_rows_path_vertical(vehicle_sheet)
        picture_columns = PICTURE_COLUMNS
    else:
        vehicle_columns = _vehicle_columns_for_version(template_version)
        vehicle_rows, vehicle_errors = _parse_vehicle_rows(vehicle_sheet, vehicle_columns)
        picture_columns = PICTURE_COLUMNS
    if PICTURE_SHEET in workbook.sheetnames:
        picture_rows, picture_errors = _parse_picture_rows(
            workbook[PICTURE_SHEET],
            known_ids={row["local_id"] for row in vehicle_rows},
            columns=picture_columns,
        )
    elif template_version in {TEMPLATE_VERSION, SIMPLE_VERTICAL_TEMPLATE_VERSION}:
        picture_rows, picture_errors = [], []
    else:
        picture_rows, picture_errors = [], [_row_error(1, PICTURE_SHEET, "missing_sheet", f"missing required sheet: {PICTURE_SHEET}")]
    errors = [*vehicle_errors, *picture_errors]
    duplicate_ids = _duplicates(row["local_id"] for row in vehicle_rows)
    for local_id in duplicate_ids:
        rows = [row["row_number"] for row in vehicle_rows if row["local_id"] == local_id]
        errors.append(_row_error(rows[0], "local_id", "duplicate_local_id", f"local_id 在车辆信息表中重复：{local_id}"))
    pictures_by_id: dict[str, list[dict[str, Any]]] = {}
    for picture in picture_rows:
        pictures_by_id.setdefault(picture["local_id"], []).append(picture)

    records: list[dict[str, Any]] = []
    public_rows: list[dict[str, Any]] = []
    observed_at = _now_iso()
    for row in vehicle_rows:
        local_id = row["local_id"]
        existing = store.get_item(local_id, include_archived=True)
        if existing:
            source = existing.get("source") if isinstance(existing.get("source"), dict) else {}
            if str(source.get("type") or "") != "manual":
                errors.append(_row_error(row["row_number"], "local_id", "existing_not_manual", f"local_id 已存在但不是本地手动车辆，不能通过本地 Excel 覆盖：{local_id}"))
                continue
        record = _vehicle_row_to_record(row, pictures_by_id.get(local_id, []), observed_at=observed_at, existing=existing)
        records.append(record)
        detail = record["source_payloads"]["vehicle_detail"]["payload"]
        public_rows.append(
            {
                "row_number": row["row_number"],
                "local_id": local_id,
                "action": "update" if existing else "create",
                "name": _vehicle_display_name(detail, fallback=local_id),
                "picture_count": len(pictures_by_id.get(local_id, [])),
                "source": {"type": "manual", "provider": "manual", "ingest_channel": "manual_input"},
            }
        )
    ok = not errors
    return {
        "ok": ok,
        "template_version": template_version,
        "summary": {
            "vehicle_count": len(vehicle_rows),
            "picture_count": len(picture_rows),
            "error_count": len(errors),
            "mode": "local_manual_vehicle_excel",
        },
        "vehicles": public_rows,
        "errors": errors,
        "records": records if ok else [],
    }


def _parse_vehicle_rows(sheet: Worksheet, columns: tuple[ColumnSpec, ...]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header, header_errors = _header_map(sheet, columns, VEHICLE_SHEET)
    missing = _missing_headers(header, columns, VEHICLE_SHEET)
    errors = [*header_errors, *missing, *_unknown_headers(header, columns, VEHICLE_SHEET)]
    rows: list[dict[str, Any]] = []
    if missing or header_errors:
        return rows, errors
    for row_number, values in _iter_data_rows(sheet, header):
        if all(value in (None, "") for value in values.values()):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            errors.append(_row_error(row_number, "workbook", "too_many_rows", f"单次最多导入 {MAX_IMPORT_ROWS} 台车辆"))
            break
        parsed: dict[str, Any] = {"row_number": row_number, "detail": {}, "annotations": {}, "manual": {}, "field_provenance": {}}
        row_errors: list[dict[str, Any]] = []
        for spec in columns:
            raw = values.get(spec.header)
            value, value_error = _normalize_value(raw, spec)
            if spec.required and value in (None, "", [], {}):
                row_errors.append(_row_error(row_number, spec.header, "required_missing", f"{spec.header} 不能为空"))
                continue
            if value_error:
                row_errors.append(_row_error(row_number, spec.header, "invalid_value", value_error))
                continue
            if value in (None, "", [], {}):
                continue
            if spec.target == "local_id":
                local_id = str(value)
                if not SAFE_PRODUCT_ID_RE.fullmatch(local_id):
                    row_errors.append(_row_error(row_number, spec.header, "unsafe_local_id", f"local_id 只能包含字母、数字、下划线、点和短横线：{local_id}"))
                parsed["local_id"] = local_id
            elif spec.target.startswith("detail."):
                path = spec.target.removeprefix("detail.")
                _set_path(parsed["detail"], path, value)
                parsed["field_provenance"][f"source_payloads.vehicle_detail.payload.{path}"] = _field_provenance(
                    sheet=VEHICLE_SHEET,
                    row=row_number,
                    header=spec.header,
                )
            elif spec.target.startswith("annotations."):
                key = spec.target.removeprefix("annotations.")
                parsed["annotations"][key] = _annotation_value(key, value)
            elif spec.target.startswith("manual."):
                _set_path(parsed["manual"], spec.target.removeprefix("manual."), value)
        if row_errors:
            errors.extend(row_errors)
            continue
        if parsed.get("local_id"):
            rows.append(parsed)
    return rows, errors


def _parse_vehicle_rows_simple_vertical(sheet: Worksheet, *, columns: tuple[ColumnSpec, ...]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header, header_errors = _vertical_header_map(sheet, expected=VERTICAL_VEHICLE_HEADERS)
    missing = [
        _row_error(1, name, "missing_column", f"{VEHICLE_SHEET} 缺少列：{name}")
        for name in VERTICAL_VEHICLE_HEADERS
        if name not in header
    ]
    errors = [*header_errors, *missing, *_vertical_unknown_headers(header, expected=VERTICAL_VEHICLE_HEADERS)]
    rows_by_id: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    seen_labels: set[tuple[str, str]] = set()
    field_specs = _vehicle_field_specs_by_label(columns)
    if missing or header_errors:
        return [], errors
    for row_number, values in _iter_data_rows(sheet, header):
        if all(value in (None, "") for value in values.values()):
            continue
        raw_local_id = str(values.get(VERTICAL_VEHICLE_ID_HEADER) or "").strip()
        field_label = str(values.get("填写项") or "").strip()
        if not raw_local_id:
            errors.append(_row_error(row_number, "车辆编号", "required_missing", "车辆编号不能为空"))
            continue
        if not SAFE_PRODUCT_ID_RE.fullmatch(raw_local_id):
            errors.append(_row_error(row_number, "车辆编号", "unsafe_local_id", f"车辆编号只能包含字母、数字、下划线、点和短横线：{raw_local_id}"))
            continue
        if not field_label:
            errors.append(_row_error(row_number, "填写项", "required_missing", "填写项不能为空"))
            continue
        spec = field_specs.get(field_label)
        if spec is None:
            errors.append(_row_error(row_number, "填写项", "unknown_field_item", f"填写项不在当前模板中：{field_label}"))
            continue
        duplicate_key = (raw_local_id, field_label)
        if duplicate_key in seen_labels:
            errors.append(_row_error(row_number, "填写项", "duplicate_field_item", f"同一车辆编号下填写项重复：{raw_local_id} / {field_label}"))
            continue
        seen_labels.add(duplicate_key)
        value, value_error = _normalize_value(values.get("填写内容"), spec)
        if spec.required and value in (None, "", [], {}):
            errors.append(_row_error(row_number, "填写内容", "required_missing", f"{field_label} 不能为空"))
            continue
        if value_error:
            errors.append(_row_error(row_number, "填写内容", "invalid_value", value_error.replace(spec.header, field_label)))
            continue
        if raw_local_id not in rows_by_id:
            if len(rows_by_id) >= MAX_IMPORT_ROWS:
                errors.append(_row_error(row_number, "workbook", "too_many_rows", f"单次最多导入 {MAX_IMPORT_ROWS} 台车辆"))
                break
            rows_by_id[raw_local_id] = {"row_number": row_number, "local_id": raw_local_id, "detail": {}, "annotations": {}, "manual": {}, "field_provenance": {}}
            order.append(raw_local_id)
        if value in (None, "", [], {}):
            continue
        parsed = rows_by_id[raw_local_id]
        _apply_vertical_value(parsed, spec, value, row_number=row_number, provenance_header=field_label)
    for local_id in order:
        row = rows_by_id[local_id]
        if _get_path(row.get("detail") if isinstance(row.get("detail"), dict) else {}, "baseCarInfo.name.displayValue") in (None, "", [], {}):
            errors.append(_row_error(row["row_number"], "车辆展示名称", "required_missing", f"{local_id} 缺少必填项：车辆展示名称"))
    return [rows_by_id[local_id] for local_id in order], errors


def _parse_vehicle_rows_path_vertical(sheet: Worksheet) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header, header_errors = _vertical_header_map(sheet, expected=PATH_VERTICAL_VEHICLE_HEADERS)
    missing = [
        _row_error(1, name, "missing_column", f"{VEHICLE_SHEET} 缺少列：{name}")
        for name in PATH_VERTICAL_VEHICLE_HEADERS
        if name not in header
    ]
    errors = [*header_errors, *missing, *_vertical_unknown_headers(header, expected=PATH_VERTICAL_VEHICLE_HEADERS)]
    rows_by_id: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    seen_paths: set[tuple[str, str]] = set()
    field_specs = _vehicle_field_specs_by_target()
    if missing or header_errors:
        return [], errors
    for row_number, values in _iter_data_rows(sheet, header):
        if all(value in (None, "") for value in values.values()):
            continue
        raw_local_id = str(values.get("local_id") or "").strip()
        field_path = str(values.get("字段路径") or "").strip()
        if not raw_local_id:
            errors.append(_row_error(row_number, "local_id", "required_missing", "local_id 不能为空"))
            continue
        if not SAFE_PRODUCT_ID_RE.fullmatch(raw_local_id):
            errors.append(_row_error(row_number, "local_id", "unsafe_local_id", f"local_id 只能包含字母、数字、下划线、点和短横线：{raw_local_id}"))
            continue
        if not field_path:
            errors.append(_row_error(row_number, "字段路径", "required_missing", "字段路径不能为空"))
            continue
        spec = field_specs.get(field_path)
        if spec is None:
            errors.append(_row_error(row_number, "字段路径", "unknown_field_path", f"字段路径不在当前模板中：{field_path}"))
            continue
        duplicate_key = (raw_local_id, field_path)
        if duplicate_key in seen_paths:
            errors.append(_row_error(row_number, "字段路径", "duplicate_field_path", f"同一 local_id 下字段路径重复：{raw_local_id} / {field_path}"))
            continue
        seen_paths.add(duplicate_key)
        value, value_error = _normalize_value(values.get("字段值"), spec)
        if spec.required and value in (None, "", [], {}):
            errors.append(_row_error(row_number, "字段值", "required_missing", f"{field_path} 不能为空"))
            continue
        if value_error:
            errors.append(_row_error(row_number, "字段值", "invalid_value", value_error))
            continue
        if raw_local_id not in rows_by_id:
            if len(rows_by_id) >= MAX_IMPORT_ROWS:
                errors.append(_row_error(row_number, "workbook", "too_many_rows", f"单次最多导入 {MAX_IMPORT_ROWS} 台车辆"))
                break
            rows_by_id[raw_local_id] = {"row_number": row_number, "local_id": raw_local_id, "detail": {}, "annotations": {}, "manual": {}, "field_provenance": {}}
            order.append(raw_local_id)
        if value in (None, "", [], {}):
            continue
        parsed = rows_by_id[raw_local_id]
        _apply_vertical_value(parsed, spec, value, row_number=row_number, provenance_header=spec.target)
    for local_id in order:
        row = rows_by_id[local_id]
        if _get_path(row.get("detail") if isinstance(row.get("detail"), dict) else {}, "baseCarInfo.name.displayValue") in (None, "", [], {}):
            errors.append(_row_error(row["row_number"], "detail.baseCarInfo.name.displayValue", "required_missing", f"{local_id} 缺少必填字段：detail.baseCarInfo.name.displayValue"))
    return [rows_by_id[local_id] for local_id in order], errors


def _parse_picture_rows(sheet: Worksheet, *, known_ids: set[str], columns: tuple[ColumnSpec, ...]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header, header_errors = _header_map(sheet, columns, PICTURE_SHEET)
    missing = _missing_headers(header, columns, PICTURE_SHEET)
    errors = [*header_errors, *missing, *_unknown_headers(header, columns, PICTURE_SHEET)]
    rows: list[dict[str, Any]] = []
    if missing or header_errors:
        return rows, errors
    for row_number, values in _iter_data_rows(sheet, header):
        if all(value in (None, "") for value in values.values()):
            continue
        if len(rows) >= MAX_PICTURE_ROWS:
            errors.append(_row_error(row_number, "workbook", "too_many_rows", f"单次最多导入 {MAX_PICTURE_ROWS} 条车辆图片"))
            break
        parsed: dict[str, Any] = {"row_number": row_number}
        row_errors: list[dict[str, Any]] = []
        for spec in columns:
            value, value_error = _normalize_value(values.get(spec.header), spec)
            if spec.required and value in (None, "", [], {}):
                row_errors.append(_row_error(row_number, spec.header, "required_missing", f"{spec.header} 不能为空"))
                continue
            if value_error:
                row_errors.append(_row_error(row_number, spec.header, "invalid_value", value_error))
                continue
            if value not in (None, "", [], {}):
                parsed[spec.target] = value
        local_id = str(parsed.get("local_id") or "")
        if local_id and local_id not in known_ids:
            row_errors.append(_row_error(row_number, "local_id", "unknown_local_id", f"车辆图片表引用了不存在的 local_id：{local_id}"))
        if not parsed.get("pictureUrl") and not parsed.get("bigPictureUrl"):
            picture_url_field = next((spec.header for spec in columns if spec.target == "pictureUrl"), "图片URL pictureUrl")
            row_errors.append(_row_error(row_number, picture_url_field, "picture_url_missing", "图片URL和大图URL至少填写一个"))
        if row_errors:
            errors.extend(row_errors)
            continue
        if local_id:
            rows.append(parsed)
    return rows, errors


def _vehicle_row_to_record(
    row: dict[str, Any],
    picture_rows: list[dict[str, Any]],
    *,
    observed_at: str,
    existing: dict[str, Any] | None = None,
) -> dict[str, Any]:
    pictures = []
    for index, picture in enumerate(sorted(picture_rows, key=lambda item: int(item.get("sort") or 0)), start=1):
        url = str(picture.get("pictureUrl") or picture.get("bigPictureUrl") or "").strip()
        entry = {
            "pictureId": f"excel_{row['local_id']}_{index}",
            "pictureUrl": url,
            "source": "manual_excel_import",
        }
        if picture.get("bigPictureUrl"):
            entry["bigPictureUrl"] = str(picture.get("bigPictureUrl"))
        if picture.get("sort") not in (None, ""):
            entry["sort"] = int(picture.get("sort") or 0)
        if picture.get("description"):
            entry["description"] = str(picture.get("description"))
        pictures.append(entry)
    detail_patch = copy.deepcopy(row.get("detail") if isinstance(row.get("detail"), dict) else {})
    annotations = row.get("annotations") if isinstance(row.get("annotations"), dict) else {}
    manual = row.get("manual") if isinstance(row.get("manual"), dict) else {}
    if isinstance(existing, dict):
        patch: dict[str, Any] = {
            "vehicle_detail_patch": detail_patch,
            "annotations": annotations,
            "manual_annotations": manual,
        }
        if picture_rows:
            patch["vehicle_pictures_patch"] = pictures
        record = apply_admin_vehicle_update(copy.deepcopy(existing), patch, observed_at=observed_at)
        provenance = record.setdefault("extensions", {}).setdefault("manual", {}).setdefault("field_provenance", {})
        for path, entry in (row.get("field_provenance") if isinstance(row.get("field_provenance"), dict) else {}).items():
            provenance[path] = copy.deepcopy(entry)
    else:
        record = create_manual_vehicle(
            record_id=str(row["local_id"]),
            vehicle_detail_payload=detail_patch,
            pictures_payload=pictures,
            observed_at=observed_at,
            field_provenance=copy.deepcopy(row.get("field_provenance") if isinstance(row.get("field_provenance"), dict) else {}),
        )
        if annotations or manual:
            record = apply_admin_vehicle_update(record, {"annotations": annotations, "manual_annotations": manual}, observed_at=observed_at)
    record.setdefault("metadata", {})["import_source"] = "local_manual_vehicle_excel"
    record["metadata"]["excel_row_number"] = row.get("row_number")
    return record


def _write_header(sheet: Worksheet, columns: tuple[ColumnSpec, ...]) -> None:
    sheet.append([spec.header for spec in columns])


def _header_map(sheet: Worksheet, columns: tuple[ColumnSpec, ...], sheet_name: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
    result: dict[str, int] = {}
    errors: list[dict[str, Any]] = []
    for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1), []), start=1):
        header = str(cell.value or "").strip()
        if header:
            if header in result:
                errors.append(_row_error(1, header, "duplicate_column", f"{sheet_name} 存在重复列：{header}"))
                continue
            result[header] = index
    return result, errors


def _vertical_header_map(sheet: Worksheet, *, expected: tuple[str, ...]) -> tuple[dict[str, int], list[dict[str, Any]]]:
    result: dict[str, int] = {}
    errors: list[dict[str, Any]] = []
    for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1), []), start=1):
        raw_header = str(cell.value or "").strip()
        header = _normalize_vertical_header(raw_header)
        if header:
            if header in result:
                errors.append(_row_error(1, raw_header or header, "duplicate_column", f"{VEHICLE_SHEET} 存在重复列：{raw_header or header}"))
                continue
            result[header] = index
    return result, errors


def _normalize_vertical_header(header: str) -> str:
    clean_header = str(header or "").strip()
    return VERTICAL_VEHICLE_HEADER_ALIASES.get(clean_header, clean_header)


def _vertical_unknown_headers(header: dict[str, int], *, expected: tuple[str, ...]) -> list[dict[str, Any]]:
    expected_names = set(expected)
    return [
        _row_error(1, name, "unknown_column", f"{VEHICLE_SHEET} 包含未识别列：{name}")
        for name in header
        if name not in expected_names
    ]


def _missing_headers(header: dict[str, int], columns: tuple[ColumnSpec, ...], sheet_name: str) -> list[dict[str, Any]]:
    errors = []
    for spec in columns:
        if spec.header not in header:
            errors.append(_row_error(1, spec.header, "missing_column", f"{sheet_name} 缺少列：{spec.header}"))
    return errors


def _unknown_headers(header: dict[str, int], columns: tuple[ColumnSpec, ...], sheet_name: str) -> list[dict[str, Any]]:
    expected = {spec.header for spec in columns}
    errors = []
    for name in header:
        if name not in expected:
            errors.append(_row_error(1, name, "unknown_column", f"{sheet_name} 包含未识别列：{name}"))
    return errors


def _iter_data_rows(sheet: Worksheet, header: dict[str, int]) -> Any:
    for row in sheet.iter_rows(min_row=2):
        values = {name: _cell_value(row[index - 1].value) for name, index in header.items()}
        yield int(row[0].row), values


def _normalize_value(value: Any, spec: ColumnSpec) -> tuple[Any, str | None]:
    if value in (None, ""):
        return None, None
    if spec.value_type == "text" or spec.value_type == "id":
        return str(value).strip(), None
    if spec.value_type == "list":
        return _split_list(value), None
    if spec.value_type in {"number", "integer"}:
        try:
            text = str(value).strip()
            if text == "":
                return None, None
            number = float(text)
            if spec.value_type == "integer":
                if not number.is_integer():
                    return None, f"{spec.header} 必须是整数"
                return int(number), None
            return int(number) if number.is_integer() else number, None
        except (TypeError, ValueError):
            return None, f"{spec.header} 必须是数字"
    return value, None


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def _split_list(value: Any) -> list[str]:
    if isinstance(value, (list, tuple)):
        candidates = value
    else:
        candidates = re.split(r"[;；,，\n]+", str(value))
    return [str(item).strip() for item in candidates if str(item or "").strip()]


def _annotation_value(key: str, value: Any) -> Any:
    if key == "additional_details" and value not in (None, "", [], {}):
        return {"excel_import_notes": str(value).strip()}
    return value


def _apply_vertical_value(parsed: dict[str, Any], spec: ColumnSpec, value: Any, *, row_number: int, provenance_header: str) -> None:
    if spec.target.startswith("detail."):
        path = spec.target.removeprefix("detail.")
        _set_path(parsed["detail"], path, value)
        parsed["field_provenance"][f"source_payloads.vehicle_detail.payload.{path}"] = _field_provenance(
            sheet=VEHICLE_SHEET,
            row=row_number,
            header=provenance_header,
        )
    elif spec.target.startswith("annotations."):
        key = spec.target.removeprefix("annotations.")
        parsed["annotations"][key] = _annotation_value(key, value)
    elif spec.target.startswith("manual."):
        _set_path(parsed["manual"], spec.target.removeprefix("manual."), value)


def _vehicle_display_name(detail: dict[str, Any], *, fallback: str) -> str:
    base = detail.get("baseCarInfo") if isinstance(detail.get("baseCarInfo"), dict) else {}
    name = base.get("name")
    if isinstance(name, dict):
        display = str(name.get("displayValue") or "").strip()
        if display:
            return display
        joined = " ".join(str(name.get(key) or "").strip() for key in ("brandName", "seriesName", "modelName") if str(name.get(key) or "").strip())
        if joined:
            return joined
    if isinstance(name, str) and name.strip():
        return name.strip()
    return str(fallback)


def _set_path(root: dict[str, Any], path: str, value: Any) -> None:
    current = root
    parts = path.split(".")
    for part in parts[:-1]:
        child = current.get(part)
        if not isinstance(child, dict):
            child = {}
            current[part] = child
        current = child
    current[parts[-1]] = value


def _get_path(root: dict[str, Any], path: str) -> Any:
    current: Any = root
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _field_provenance(*, sheet: str, row: int, header: str) -> dict[str, Any]:
    return {
        "source": "manual_excel_import",
        "ingest_channel": "manual_input",
        "original_path": f"{sheet}!{header}@row{row}",
        "recorded_at": _now_iso(),
    }


def _template_version(workbook: Any) -> str:
    if FIELD_SHEET not in workbook.sheetnames:
        return ""
    sheet = workbook[FIELD_SHEET]
    if str(sheet["A1"].value or "").strip() in {"template_version", "模板版本"}:
        return str(sheet["B1"].value or "").strip()
    return ""


def _vehicle_columns_for_version(template_version: str) -> tuple[ColumnSpec, ...]:
    if template_version == LEGACY_TEMPLATE_VERSION:
        return LEGACY_VEHICLE_COLUMNS
    return VEHICLE_COLUMNS


def _vehicle_field_specs_by_target(columns: tuple[ColumnSpec, ...] = VEHICLE_COLUMNS) -> dict[str, ColumnSpec]:
    return {spec.target: spec for spec in columns if spec.target != "local_id"}


def _vehicle_field_specs_by_label(columns: tuple[ColumnSpec, ...] = VEHICLE_COLUMNS) -> dict[str, ColumnSpec]:
    return {_simple_field_label(spec): spec for spec in columns if spec.target != "local_id"}


def _simple_field_label(spec: ColumnSpec) -> str:
    if "｜" in str(spec.header or ""):
        return str(spec.header)
    return _SIMPLE_FIELD_LABELS_BY_TARGET.get(spec.target, str(spec.header or spec.target).split(" ", 1)[0])


def _duplicates(values: Any) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for value in values:
        text = str(value or "")
        if not text:
            continue
        if text in seen and text not in duplicates:
            duplicates.append(text)
        seen.add(text)
    return duplicates


def _row_error(row_number: int, field: str, code: str, message: str) -> dict[str, Any]:
    return {"row_number": int(row_number), "field": str(field), "code": str(code), "message": str(message)}


def _record_row_number(record: Any) -> int:
    metadata = record.get("metadata") if isinstance(record, dict) and isinstance(record.get("metadata"), dict) else {}
    try:
        return int(metadata.get("excel_row_number") or 0)
    except (TypeError, ValueError):
        return 0


def _public_preview(artifact: dict[str, Any]) -> dict[str, Any]:
    return {
        "ok": bool(artifact.get("ok")),
        "preview_id": str(artifact.get("preview_id") or ""),
        "template_version": str(artifact.get("template_version") or ""),
        "summary": copy.deepcopy(artifact.get("summary") if isinstance(artifact.get("summary"), dict) else {}),
        "vehicles": copy.deepcopy(artifact.get("vehicles") if isinstance(artifact.get("vehicles"), list) else []),
        "errors": copy.deepcopy(artifact.get("errors") if isinstance(artifact.get("errors"), list) else []),
        "mode": "local_manual_vehicle_excel",
    }


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f"{path.name}.{uuid4().hex}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")
