"""Customer-data extraction and Excel persistence for WeChat customer service."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


HEADERS = [
    "received_at",
    "source_target",
    "message_ids",
    "name",
    "phone",
    "address",
    "product",
    "quantity",
    "spec",
    "budget",
    "note",
    "raw_text",
]

FIELD_LABELS = {
    "name": ["姓名", "客户姓名", "联系人", "收件人", "名字"],
    "phone": ["电话", "手机号", "手机", "联系电话", "联系方式"],
    "address": ["地址", "收货地址", "寄送地址", "详细地址"],
    "product": ["产品", "商品", "需求", "采购", "品名"],
    "quantity": ["数量", "件数", "采购量", "数量需求"],
    "spec": ["规格", "型号", "尺寸", "参数"],
    "budget": ["预算", "价格", "报价", "费用"],
    "note": ["备注", "说明", "其他", "补充"],
}


@dataclass(frozen=True)
class CustomerExtraction:
    fields: dict[str, str]
    missing_required_fields: list[str]
    is_customer_data: bool
    complete: bool


def extract_customer_data(text: str, required_fields: list[str] | None = None) -> CustomerExtraction:
    required = required_fields or ["name", "phone"]
    normalized = normalize_text(text)
    fields: dict[str, str] = {}

    phone = extract_phone(normalized)
    if phone:
        fields["phone"] = phone

    for field, labels in FIELD_LABELS.items():
        if field == "phone" and fields.get("phone"):
            continue
        value = extract_labeled_value(normalized, labels)
        if value:
            fields[field] = value

    if not fields.get("name") and should_infer_unlabeled_name(normalized, fields):
        inferred_name = extract_unlabeled_name(normalized)
        if inferred_name:
            fields["name"] = inferred_name

    is_customer_data = has_customer_data_signal(normalized, fields)
    missing = [field for field in required if not fields.get(field)]
    return CustomerExtraction(
        fields=fields,
        missing_required_fields=missing,
        is_customer_data=is_customer_data,
        complete=is_customer_data and not missing,
    )


def has_customer_data_signal(text: str, fields: dict[str, str]) -> bool:
    if fields.get("phone") or fields.get("name") or fields.get("address"):
        return True
    explicit_keywords = [
        "客户资料",
        "客户信息",
        "收货信息",
        "联系方式",
        "联系人",
        "收件人",
        "收货地址",
        "联系电话",
        "手机号",
    ]
    return any(keyword in text for keyword in explicit_keywords)


def should_infer_unlabeled_name(text: str, fields: dict[str, str]) -> bool:
    if fields.get("phone") or fields.get("address") or fields.get("product") or fields.get("quantity"):
        return True
    explicit_keywords = [
        "客户资料",
        "客户信息",
        "收货信息",
        "联系方式",
        "联系人",
        "收件人",
        "收货地址",
        "联系电话",
        "手机号",
    ]
    if any(keyword in text for keyword in explicit_keywords):
        return True
    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"[\u4e00-\u9fff·]{2,6}", compact):
        return is_plain_chinese_name(compact)
    return False


def append_customer_row(
    workbook_path: Path,
    sheet_name: str,
    source_target: str,
    message_ids: list[str],
    raw_text: str,
    fields: dict[str, str],
) -> dict[str, Any]:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        workbook = Workbook()
        workbook.active.title = sheet_name

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    ensure_headers(sheet)
    row = {
        "received_at": datetime.now().isoformat(timespec="seconds"),
        "source_target": source_target,
        "message_ids": ",".join(message_ids),
        "raw_text": raw_text,
        **fields,
    }
    sheet.append([row.get(header, "") for header in HEADERS])
    workbook.save(workbook_path)
    return {
        "ok": True,
        "workbook_path": str(workbook_path),
        "sheet_name": sheet_name,
        "row_number": sheet.max_row,
        "fields": fields,
    }


def ensure_headers(sheet: Any) -> None:
    current = [sheet.cell(row=1, column=index + 1).value for index in range(len(HEADERS))]
    if current == HEADERS:
        return
    if sheet.max_row == 1 and all(value is None for value in current):
        for index, header in enumerate(HEADERS, start=1):
            sheet.cell(row=1, column=index).value = header
        return
    sheet.insert_rows(1)
    for index, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=index).value = header


def normalize_text(text: str) -> str:
    replacements = {
        "：": ":",
        "；": ";",
        "，": ",",
        "。": ".",
        "、": ",",
        "\r\n": "\n",
        "\r": "\n",
    }
    normalized = text
    for old, new in replacements.items():
        normalized = normalized.replace(old, new)
    return normalized.strip()


def extract_phone(text: str) -> str:
    match = re.search(r"(?<!\d)(1[3-9]\d{9})(?!\d)", text)
    return match.group(1) if match else ""


def extract_labeled_value(text: str, labels: list[str]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    pattern = re.compile(
        rf"(?:^|[\n,;，；\s])(?:{label_pattern})\s*(?:[:=：]|是|为)?\s*"
        rf"(.+?)(?=(?:\n|[,;，；]|\s+)(?:{all_label_pattern()})\s*(?:[:=：]|是|为)?|$)",
        re.IGNORECASE,
    )
    match = pattern.search(text)
    if not match:
        return ""
    value = match.group(1).strip(" ,;.，；。:：")
    return cleanup_field_value(value)[:200]


def cleanup_field_value(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip(" ,;.，；。:：")
    value = re.sub(r"(?:姓名|客户姓名|联系人|收件人|名字|电话|手机号|手机|联系电话|联系方式)\s*$", "", value)
    return value.strip(" ,;.，；。:：")


def extract_unlabeled_name(text: str) -> str:
    explicit_patterns = [
        r"(?:我叫|我是|本人|名字叫|联系人是|联系人|收件人是|收件人)\s*([\u4e00-\u9fff·]{2,8})",
        r"(?:name)\s*[:=：]?\s*([A-Za-z][A-Za-z .'-]{1,40})",
    ]
    for pattern in explicit_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return normalize_name(match.group(1))

    for raw_line in reversed(re.split(r"[\n,;，；]", text)):
        line = raw_line.strip()
        line = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", "", line).strip()
        if not line:
            continue
        if is_plain_chinese_name(line):
            return normalize_name(line)
    return ""


def normalize_name(value: str) -> str:
    value = value.strip(" ,;.，；。:：")
    value = re.sub(r"(先生|女士|小姐|老板|总)$", "", value)
    return value.strip(" ,;.，；。:：")


def is_plain_chinese_name(value: str) -> bool:
    if not re.fullmatch(r"[\u4e00-\u9fff·]{2,6}", value):
        return False
    blocked_words = {
        "产品",
        "商品",
        "冰箱",
        "商用冰箱",
        "滤芯",
        "地址",
        "电话",
        "报价",
        "价格",
        "数量",
        "规格",
        "预算",
        "客户资料",
        "客户信息",
        "收货信息",
        "联系方式",
        "净水器滤芯",
        "测试路",
        "你好",
        "您好",
        "在吗",
        "有人吗",
        "请问",
        "谢谢",
        "辛苦",
        "老板",
        "客服",
        "复杂",
        "推荐",
        "想找",
        "找个",
        "能放",
        "饮料",
        "冷柜",
        "冰柜",
        "随便",
        "看看",
        "挺快",
        "先看",
        }
    if value in blocked_words:
        return False
    blocked_fragments = {
        "地方",
        "接近",
        "小店",
        "饮料",
        "冷柜",
        "办公室",
        "仓库",
        "快递",
        "舒服",
        "腰疼",
        "随便",
        "看看",
        "预算",
        "没定",
        "太贵",
        "先看",
        "东西",
        "靠谱",
        "可靠",
        "忙晕",
        "质量",
        "你好",
        "您好",
        "在吗",
        "有人",
        "请问",
        "谢谢",
        "辛苦",
        "客服",
        "老板",
    }
    if any(fragment in value for fragment in blocked_fragments):
        return False
    return True


def all_label_pattern() -> str:
    labels = []
    for items in FIELD_LABELS.values():
        labels.extend(items)
    return "|".join(re.escape(label) for label in sorted(set(labels), key=len, reverse=True))
