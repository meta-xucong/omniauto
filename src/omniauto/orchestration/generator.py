"""Deterministic script generator."""

from __future__ import annotations

from pathlib import Path
from typing import Any, List, Tuple

from ..high_level.task_planner import TaskPlanner


class ScriptGenerator:
    """Generate deterministic OmniAuto workflow scripts."""

    def __init__(self, model: Any = None) -> None:
        self.planner = TaskPlanner(model)
        self.model = model

    def generate(self, task_description: str, output_path: str) -> str:
        steps = self.planner.plan(task_description)
        if any(step.get("type") == "build_excel_report" for step in steps):
            return self._write_local_excel_report_template(task_description, steps, output_path)
        return self._write_atomic_template(task_description, steps, output_path)

    def _write_atomic_template(self, task_description: str, steps: List[dict], output_path: str) -> str:
        imports, statements = self._build_atomic_steps(steps)
        lines = [
            "# Auto-generated OmniAuto atomic script",
            f"# Task: {task_description}",
            "",
            "from omniauto.core.state_machine import Workflow",
        ]
        lines.extend(sorted(imports))
        lines.extend(
            [
                "",
                "requires_browser = True",
                "",
                'workflow = Workflow(task_id="auto_task")',
            ]
        )
        lines.extend(statements)
        lines.append("")

        code = "\n".join(lines)
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_text(code, encoding="utf-8")
        return output_path

    def _write_local_excel_report_template(self, task_description: str, steps: List[dict], output_path: str) -> str:
        input_path = next((step.get("path", "") for step in steps if step.get("type") == "load_json_report"), "")
        report_step = next((step for step in steps if step.get("type") == "build_excel_report"), {})
        workbook_path = report_step.get("output_path", "")

        code = f'''# Auto-generated OmniAuto local report workflow
# Task: {task_description}

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from omniauto.core.context import TaskContext
from omniauto.core.state_machine import AtomicStep, Workflow

requires_browser = False

INPUT_PATH = Path(r"{input_path}")
OUTPUT_PATH = Path(r"{workbook_path}")


def _parse_param_map(params: list[str]) -> dict[str, str]:
    result: dict[str, str] = {{}}
    for raw in params or []:
        text = (raw or "").replace("\\r", "\\n")
        parts = [part.strip() for part in re.split(r"[\\t\\n]+", text) if part.strip()]
        if len(parts) >= 2:
            result[parts[0]] = " / ".join(parts[1:])
    return result


def _first_non_empty_line(text: str) -> str:
    for line in (text or "").splitlines():
        line = line.strip()
        if line:
            return line
    return ""


def _build_rows(data: dict) -> list[dict]:
    source_items = list(data.get("items") or data.get("top_cheapest_items") or data.get("all_items") or [])
    ordered = sorted(
        source_items,
        key=lambda item: (
            item.get("price_num") is None,
            item.get("price_num") if item.get("price_num") is not None else 10**9,
            item.get("title") or "",
        ),
    )

    rows: list[dict] = []
    seen_links: set[str] = set()
    for item in ordered:
        link = item.get("link") or ""
        if link and link in seen_links:
            continue
        if link:
            seen_links.add(link)
        detail = item.get("detail") or {{}}
        param_map = _parse_param_map(detail.get("params") or [])
        params_summary = "；".join(f"{{k}}: {{v}}" for k, v in param_map.items())
        shop_subject = item.get("shop_name") or _first_non_empty_line(detail.get("location") or "")
        rows.append({{
            "价格(元)": item.get("price_num"),
            "价格文本": (item.get("price_text") or "").replace("\\n", ""),
            "商品标题": item.get("title") or "",
            "店铺主体": shop_subject,
            "来源页": item.get("source_page"),
            "品牌": param_map.get("品牌", ""),
            "材质": param_map.get("材质", ""),
            "风格": param_map.get("风格", ""),
            "颜色": param_map.get("颜色", ""),
            "尺寸": param_map.get("尺寸", ""),
            "参数摘要": params_summary,
            "1688链接": link,
        }})

    for index, row in enumerate(rows, 1):
        row["排名"] = index
        row["TOP10"] = "是" if index <= 10 else ""
    return rows


def _write_workbook(data: dict, rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PriceSorted"
    summary = wb.create_sheet("Summary")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    top10_fill = PatternFill("solid", fgColor="FFF2CC")
    link_font = Font(color="0563C1", underline="single")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    summary_rows = [
        ("关键词", data.get("keyword")),
        ("站点", data.get("site")),
        ("抓取页数", data.get("list_pages_completed")),
        ("排序后商品数", len(rows)),
        ("前10最低价数量", min(10, len(rows))),
        ("详情完成数", data.get("detail_sample_completed")),
        ("停止原因", data.get("stopped_reason") or "无"),
        ("生成时间", data.get("generated_at") or datetime.now().isoformat()),
    ]
    for row_index, (key, value) in enumerate(summary_rows, 1):
        summary.cell(row_index, 1, key)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).fill = header_fill
        summary.cell(row_index, 1).font = header_font
        summary.cell(row_index, 1).alignment = center_alignment
        summary.cell(row_index, 2).alignment = body_alignment
        summary.cell(row_index, 1).border = border
        summary.cell(row_index, 2).border = border
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 52
    summary.freeze_panes = "A2"

    headers = ["排名", "TOP10", "价格(元)", "价格文本", "商品标题", "店铺主体", "来源页", "品牌", "材质", "风格", "颜色", "尺寸", "参数摘要", "1688链接"]
    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(1, col_index, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    for row_index, row in enumerate(rows, 2):
        is_top10 = row.get("TOP10") == "是"
        for col_index, header in enumerate(headers, 1):
            value = row.get(header, "")
            cell = ws.cell(row_index, col_index, value)
            cell.border = border
            if header in {{"排名", "TOP10", "价格(元)", "价格文本", "来源页"}}:
                cell.alignment = center_alignment
            else:
                cell.alignment = body_alignment
            if is_top10:
                cell.fill = top10_fill
            if header == "1688链接" and value:
                cell.hyperlink = value
                cell.font = link_font
                cell.value = "打开1688商品页"
            if header == "价格(元)" and isinstance(value, (int, float)):
                cell.number_format = "0.00"

    ws.auto_filter.ref = f"A1:{{get_column_letter(len(headers))}}{{max(2, len(rows) + 1)}}"
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90

    widths = {{
        "A": 8, "B": 8, "C": 12, "D": 12, "E": 42, "F": 24, "G": 10,
        "H": 14, "I": 12, "J": 12, "K": 22, "L": 22, "M": 52, "N": 20,
    }}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 24
    for index in range(2, len(rows) + 2):
        ws.row_dimensions[index].height = 36

    wb.save(output_path)


async def build_excel_report(context: TaskContext) -> dict:
    if not INPUT_PATH.exists():
        return {{"success": False, "error": f"input_not_found: {{INPUT_PATH}}"}}

    data = json.loads(INPUT_PATH.read_text(encoding="utf-8"))
    rows = _build_rows(data)
    if not rows:
        return {{"success": False, "error": "no_rows_built"}}

    _write_workbook(data, rows, OUTPUT_PATH)
    return {{
        "success": True,
        "output_path": str(OUTPUT_PATH),
        "row_count": len(rows),
    }}


workflow = Workflow(task_id="auto_task")
workflow.add_step(
    AtomicStep(
        "build_excel_report",
        build_excel_report,
        lambda result: bool(result.get("success")) and bool(result.get("output_path")),
        retry=0,
        description="build local excel report",
    )
)
'''

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_text(code, encoding="utf-8")
        return output_path

    def _build_atomic_steps(self, steps: List[dict]) -> Tuple[set[str], List[str]]:
        imports: set[str] = set()
        statements: List[str] = []
        for step in steps:
            import_line, statement = self._render_step(step)
            imports.add(import_line)
            statements.append(statement)
        return imports, statements

    def _render_step(self, step: dict) -> Tuple[str, str]:
        stype = step.get("type")

        if stype == "navigate":
            url = step.get("url", "https://example.com")
            return (
                "from omniauto.steps.navigate import NavigateStep",
                f"workflow.add_step(NavigateStep({url!r}))",
            )

        if stype == "click":
            selector = step.get("selector", "button")
            return (
                "from omniauto.steps.click import ClickStep",
                f"workflow.add_step(ClickStep({selector!r}))",
            )

        if stype == "type":
            selector = step.get("selector", "input")
            text = step.get("text", "")
            return (
                "from omniauto.steps.type import TypeStep",
                f"workflow.add_step(TypeStep({selector!r}, {text!r}, interval=(0.05, 0.15)))",
            )

        if stype == "extract_text":
            selector = step.get("selector", "body")
            return (
                "from omniauto.steps.extract import ExtractTextStep",
                f"workflow.add_step(ExtractTextStep({selector!r}))",
            )

        if stype == "hotkey":
            keys = step.get("keys", [])
            key_args = ", ".join(repr(key) for key in keys)
            return (
                "from omniauto.steps.hotkey import HotkeyStep",
                f"workflow.add_step(HotkeyStep({key_args}))",
            )

        if stype == "screenshot":
            output_dir = step.get("output_dir", "./screenshots")
            return (
                "from omniauto.steps.screenshot import ScreenshotStep",
                f"workflow.add_step(ScreenshotStep({output_dir!r}))",
            )

        raise ValueError(f"unsupported step type: {stype!r}")
